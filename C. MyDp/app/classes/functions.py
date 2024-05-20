import pandas as pd
import openpyxl
import pyreadstat
import io
import numpy as np


def convertSav(file):

    strFileName = file.filename

    xlsx = io.BytesIO(file.file.read())

    # wb = openpyxl.load_workbook(f'{strFileName}.xlsx')
    wb = openpyxl.load_workbook(xlsx)

    wsData = wb['Data']
    wsQres = wb['Question']

    mergedCells = list()
    for group in wsData.merged_cells.ranges:
        mergedCells.append(group)

    for group in mergedCells:
        wsData.unmerge_cells(str(group))

    # print("Maximum rows before removing:", wsData.max_row)

    wsData.delete_rows(1, 4)
    wsData.delete_rows(2, 1)

    for icol in range(1, wsData.max_column + 1):
        if wsData.cell(row=1, column=icol).value is None:
            wsData.cell(row=1, column=icol).value = f'{wsData.cell(row=1, column=icol-1).value}_{wsData.cell(row=2, column=icol).value}'

    wsData.delete_rows(2, 1)

    # print("Maximum rows after removing:", wsData.max_row)

    # wb.save(f'{strFileName}_new.xlsx')

    data = wsData.values
    columns = next(data)[0:]
    dfData = pd.DataFrame(data, columns=columns)

    lstDrop = [
        'Approve',
        'Reject',
        'Re - do request', 'Re-do request',
        'Reason to reject',
        'Memo',
        'No.',
        'Date',
        'Country',
        'Channel',
        'Chain / Type',
        'Distributor',
        'Method',
        'Panel FB',
        'Panel Email',
        'Panel Phone',
        'Panel Age',
        'Panel Gender',
        'Panel Area',
        'Panel Income',
        'Login ID',
        'User name',
        'Store ID',
        'Store Code',
        'Store name',
        'Store level',
        'District',
        'Ward',
        'Store address',
        'Area group',
        'Store ranking',
        'Region 2',
        'Nhóm cửa hàng',
        'Nhà phân phối',
        'Manager',
        'Telephone number',
        'Contact person',
        'Email',
        'Others 1',
        'Others 2',
        'Others 3',
        'Others 4',
        'Check in',
        'Store Latitude',
        'Store Longitude',
        'User Latitude',
        'User Longitude',
        'Check out',
        'Distance',
        'Task duration',

        'Panel ID',
        'InterviewerID',
        'InterviewerName',
        'RespondentName',
    ]

    for col in dfData.columns:
        if col in lstDrop or '_Images' in col:
            dfData.drop(col, inplace=True, axis=1)

    # dfData.rename(columns={'Panel ID': 'PanelID'}, inplace=True)

    data = wsQres.values
    columns = next(data)[0:]
    dfQres = pd.DataFrame(data, columns=columns)

    # dfQres.to_excel(f'{strFileName}_Qres.xlsx')

    dictQres = dict()
    for idx in dfQres.index:

        strMatrix = ('' if dfQres.loc[idx, 'Question(Matrix)'] is None else f"{dfQres.loc[idx, 'Question(Matrix)']}")
        strNormal = dfQres.loc[idx, 'Question(Normal)']
        strQreName = str(dfQres.loc[idx, 'Name of items'])
        strQreName = strQreName.replace('Rank_', 'Rank') if 'Rank_' in strQreName else strQreName

        dictQres[strQreName] = {
            'type': dfQres.loc[idx, 'Question type'],
            'label': f'{strMatrix}_{strNormal}' if strMatrix != '' else f'{strNormal}',
            'isMAMatrix': True if strMatrix != '' and dfQres.loc[idx, 'Question type'] == 'MA' else False,
            'cats': {}
        }

        for col in dfQres.columns:
            if col not in ['Name of items', 'Question type', 'Question(Matrix)', 'Question(Normal)'] \
                    and dfQres.loc[idx, col] is not None:
                dictQres[strQreName]['cats'].update({int(col): str(dfQres.loc[idx, col])})


    lstMatrixHeader = list()
    for k in dictQres.keys():
        if dictQres[k]['isMAMatrix'] and len(dictQres[k]['cats'].keys()):
            lstMatrixHeader.append(k)

    for i in lstMatrixHeader:
        for code in dictQres[i]['cats'].keys():

            strQreLbl = dictQres[i]['label']
            strAttLbl = str(dictQres[f'{i}_{code}']['label'])
            strAttLbl = strAttLbl.replace(strQreLbl.rsplit('_', 1)[0], '')[1:]

            dictQres[f'{i}_{code}']['label'] = f"{strQreLbl}_{strAttLbl}"

            dictQres[f'{i}_{code}']['cats'].update({1: strAttLbl})


    column_labels = list()
    variable_value_labels = dict()
    for col in dfData.columns:
        if col in dictQres.keys():
            column_labels.append(dictQres[col]['label'])
            variable_value_labels[col] = dictQres[col]['cats']

        else:
            column_labels.append(col)


    dfData.replace({None: np.nan}, inplace=True)

    savFile = pyreadstat.write_sav(dfData, f"{strFileName.replace('xlsx', '')}sav",
                                   column_labels=column_labels, variable_value_labels=variable_value_labels)

    wb.close()

    print('sav saved.')
    return f"{strFileName.replace('xlsx', '')}sav", savFile




