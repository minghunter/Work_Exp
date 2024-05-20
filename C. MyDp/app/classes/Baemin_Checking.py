from builtins import Exception

import pandas as pd
import openpyxl
import io
import numpy as np
import openpyxl.styles
from datetime import datetime
import re

TEMP_VAL = 999999999


class BaeminCheck:

    def __init__(self):

        self.strFileName = None
        self.xlsx = None


    def load(self, file):

        self.strFileName = str(file.filename).replace('.xlsx', '_Output.xlsx')
        self.xlsx = io.BytesIO(file.file.read())


    def check(self):

        wb = openpyxl.load_workbook(self.xlsx, data_only=True)

        wsData = wb['Data']
        wsCheck = wb.copy_worksheet(wsData)

        wsCheck.delete_rows(1, 1)

        data = wsCheck.values
        columns = next(data)[0:]
        dfCheck = pd.DataFrame(data, columns=columns)

        wb.remove(wsCheck)

        dfCheck = dfCheck.dropna(how='all')

        dfCheck.replace({np.nan: None}, inplace=True)
        dfCheck.replace({None: 'NULL'}, inplace=True)

        dfCheck['App'] = [np.nan if str(a).upper() not in ['GRAB', 'SHOPEEFOOD', 'GOJEK'] else a for a in dfCheck['App']]

        dfCheck['Date'] = [np.nan if not self.validateDateFormat(a) else a for a in dfCheck['Date']]

        validTime = re.compile(r'^(([0-1]\d)|(2[0-3]))h[0-5]\d$')
        dfCheck['Time'] = [np.nan if not validTime.match(a) else a for a in dfCheck['Time']]

        dfCheck['Area'] = [np.nan if not self.validateDistrict(a) else a for a in dfCheck['Area']]

        dfCheck['AFV range'] = [np.nan if not self.validate_AFV_range(a, b) else a for a, b in zip(dfCheck['AFV range'], dfCheck['AFV'])]

        dfCheck['AOV'] = [np.nan if not self.validateAOV(a, [b, c, d]) else a for a, b, c, d in
                          zip(dfCheck['AOV'], dfCheck['AFV'], dfCheck['Delivery fee'], dfCheck['Applicable fee'])]


        # AFV
        dfCheck = self.validateAFV(dfCheck)

        # Food discount
        dfCheck = self.validate_Food_Discount(dfCheck)

        dfCheck['% Discount/AFV'] = [np.nan if not self.validate_pct_Discount_AFV(a, b, c) else a for a, b, c in
                                     zip(dfCheck['% Discount/AFV'], dfCheck['Food discount'], dfCheck['AFV'])]

        dfCheck['Giảm giá món ăn (coupon)'] = [np.nan if not self.validateCoupon(a) else a for a in dfCheck['Giảm giá món ăn (coupon)']]

        dfCheck['Delivery fee'] = [np.nan if not self.validate_Fee(a) else a for a in dfCheck['Delivery fee']]

        dfCheck['Applicable fee'] = [np.nan if not self.validate_Fee(a) else a for a in dfCheck['Applicable fee']]

        dfCheck['Delivery discount'] = [np.nan if not self.validate_discount(a, b, c) else a for a, b, c in
                                        zip(dfCheck['Delivery discount'], dfCheck['Delivery fee'], dfCheck['Applicable fee'])]

        dfCheck['Shopee Xu Discount'] = [np.nan if not self.validate_XuDisc(a) else a for a in dfCheck['Shopee Xu Discount']]

        # Tổng giảm Đơn hàng
        dfCheck = self.validate_Total_Discount(dfCheck)

        # AOV after discount
        dfCheck = self.validate_AOV_After_Discount(dfCheck)

        dfCheck['Merchant'] = [np.nan if not self.validate_Merchant(a) else a for a in dfCheck['Merchant']]

        # No. items
        dfCheck = self.validate_No_Items(dfCheck)

        # Quantity & AFV
        dfCheck = self.validate_Quantity_AFV(dfCheck)


        dfCheck = dfCheck[dfCheck.isna().any(axis=1)][:]

        my_red = openpyxl.styles.colors.Color(rgb='FF0000')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)

        for idx in dfCheck.index:
            for i, v in enumerate(list(dfCheck.columns)):

                if v is None:
                    continue

                if pd.isnull(dfCheck.at[idx, v]):
                    irow = idx + 3
                    icol = i + 1

                    wsData.cell(irow, 3).fill = my_fill
                    wsData.cell(1, icol).fill = my_fill
                    wsData.cell(irow, icol).fill = my_fill





        wb.save(self.strFileName)
        wb.close()


    @staticmethod
    def validateDateFormat(strDate):
        strFormat = '%d/%m/%Y'

        try:
            return bool(datetime.strptime(strDate, strFormat))
        except ValueError:
            return False
        except Exception:
            return False


    @staticmethod
    def validateDistrict(strDist):
        try:

            lstDist = [
                'Quận 1',
                'Quận 2',
                'Quận 3',
                'Quận 4',
                'Quận 5',
                'Quận 6',
                'Quận 7',
                'Quận 8',
                'Quận 9',
                'Quận 10',
                'Quận 11',
                'Quận 12',
                'Quận Tân Bình',
                'Quận Phú Nhuận',
                'Quận Gò Vấp',
                'Quận Bình Thạnh',
                'Quận Bình Tân',
                'Quận Tân Phú',
                'Thành phố Thủ Đức',
                'Huyện Hóc Môn',
                'Huyện Củ Chi',
                'Huyện Nhà Bè',
                'Huyện Bình Chánh',
                'Huyện Cần Giờ',
            ]

            if str(strDist).strip() in lstDist:
                return True
            else:
                return False

        except Exception:
            return False


    @staticmethod
    def validate_AFV_range(strAFVRng, intAFV):
        try:
            if not isinstance(intAFV, (float, int)):
                return False

            if intAFV > 160000:
                strCheck = '160+'
            elif intAFV > 120000:
                strCheck = '120-160'
            elif intAFV > 80000:
                strCheck = '80-120'
            elif intAFV > 60000:
                strCheck = '60-80'
            elif intAFV > 40000:
                strCheck = '40-60'
            elif intAFV <= 40000:
                strCheck = 'Under 40'
            else:
                strCheck = ''

            if strAFVRng != strCheck:
                return False
            else:
                return True

        except Exception:
            return False


    @staticmethod
    def validateAOV(a, lst):

        lst = list(map(lambda x: float(str(x).replace('NULL', '0')), lst))

        try:
            if a != sum(lst):
                return False
            else:
                return True

        except TypeError:
            return False
        except Exception:
            return False


    @staticmethod
    def validateAFV(dfCheck):

        df = dfCheck.copy()

        try:
            df.replace({'NULL': 0}, inplace=True)

            df['AFV'] = [np.nan if not float(a).is_integer() else a for a in df['AFV']]

            df['AFV_Sum'] = [0] * df.shape[0]

            for idx in df.index:

                for i in range(1, 21):

                    if isinstance(df.loc[idx, f'AFV {i}'], (int, float)):
                        df.loc[idx, 'AFV_Sum'] += df.loc[idx, f'AFV {i}']
                    else:
                        df.loc[idx, f'AFV {i}'] = np.nan

            df['AFV'] = [np.nan if a != b else a for a, b in zip(df['AFV'], df['AFV_Sum'])]

            df.loc[df['AFV'] < 1000, ['AFV']] = np.nan

            df['AFV'] = [np.nan if 0 > b - a > 60000 else a for a, b in zip(df['AFV'], df['AOV'])]

            dfCheck['AFV'] = df['AFV']

        except Exception:
            dfCheck['AFV'] = [np.nan] * dfCheck.shape[0]


        return dfCheck


    @staticmethod
    def validate_Food_Discount(dfCheck):

        df = dfCheck.copy()

        try:

            df.loc[df['Food discount'] == 'NULL', ['Food discount']] = TEMP_VAL

            df['Food discount'] = [np.nan if not float(a).is_integer() else a for a in df['Food discount']]

            df.loc[df['Food discount'] < 1000, ['Food discount']] = np.nan

            df['Food discount'] = [np.nan if TEMP_VAL != a > b else a for a, b in zip(df['Food discount'], df['AFV'])]

            df.loc[df['Food discount'] == TEMP_VAL, ['Food discount']] = 'NULL'

            dfCheck['Food discount'] = df['Food discount']

        except Exception:
            dfCheck['Food discount'] = [np.nan] * dfCheck.shape[0]

        return dfCheck


    @staticmethod
    def validate_pct_Discount_AFV(a, b, c):
        try:
            if b == 'NULL':
                b = 0

            if a == b/c:
                return True
            else:
                return False

        except Exception:
            return False


    @staticmethod
    def validateCoupon(val):
        try:
            if val != 'NULL':

                if not isinstance(val, (int, float)):
                    return False

                if val < 500:
                    return False

                if not float(val).is_integer():
                    return False

            return True

        except Exception:
            return False

    @staticmethod
    def validate_Fee(val):
        try:
            if val != 'NULL':

                if not isinstance(val, (int, float)):
                    return False

                if val < 1000:
                    return False

                if val > 60000:
                    return False

                if not float(val).is_integer():
                    return False

            return True

        except Exception:
            return False


    @staticmethod
    def validate_discount(a, b, c):
        try:
            if a != 'NULL':

                if not isinstance(a, (int, float)):
                    return False

                if a < 500:
                    return False

                if not float(a).is_integer():
                    return False

                if pd.isnull(a) or pd.isnull(b) or pd.isnull(c):
                    return False

                if a > float(str(b).replace('NULL', '0')) + float(str(c).replace('NULL', '0')):
                    return False

            return True

        except Exception:
            return False


    @staticmethod
    def validate_XuDisc(val):
        try:
            if val != 'NULL':

                if not isinstance(val, (int, float)):
                    return False

                if val < 100:
                    return False

                if not float(val).is_integer():
                    return False

            return True

        except Exception:
            return False


    @staticmethod
    def validate_Total_Discount(dfCheck):

        df = dfCheck.copy()

        try:
            df.loc[df['Tổng giảm Đơn hàng'] == 'NULL', ['Tổng giảm Đơn hàng']] = TEMP_VAL

            df.loc[df['Food discount'] == 'NULL', ['Food discount']] = 0
            df.loc[df['Giảm giá món ăn (coupon)'] == 'NULL', ['Giảm giá món ăn (coupon)']] = 0
            df.loc[df['Delivery discount'] == 'NULL', ['Delivery discount']] = 0
            df.loc[df['Shopee Xu Discount'] == 'NULL', ['Shopee Xu Discount']] = 0

            df['Total_Discount_Check'] = df['Food discount'] + df['Giảm giá món ăn (coupon)'] + df['Delivery discount'] + df['Shopee Xu Discount']

            df.loc[(df['Total_Discount_Check'] != df['Tổng giảm Đơn hàng']) & (df['Tổng giảm Đơn hàng'] != TEMP_VAL), ['Tổng giảm Đơn hàng']] = np.nan

            df.loc[df['Tổng giảm Đơn hàng'] == TEMP_VAL, ['Tổng giảm Đơn hàng']] = 'NULL'

            dfCheck['Tổng giảm Đơn hàng'] = df['Tổng giảm Đơn hàng']

        except Exception:
            dfCheck['Tổng giảm Đơn hàng'] = [np.nan] * dfCheck.shape[0]

        return dfCheck


    @staticmethod
    def validate_AOV_After_Discount(dfCheck):

        df = dfCheck.copy()

        try:

            df.loc[df['AOV after discount'] == 'NULL', ['AOV after discount']] = TEMP_VAL

            df.loc[df['AOV'] == 'NULL', ['AOV']] = 0
            df.loc[df['Food discount'] == 'NULL', ['Food discount']] = 0
            df.loc[df['Giảm giá món ăn (coupon)'] == 'NULL', ['Giảm giá món ăn (coupon)']] = 0
            df.loc[df['Delivery discount'] == 'NULL', ['Delivery discount']] = 0
            df.loc[df['Shopee Xu Discount'] == 'NULL', ['Shopee Xu Discount']] = 0

            df['AOV_after_discount'] = df['AOV'] - df['Food discount'] - df['Giảm giá món ăn (coupon)'] - df['Delivery discount'] - df['Shopee Xu Discount']

            df.loc[(df['AOV after discount'] != df['AOV_after_discount']) & (df['AOV after discount'] != TEMP_VAL), ['AOV after discount']] = np.nan
            df.loc[df['AOV after discount'] < 0, ['AOV after discount']] = np.nan

            df.loc[df['AOV after discount'] == TEMP_VAL, ['AOV after discount']] = 'NULL'

            dfCheck['AOV after discount'] = df['AOV after discount']

        except Exception:
            dfCheck['AOV after discount'] = [np.nan] * dfCheck.shape[0]

        return dfCheck


    @staticmethod
    def validate_Merchant(val):

        try:

            lst = [
                '-18⁰C',
                '3 Râu - Gà Rán, Pizza & Trà sữa',
                'A Lử Bún Đậu - Bún chả Hà Nội',
                'A Ngáo - Trà sữa & Ăn vặt',
                'A Phón - Cơm chiên & Hủ tiếu xào',
                'A Tài - Bún thịt nướng',
                'A Tùng - Bánh mì bò nướng bơ Cambodia',
                'Ahacook',
                'Ẩm thực Phan Rang',
                'Ăn vặt Sài Gòn 1992',
                'Ande - Cơm da gà mật ong',
                'Anh Ba Gạo',
                'AZ Tea',
                'Bà Bắc - Bánh tráng cuốn trộn',
                'Bà Bảy - Bún bò',
                'Balzar De Café',
                'Bánh bạch tuộc Takoyaki Sami',
                'Bánh bao Thọ Phát',
                'Bánh canh cua Lộc Vừng',
                'Bánh cuốn Ba Miền',
                'Bánh cuốn gia truyền Hà Nội',
                'Bánh Deli',
                'Bánh mì 1 phút 30 giây',
                'Bánh mì A Vĩ',
                'Bánh mì An An',
                'Bánh mì BM',
                'Bánh mì Chim Chạy',
                'Bánh mì Hà Nội',
                'Bánh mì Huỳnh Hoa',
                'Bánh mì Mũ Rơm',
                'Bánh Mì Ơi',
                'Bánh mì PewPew',
                'Bánh mì que Pháp',
                'Bánh mì Thổ Nhĩ Kỳ An Na',
                'Bánh mì Tuấn Mập',
                'Bánh tráng chấm Cô Gánh',
                'Bánh tráng Deli & Xiên que',
                'Bánh Tráng Nướng Đà Lạt - Đường Số 8',
                'Bánh tráng trộn thịt bằm Quang Huy',
                'Bathucha - Bánh chuối Thái',
                'Beno - Mỳ ý sốt bò Mỹ',
                'Bento Delichi',
                'Bố già quán',
                'Bò Kho & Cà Ri Gà Hương Nga',
                'Bonchon Chicken',
                'Bột - Healthy & Weight Loss Food',
                'Bready - Bánh mì tươi Burger đĩa bay',
                'Bún bò 229',
                'Bún bò 31A',
                'Bún bò Ánh Thương',
                'Bún bò Cậu Hai',
                'Bún bò Đông Ba Gia Hội',
                'Bún bò Hoa Lâm',
                'Bún bò huế 14B',
                'Bún bò huế Cô Ba',
                'Bún bò huế Đông Ba',
                'Bún bò Thiên Lý',
                'Bún cay Thái 2 Thuận',
                'Bún chả Hà Nội 1982',
                'Bún chả Hồ Gươm',
                'Bún chả sứa Nha Trang ',
                'Bún chả sứa Nha Trang Mến',
                'Bún đậu Homemade',
                'Bún đậu mắm tôm A Chảnh',
                'Bún đậu mắm tôm Mẹt',
                'Bún Riêu Bếp Trưởng - Bánh Khọt Vũng Tàu',
                'Bún riêu bún bò Cô Lan',
                'Bún riêu canh bún 30',
                'Bún riêu cua Cô Tuyền',
                'Bún riêu cua ốc 66',
                'Bún riêu gánh Bến Thành',
                'Bùn riêu Nguyễn Cảnh Chân',
                'Bún riêu O Nhi',
                'Bún thái hải sản Vân Trường',
                'Bún thái ngon Muối Ớt Xanh',
                'Bún thái và bún mắm Dung',
                'Bún thịt nướng Cô Tín',
                'Bún thịt nướng Hồng Ân',
                'Bún thịt nướng Kiều Bảo',
                'Bún xương bò O Thanh',
                'Burger King',
                'Busan Korean Food',
                'Cà ri gà 1357',
                'Café Amazon',
                'Café IYO - cà phê kem muối',
                'Canopee',
                'Caztus Ice Blended',
                'Chang Hi',
                'Cháo dinh dưỡng Việt Soup',
                'Cháo Ếch Singapore Việt Sing',
                'Cháo lòng - Bún mọc - Bún vịt Phương Nghi',
                'Cháo sườn Bé Hiền',
                'Cháo sườn Cô Giang',
                'Cháo Thuần Việt - Cháo Dinh Dưỡng',
                'Chè Bưởi Đồng Tháp',
                'Chè bưởi Vĩnh Long',
                'Chè khúc bạch Thanh',
                'Chè Ngon 3N',
                'Chè tàu hủ Bà Bồng TS',
                'Cheese Coffee',
                'Chu Dimsum House',
                'Chuk Chuk - Trà Và Cà phê',
                'ChungChun Korean Hotdog',
                'Chuti Korean Food',
                'Cô Hai Quán - Bún Cá Nha Trang & Nem Nướng Nha Trang',
                'Cỏ Mây- Cơm Văn Phòng & Bún Chả Cá',
                'Cô Phương - Cơm Chiên & Nui xào',
                'Coco summer - Trái cây nhanh',
                'Cocoboba - Nước dừa trân châu dừa',
                'Coconino - Tea & Cheese',
                'Cơm 79',
                'Cơm chay Diệu Vy',
                'Cơm chiên Linh Đông',
                'Cơm chú Fòng',
                'Cơm gà - Cháo ếch Singapore 68',
                'Cơm gà Đại Náo',
                'Cơm gà Đệ Nhất',
                'Cơm gà Hải Nam',
                'Cơm gà Tam Kỳ',
                'Cơm gà Tân Hải Nam',
                'Cơm gà xối mỡ 142',
                'Cơm gà xối mỡ Nũng Nịu',
                'Cơm gà xối mỡ quán Cô Ba',
                'Cơm gà xối mỡ Thạch Lam',
                'Cơm nhà Phố Thị',
                'Cơm Niêu Phương Bắc',
                'Cơm niêu Thiên Lý',
                'Cơm Sài Gòn',
                'Cơm tấm bụi Sài Gòn',
                'Cơm tấm Cali',
                'Cơm tấm Chị Hai',
                'Cơm tấm Cô Hoa',
                'Cơm tấm Kim Tiền 2',
                'Cơm tấm Kim Tiền 3',
                'Cơm tấm Làng',
                'Cơm tấm Long Xuyên',
                'Cơm tấm Mai',
                'Cơm tấm Mây',
                'Cơm tấm Minh Long',
                'Cơm tấm Ngô Quyền',
                'Cơm tấm Nhớ',
                'Cơm tấm Ni Mập',
                'Cơm tấm Phúc Lộc Thọ',
                'Cơm tấm Thuận Kiều',
                'Con Gà Mái - Cơm gà Phú Yên',
                'Con Sói - Sữa tươi trân châu đường đen',
                'Cộng Cà Phê',
                'Đại Kê Quán - Cơm Gà Phú Yên & Mì Xào',
                'Daily - Hủ tiếu & cà phê',
                'Đậu M Mix - Đậu Nành & Rau Má',
                'Đen Đá Café',
                'Dì Ba - Nem nướng Nha Trang',
                'Dì Bảy - Bún Mắm Nêm Đà Nẵng',
                'Domino\'s Pizza',
                'Double Tea',
                'Đức Ký mì gia',
                'Effoc Coffee',
                'FOX Tea - Trà sữa & ăn vặt',
                'Funny Beef',
                'Gà Cơ Bắp',
                'Gà Delichi - gà lên mâm & gà bó xôi',
                'Gà nướng Cái Bang',
                'Gà Nướng Ò Ó O',
                'Gà rán Chicken Plus',
                'Gà ta Ngon Số 1 - Cơm gà & Cháo gà & Gỏi',
                'GAM Coffee - Arabica Lắc Lọ Lem',
                'GENKI',
                'Gờ Cafe',
                'Góc Bánh Mì Chảo',
                'Gong Cha',
                'Guchi - Burger - Tokbokki Kimbap Và Cơm Trộn',
                'Gusto Food & Drink',
                'Guta Café',
                'Hai Tea - Tea & Coffee',
                'Hamburger Bò miếng Hapi',
                'Hancook Korea fast food',
                'Hanuri - Quán ăn Hàn Quốc',
                'Haocha Milk Tea',
                'Helios - Tiệm Ăn vặt 1999',
                'Highlands Coffee',
                'Highlands Coffee',
                'Hồng trà Ngô Gia',
                'Hot & Cold - Trà sữa & Xiên que',
                'HP Cơm tấm',
                'Hương Ký 9 - Cơm gà',
                'Jollibee',
                'Kem bơ - Trái cây tô 251',
                'Kem Bơ 251',
                'KFC',
                'Khoai Bistro - Bánh Mì Chảo, Beefsteak & Mì Ý',
                'Kimbap City',
                'KOI Thé',
                'Laha Cafe & Trà Sữa',
                'Lavida Coffee & Tea',
                'Loca - Bánh Canh Cá Lóc',
                'Lotteria',
                'McDonald\'s',
                'Mì cay Seoul',
                'Mì ốc hến Dì Lan',
                'Mì trộn Tên Lửa',
                'Mì Xí Mứng Chú Cuội ',
                'Mì Ý Double B',
                'Milano Coffee',
                'Mộc Vị Quán - Cơm dĩa nóng & mì quảng',
                'Neca Fresh Yogurt',
                'Ohzee - Bánh bèo & bánh bột lọc',
                'Otoke chicken',
                'Panda Coffee & Tea Express',
                'Papa Chicken',
                'PapaXốt - Cơm Xèo & Bò Bít Tết',
                'Passio Coffee',
                'Phát Ký - Cơm Chiên, Mì Xào & Hủ Tiếu Xào',
                'Phở 24',
                'Phở bò Cô Trang',
                'Phở bò Đan Phượng ',
                'Phúc Long',
                'Pizza 4P\'s',
                'Pizza Hut',
                'Popeyes',
                'Quán ăn Đức Ân 86',
                'Quán cơm Ngọc Hân',
                'Quán cơm Nguyễn Ký',
                'Quán Con Gà Mái - Cơm gà Phú Yên',
                'Quán Hủ Tiếu Mì Hoành 008',
                'Quán Hủ Tiếu Mực Cô Út Saigon',
                'R&B Tea',
                'Râu Cam - Gà rán, Pizza & Trà sữa',
                'Rau Má Mix',
                'Rau Má Pha Sài Gòn 1982',
                'Royaltea',
                'Sasin - Mì cay 7 cấp độ Hàn Quốc',
                'Shilin',
                'Shilin - Cơm Gà - Gà rán, Trà Đào Đài Loan',
                'Sóng Sánh Milk Tea',
                'Starbucks Coffee',
                'Sữa chua trân châu Hạ Long',
                'Subin Steak',
                'Sủi Cảo 193',
                'Sukiya',
                'Súp cua Mộc',
                'Súp cua Thu Hiền',
                'Susu\'s Burger & Bánh Tráng Long An',
                'Tân Phước Takoyaki',
                'Tàu Hũ Xe Lam',
                'Texas Chicken',
                'Thành Đạt - Hủ tiếu Nam Vang',
                'Thanh Hương - Bánh Xèo, Bánh Khọt, Bún Mắm Miền Tây',
                'The "MY QUẢNG" House & Coffee Bar',
                'The 1988 - Trà sữa & ăn vặt',
                'The Alley - Trà sữa Đài Loan',
                'The Coffee Cean & Tea Leaf',
                'The Coffee House',
                'Thế Giới Cơm Tấm - Bún Thịt Nướng - Bánh Mì Thịt Nướng',
                'The Hideout Tea & Coffee',
                'The Pizza Company',
                'Thèm - Trà Sữa Chống Khẩu Nghiệp',
                'Tiến Hải Quán - Bún đậu mắm tôm',
                'Tiger Sugar - Đường Nâu Sữa Đài Loan',
                'Toco Toco Bubble Tea',
                'Tous Le Jours',
                'Trà dâu Mr.Đỏ',
                'Trà sữa - Hồng trà Ngô Gia',
                'Trà Sữa & Milo Tutimi',
                'Trà sữa A Ngáo',
                'Trà sữa Ai Cha',
                'Trà sữa Bobapop',
                'Trà sữa Comebuy',
                'Trà sữa Dhi',
                'Trà Sữa Fozu',
                'Trà sữa Fulong Tea',
                'Trà sữa Heekcaa ',
                'Trà sữa Huy\'s',
                'Trà sữa MayCha',
                'Trà sữa Miutea',
                'Trà sữa Mộc',
                'Trà sữa Nọng',
                'Trà sữa Sacha',
                'Trà sữa Sunday Basic',
                'Trà sữa Te Amo',
                'Trà sữa Tiên Hưởng',
                'Trà sữa Tròn 94',
                'Trùm Mì Trộn',
                'Vitamin Bar - Sinh Tố',
                'Wow Chicken Gà Rán & Cơm Gà',
                'Xôi Bình Tiên',
                'Xộp - Món Hến & Ốc ',
                'Young Coffee & Tea',
            ]

            lstUpper = [str(i).upper() for i in lst]

            if str(val).upper() not in lstUpper:
                return False

            return True

        except Exception:
            return False


    @staticmethod
    def validate_No_Items(dfCheck):

        df = dfCheck.copy()

        # try:
        df['No_items'] = [0] * df.shape[0]

        for idx in df.index:
            for i in range(1, 21):

                if isinstance(df.loc[idx, f'Quantity {i}'], (int, float)) or df.loc[idx, f'Quantity {i}'] == 'NULL':

                    if df.loc[idx, f'Quantity {i}'] == 'NULL':
                        df.loc[idx, f'Quantity {i}'] = 0

                    df.loc[idx, 'No_items'] = df.loc[idx, 'No_items'] + df.loc[idx, f'Quantity {i}']

                else:
                    df.loc[idx, f'Quantity {i}'] = np.nan



            if not isinstance(df.loc[idx, 'No. items'], (int, float)):
                df.loc[idx, 'No. items'] = np.nan


        df.loc[(df['No. items'] == 0) | (df['No. items'] > 50), ['No. items']] = np.nan
        df.loc[(df['No. items'] != df['No_items']), ['No. items']] = np.nan


        dfCheck['No. items'] = df['No. items']

        # except Exception:
        #
        #     dfCheck['No. items'] = [np.nan] * dfCheck.shape[0]


        return dfCheck


    @staticmethod
    def validate_Quantity_AFV(dfCheck):

        df = dfCheck.copy()

        for i in range(1, 21):
            try:
                df.loc[df[f'Quantity {i}'] == 'NULL', [f'Quantity {i}']] = TEMP_VAL
                df.loc[df[f'AFV {i}'] == 'NULL', [f'AFV {i}']] = TEMP_VAL

                df.loc[(df[f'Quantity {i}'] > 50) & (df[f'Quantity {i}'] != TEMP_VAL), [f'Quantity {i}']] = np.nan

                for idx in df.index:

                    if isinstance(df.loc[idx, f'AFV {i}'], (int, float)):

                        if not float(df.loc[idx, f'AFV {i}']).is_integer():
                            df.loc[idx, f'AFV {i}'] = np.nan

                    else:
                        df.loc[idx, f'AFV {i}'] = np.nan


                df.loc[(df[f'Quantity {i}'] == 0) & (df[f'Food {i}'] != 'NULL'), [f'Quantity {i}']] = np.nan
                df.loc[(df[f'AFV {i}'] == 0) & (df[f'Food {i}'] != 'NULL'), [f'AFV {i}']] = np.nan

                df.loc[df[f'Quantity {i}'] == TEMP_VAL, [f'Quantity {i}']] = 'NULL'
                df.loc[df[f'AFV {i}'] == TEMP_VAL, [f'AFV {i}']] = 'NULL'

                dfCheck[f'Quantity {i}'] = df[f'Quantity {i}']
                dfCheck[f'AFV {i}'] = df[f'AFV {i}']

            except Exception:
                dfCheck[f'AFV {i}'] = [np.nan] * dfCheck.shape[0]

        return dfCheck