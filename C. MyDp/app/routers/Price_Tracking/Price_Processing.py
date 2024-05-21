import pandas as pd
import numpy as np
import json
import os
from datetime import datetime
from openpyxl.styles import Alignment, NamedStyle
from openpyxl.styles.fills import PatternFill
from openpyxl.utils.cell import get_column_letter
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive


class PriceProcess:
    def __int__(self, logger):
        self.logger = logger


    def convert_qme_file_to_df(self, qme_xlsx, int_week: int, is_bhx) -> pd.DataFrame:

        self.logger.info(f'Convert Q&me xlsx file to dataframe')

        if is_bhx:
            df_upload = pd.read_excel(qme_xlsx)
            df_upload.rename(columns={'Ngày Crawl data': 'Latest visit', 'Trạng thái': 'Trạng thái không bán?', 'Link on website BHX': 'Store Name'}, inplace=True)

            df_upload.loc[(df_upload['Giá bán'].isnull() & df_upload['Giá gốc'].isnull() & df_upload['Trạng thái không bán?'].isnull()), 'Trạng thái không bán?'] = 'NA'

            df_upload['Store Name'] = 'BHX Online'
            df_upload['Giá bán'].fillna(value=0, inplace=True)
            df_upload['Giá gốc'].fillna(value=0, inplace=True)

            df_upload.replace({'ERROR': 'NA'}, inplace=True)

            lst_new_col = [
                'Visual merchandising (VM)',
                'VM-linked category',
                'VM-linked brand',
                'User name',
                'Previous visit',
                'Sản phẩm có bán không?',
                'Sản phẩm có giá dành cho thành viên Watsons?',
                'Giá thành viên',
                'Sản phẩm có giảm giá không?',
                'Thời gian áp dụng giảm giá?',
                'Từ ngày',
                'Đến ngày',
                'Giá được lấy từ online?',
                'Có hình thức khuyến mãi khác?',
                'Có [Mua ... tặng ... không?]',
                'Mua với số lượng ...',
                'Thì tặng với số lượng ...',
                'Có [Mua ... với giá ... không?]',
                'Mua với số lượng là',
                'Thì có giá là',
            ]

            df_upload = pd.concat([df_upload, pd.DataFrame(columns=lst_new_col, data=[[np.nan] * len(lst_new_col)] * df_upload.shape[0])], axis=1)

            df_upload['Visual merchandising (VM)'] = df_upload['STT']
            df_upload['VM-linked category'] = 'Chăm sóc cá nhân BHX'
            df_upload['Sản phẩm có bán không?'] = [0 if a == 0 else 1 for a in df_upload['Giá bán']]
            df_upload['Sản phẩm có giá dành cho thành viên Watsons?'] = 0
            df_upload['Giá thành viên'] = 0
            df_upload['Sản phẩm có giảm giá không?'] = [0 if a == 0 else 1 for a in df_upload['Giá gốc']]
            df_upload['Thời gian áp dụng giảm giá?'] = 0
            df_upload['Giá được lấy từ online?'] = 1
            df_upload['Có hình thức khuyến mãi khác?'] = 0
            df_upload['Có [Mua ... tặng ... không?]'] = 0
            df_upload['Mua với số lượng ...'] = 0
            df_upload['Thì tặng với số lượng ...'] = 0
            df_upload['Có [Mua ... với giá ... không?]'] = 0
            df_upload['Mua với số lượng là'] = 0
            df_upload['Thì có giá là'] = 0

        else:
            try:
                df_upload = pd.read_excel(qme_xlsx, sheet_name='Worksheet')

                df_header = df_upload.loc[[2, 3], :].T.copy()
                df_header['header'] = [a if pd.isnull(b) else b for a, b in zip(df_header[2], df_header[3])]
                lst_header = df_header['header'].values.tolist()
                lst_header[0] = 'STT'

                df_upload.drop([0, 1, 2, 3, 4, 5], inplace=True)
                df_upload.columns = lst_header
                df_upload.reset_index(drop=True, inplace=True)
                df_upload['STT'] = [int(a.split(".")[0]) for a in df_upload['Visual merchandising (VM)']]

                # df_upload.to_excel('aaa.xlsx')

            except Exception:
                df_upload = pd.read_csv(qme_xlsx)

                df_header = df_upload.loc[[2], :].T.copy()
                for idx in df_header.index:
                    if '_Act_' in str(df_header.at[idx, 2]):
                        df_header.at[idx, 2] = str(df_header.at[idx, 2]).split('_')[1]
                lst_header = df_header[2].values.tolist()
                lst_header[0] = 'STT'

                df_upload.drop([0, 1, 2], inplace=True)
                df_upload.columns = lst_header
                df_upload.reset_index(drop=True, inplace=True)
                df_upload['STT'] = [int(a.split(".")[0]) for a in df_upload['Visual merchandising (VM)']]

                for col in df_upload.columns:
                    try:
                        df_upload[col] = df_upload[col].astype(int)
                    except Exception:
                        pass

            df_upload['Store Name'] = 'Watsons'

            df_upload.loc[(df_upload['Sản phẩm có bán không?'] == 0) & (pd.isnull(df_upload['Trạng thái không bán?'])), ['Trạng thái không bán?']] = ['NA']


        # REMOVE DUPLICATE AND KEEP LATEST DATE
        df_upload['Latest visit'] = df_upload['Latest visit'].astype('datetime64[ns]')
        df_upload.sort_values(by=['Store Name', 'STT', 'Latest visit'], inplace=True)
        df_upload.drop_duplicates(subset=['Store Name', 'STT'], keep='last', inplace=True)
        df_upload.reset_index(drop=True, inplace=True)

        # df_upload.to_excel('zzzz_df_upload.xlsx')


        # VALIDATE UPLOADED FILE--------------------------------------------------------------------------------
        df_upload.loc[(df_upload['Store Name'] == 'BHX Online'), ['Giá được lấy từ online?']] = [0]
        df_upload['ERRORS_NOTE'] = [''] * df_upload.shape[0]
        df_upload["Agency's comment"] = [''] * df_upload.shape[0]

        for idx in df_upload.index:
            lst_err = list()

            if ',' in str(df_upload.at[idx, 'Giá bán']) or ';' in str(df_upload.at[idx, 'Giá bán']):
                gia_ban = int(str(df_upload.at[idx, 'Giá bán']).replace(',', '').replace(';', ''))
            else:
                gia_ban = int(df_upload.at[idx, 'Giá bán'])

            if ',' in str(df_upload.at[idx, 'Giá gốc']) or ';' in str(df_upload.at[idx, 'Giá gốc']):
                gia_goc = int(str(df_upload.at[idx, 'Giá gốc']).replace(',', '').replace(';', ''))
            else:
                gia_goc = int(df_upload.at[idx, 'Giá gốc'])

            # ADD NEW ON MAY 02, 2024
            if ',' in str(df_upload.at[idx, 'Giá thành viên']) or ';' in str(df_upload.at[idx, 'Giá thành viên']):
                gia_member = int(str(df_upload.at[idx, 'Giá thành viên']).replace(',', '').replace(';', ''))
            else:
                gia_member = int(df_upload.at[idx, 'Giá thành viên'])

            df_upload.at[idx, 'Giá bán'] = gia_ban
            df_upload.at[idx, 'Giá gốc'] = gia_goc
            df_upload.at[idx, 'Giá thành viên'] = gia_member

            ngay_pv = df_upload.at[idx, 'Latest visit']

            tu_ngay = df_upload.at[idx, 'Từ ngày']
            if tu_ngay and not pd.isnull(tu_ngay):

                try:
                    tu_ngay = datetime.strptime(tu_ngay, "%d/%m/%Y")
                except Exception:
                    # 'Mar 12; 2024'
                    tu_ngay = str(tu_ngay).replace(';', '')
                    tu_ngay = datetime.strptime(tu_ngay, "%b %d %Y")

                df_upload.at[idx, 'Từ ngày'] = tu_ngay

            den_ngay = df_upload.at[idx, 'Đến ngày']
            if den_ngay and not pd.isnull(den_ngay):
                try:
                    den_ngay = datetime.strptime(den_ngay, "%d/%m/%Y")
                except Exception:
                    # 'Mar 12; 2024'
                    den_ngay = str(den_ngay).replace(';', '')
                    den_ngay = datetime.strptime(den_ngay, "%b %d %Y")

                df_upload.at[idx, 'Đến ngày'] = den_ngay

            buy_membership = df_upload.at[idx, "Sản phẩm có giá dành cho thành viên Watsons?"]
            buy_x_get_y = df_upload.at[idx, "Có [Mua ... tặng ... không?]"]
            buy_x_with_price_y = df_upload.at[idx, "Có [Mua ... với giá ... không?]"]

            if df_upload.at[idx, 'Sản phẩm có bán không?'] == 1:
                if not gia_ban > 0:
                    lst_err.append('Missing giá bán')

                # ADD NEW ON MAY 02, 2024
                if df_upload.at[idx, 'Sản phẩm có giá dành cho thành viên Watsons?'] == 1 and not gia_member > 0:

                    lst_err.append('Missing giá thành viên')

                if gia_member >= gia_ban > 0:
                    lst_err.append('Sai data giá bán phải lớn hơn giá thành viên')

                if gia_member >= gia_goc > 0:
                    lst_err.append('Sai data giá gốc phải lớn hơn giá thành viên')
                # END ADD NEW ON MAY 02, 2024

                if df_upload.at[idx, 'Sản phẩm có giảm giá không?'] == 1 and not gia_goc > 0:
                    lst_err.append('Missing giá gốc')

                if gia_ban >= gia_goc > 0:
                    lst_err.append('Sai data giá gốc phải lớn hơn giá bán')

                if df_upload.at[idx, 'Thời gian áp dụng giảm giá?'] == 1 and (
                        pd.isnull(tu_ngay) or pd.isnull(den_ngay)):
                    lst_err.append("Missing data 'Promo Start Date' & 'Promo End Date'")

                if not pd.isnull(tu_ngay) and not pd.isnull(den_ngay):
                    if tu_ngay > den_ngay:
                        lst_err.append("Sai data 'Promo Start Date' phải trước 'Promo End Date'")

                if not pd.isnull(den_ngay):
                    if den_ngay < ngay_pv:
                        lst_err.append("Sai data 'Promo End Date' phải sau 'Interview Date'")

                if df_upload.at[idx, 'Sản phẩm có giảm giá không?'] == 1 and df_upload.at[idx, 'Có hình thức khuyến mãi khác?'] == 1:
                    lst_err.append("Sai data 'Sản phẩm có giảm giá không?' & 'Có hình thức khuyến mãi khác?' không được đồng thời = 1")


                # ADD NEW ON MAY 02, 2024
                if buy_membership == 1:

                    if gia_ban > 0 and gia_member > 0 and gia_goc > 0:
                        df_upload.at[idx, 'Giá bán'] = gia_member
                        df_upload.at[idx, 'Sản phẩm có giảm giá không?'] = 1
                        df_upload.at[idx, "Agency's comment"] = f"Non-member price: {'{:,.0f}'.format(gia_ban)}"

                    if gia_ban > 0 and gia_member > 0 and gia_goc == 0:
                        df_upload.at[idx, 'Giá gốc'] = gia_ban
                        df_upload.at[idx, 'Giá bán'] = gia_member
                        df_upload.at[idx, 'Sản phẩm có giảm giá không?'] = 1
                        df_upload.at[idx, "Agency's comment"] = f"Non-member price: {'{:,.0f}'.format(gia_ban)}"
                # END ADD NEW ON MAY 02, 2024


                if df_upload.at[idx, 'Có hình thức khuyến mãi khác?'] == 1:
                    if buy_x_get_y + buy_x_with_price_y > 1:
                        lst_err.append("Chỉ được chọn 1 trong 2 chương trình 'Mua...tặng...' hoặc 'Mua...với giá...'")
                        continue

                    if buy_x_get_y + buy_x_with_price_y != 1:
                        lst_err.append("Phải chọn 1 trong 2 chương trình 'Mua...tặng...' hoặc 'Mua...với giá...'")
                        continue

                    # Calculate 'Giá bán' & 'Giá gốc' base on 'Có hình thức khuyến mãi khác?'
                    if buy_x_get_y == 1:
                        buy_x = df_upload.at[idx, 'Mua với số lượng ...']
                        get_y = df_upload.at[idx, 'Thì tặng với số lượng ...']
                        df_upload.at[idx, 'Giá gốc'] = gia_ban
                        df_upload.at[idx, 'Giá bán'] = (gia_ban * buy_x) / (buy_x + get_y)
                        df_upload.at[idx, 'Sản phẩm có giảm giá không?'] = 1

                        # ADD NEW ON MAY 02, 2024
                        str_Agency_comment = df_upload.at[idx, "Agency's comment"]

                        if len(str_Agency_comment) == 0:
                            str_Agency_comment = f'Buy {buy_x} Get {get_y}'
                        else:
                            str_Agency_comment += f' | Buy {buy_x} Get {get_y}'
                        # END ADD NEW ON MAY 02, 2024

                        df_upload.at[idx, "Agency's comment"] = str_Agency_comment
                        continue

                    if buy_x_with_price_y == 1:
                        buy_x = df_upload.at[idx, 'Mua với số lượng là']
                        with_price_y = int(str(df_upload.at[idx, 'Thì có giá là']).replace(',', '').replace(';', ''))
                        df_upload.at[idx, 'Thì có giá là'] = with_price_y
                        df_upload.at[idx, 'Sản phẩm có giảm giá không?'] = 1

                        # ADD NEW ON MAY 02, 2024
                        str_Agency_comment = df_upload.at[idx, "Agency's comment"]

                        if len(str_Agency_comment) == 0:
                            str_Agency_comment = f'Buy {buy_x} Pay {format(with_price_y, ",.0f")}'
                        else:
                            str_Agency_comment += f' | Buy {buy_x} Pay {format(with_price_y, ",.0f")}'
                        # END ADD NEW ON MAY 02, 2024

                        df_upload.at[idx, "Agency's comment"] = str_Agency_comment

                        if buy_x == 0:
                            lst_err.append("Missing data 'Mua với số lượng là'")
                            continue

                        if with_price_y == 0:
                            lst_err.append("Missing data 'Thì có giá là'")
                            continue

                        if buy_x > 0 and with_price_y > 0:
                            df_upload.at[idx, 'Giá gốc'] = gia_ban
                            df_upload.at[idx, 'Giá bán'] = with_price_y / buy_x
                            continue
                else:
                    if buy_x_get_y + buy_x_with_price_y > 0:
                        lst_err.append("Dư data  'Mua...tặng...', 'Mua...với giá...'")

            else:
                if df_upload.at[idx, 'Trạng thái không bán?'] not in ['NA', 'OOS']:
                    lst_err.append('Missing trạng thái không bán')

                if not pd.isnull(tu_ngay) or not pd.isnull(den_ngay):
                    lst_err.append("Dư data 'Promo Start Date' & 'Promo End Date'")

                lst_col_name = [
                    'Giá bán',
                    'Sản phẩm có giá dành cho thành viên Watsons?',
                    'Giá thành viên',
                    'Sản phẩm có giảm giá không?',
                    'Giá gốc',
                    'Thời gian áp dụng giảm giá?',
                    'Giá được lấy từ online?',
                    'Có hình thức khuyến mãi khác?',
                    'Có [Mua ... tặng ... không?]',
                    'Mua với số lượng ...',
                    'Thì tặng với số lượng ...',
                    'Có [Mua ... với giá ... không?]',
                    'Mua với số lượng là',
                    'Thì có giá là',
                ]

                for col_name in lst_col_name:
                    val_col = int(df_upload.at[idx, col_name])
                    if val_col > 0:
                        lst_err.append(f"Dư data '{col_name}'")

            df_upload.at[idx, 'ERRORS_NOTE'] = ' | '.join(lst_err)

        df_upload['Giá được lấy từ online?'].replace({1: 'Online', 0: np.nan}, inplace=True)
        df_upload["Agency's comment"] = [a if pd.isnull(b) else (b if a == '' else f'{a} | {b}') for a, b in zip(df_upload["Agency's comment"], df_upload['Giá được lấy từ online?'])]
        df_upload.loc[(df_upload['Giá bán'] == 0) & (df_upload['Trạng thái không bán?'] == 'OOS'), ['Giá bán']] = [99]

        # END VALIDATE UPLOADED FILE----------------------------------------------------------------------------

        dict_col_compare = {
            "STT": "STT",
            "Week Number": None,
            "Data Collection Start Date (Fri)": "Latest visit",
            "Field Work Date": 'Latest visit',  # "Ngày khảo sát giá",
            "Competitor": "Store Name",
            "Competitor Price": "Giá bán",
            "Promo Indicator": 'Sản phẩm có giảm giá không?',
            "Promo Details OFF": None,
            "Promo Details Reg Price": "Giá gốc",
            "Competitor Promo Type": 'Sản phẩm có giảm giá không?',
            "Others Remarks": None,
            "Agency's comment": "Agency's comment",

            # Addin for internal checking
            "Promo Start Date": "Từ ngày",
            "Promo End Date": "Đến ngày",
            "ERRORS_NOTE": "ERRORS_NOTE",
        }

        lst_col_compare = list(dict_col_compare.keys())

        df_sku_data = pd.DataFrame(columns=lst_col_compare, data=[[np.nan] * len(lst_col_compare)] * df_upload.shape[0])

        for key, val in dict_col_compare.items():
            if val is not None:
                df_sku_data[key] = df_upload[val].values

                if key in ['Data Collection Start Date (Fri)', 'Field Work Date']:
                    df_sku_data[key] = df_sku_data[key].astype('datetime64[ns]')

        del df_upload

        df_sku_data['Week Number'] = int_week

        # Set 'Promo Indicator' & 'Competitor Promo Type'
        df_sku_data['Promo Indicator'].replace({1: '1 - Yes', 0: '0 - No'}, inplace=True)
        df_sku_data['Competitor Promo Type'].replace({1: 'Promo', 0: 'Non promo'}, inplace=True)

        # Set 'Promo Details OFF'
        df_sku_data['Promo Details OFF'] = [1 - (com_pri / reg_pri) if reg_pri > 0 else np.nan for com_pri, reg_pri in
                                            zip(df_sku_data['Competitor Price'],
                                                df_sku_data['Promo Details Reg Price'])]

        # Replacing
        df_sku_data['Promo Details Reg Price'].replace({0: np.nan}, inplace=True)
        df_sku_data['Competitor'].replace({'BHX Online': 'BHX', 'Watsons Bitexco': 'Watsons', 'Watsons Online': 'Watsons', 'Watsons Vincom': 'Watsons'}, inplace=True)

        df_sku_data.reset_index(drop=True, inplace=True)

        return df_sku_data


    def validate_df_sku_data(self, prj_sku_info: dict, df_sku_data: pd.DataFrame) -> list:

        lst_sku_data_err = list()

        # CHECK DUPLICATED
        df_check = df_sku_data.copy()
        store_stt = df_check['Competitor'] + '_' + df_check['STT'].astype(str)
        df_err = df_check[store_stt.isin(store_stt[store_stt.duplicated()])]

        if not df_err.empty:
            lst_sku_data_err.append(['DUPLICATED SKU', df_err.to_html(index=False)])


        # CHECK BASE MUST BE = 250 (150 for Watsons, 100 for BHX)
        df_sku_info = pd.DataFrame.from_dict(json.loads(prj_sku_info['sku_info']))
        del prj_sku_info

        df_sku_info['store_stt'] = df_sku_info['Sheet'] + '_' + df_sku_info['STT'].astype(str)
        df_err = df_sku_info.query(f"not store_stt.isin({store_stt.values.tolist()})")

        if not df_err.empty:
            count_watson = df_sku_info.query(f"store_stt.isin({store_stt.values.tolist()}) & Sheet == 'Watsons'").shape[0]
            count_bhx = df_sku_info.query(f"store_stt.isin({store_stt.values.tolist()}) & Sheet == 'BHX'").shape[0]

            df_err = df_err.loc[:, ['Sheet', 'STT', 'Product Description']]

            lst_sku_data_err.append([f'Watsons = {count_watson}/150|BHX = {count_bhx}/100|Please check MISSING SKU', df_err.to_html(index=False)])


        # CHECK MISSING VALUE
        lst_col_check = [
            'Data Collection Start Date (Fri)',
            'Field Work Date',
            'Competitor Price',
            'Promo Indicator',
            'Competitor Promo Type',
        ]

        for col_name in lst_col_check:
            val_check = self.validate_sku_data_missing_value(df_check=df_check, col_check_name=col_name)
            if val_check:
                lst_sku_data_err.append(val_check)

        return lst_sku_data_err


    @staticmethod
    def validate_sku_data_missing_value(df_check: pd.DataFrame, col_check_name: str) -> list:

        df_err = df_check.loc[df_check[col_check_name].isnull(), ['STT', 'Competitor', col_check_name]]
        if not df_err.empty:
            return ['Please check MISSING VALUE', df_err.to_html(index=False)]

        return []


    def export_xlsx_data(self, **kwargs) -> str:

        df_data: pd.DataFrame = kwargs['df_data']
        dict_dfs_prj_sku_data: dict = kwargs['dict_dfs_prj_sku_data']
        int_week: int = kwargs['int_week']
        int_pre_week: int = kwargs['int_pre_week']
        is_to_client: bool = kwargs['is_to_client']
        is_to_ggdrive: bool = kwargs['is_to_ggdrive']
        file_name: str = kwargs['file_name']

        for key, val in dict_dfs_prj_sku_data.items():

            if key == int_week:
                df_sku_data = val
                df_sku_data['Data Collection Start Date (Fri)'] = pd.to_datetime(df_sku_data['Data Collection Start Date (Fri)'])
                df_sku_data['Field Work Date'] = pd.to_datetime(df_sku_data['Field Work Date'])
                df_sku_data['Promo Start Date'] = pd.to_datetime(df_sku_data['Promo Start Date'])
                df_sku_data['Promo End Date'] = pd.to_datetime(df_sku_data['Promo End Date'])

                df_sku_data['sheet_stt'] = df_sku_data['Competitor'] + '_' + df_sku_data['STT'].astype(str)
                df_sku_data.set_index('sheet_stt', inplace=True)
                df_sku_data.drop(columns=['STT'], inplace=True)

                lst_order_col = [
                    "Sheet", "STT", "Week Number", "Data Collection Start Date (Fri)", "Field Work Date", "Market",
                    "Banner", "Article Number", "Aricle Number (banner)", "BARCODE", "Product Description",
                    "DF Product Description", "Division (Fresh/Grocery)", "Competitor", "Store Name",
                    "Competitor no.", "Competitor product information", "Competitor Price", "Checked Product Size",
                    "Checked Product UOM", "Promo Indicator", "Promo Details OFF", "Promo Details Reg Price",
                    "Competitor Promo Type", "Others Remarks", "Agency's comment", "Link",
                    "Promo Start Date", "Promo End Date", "ERRORS_NOTE"
                ]

                df_data = pd.concat([df_data, df_sku_data], axis=1)
                df_data["ERRORS_NOTE"].replace({np.nan: ''}, inplace=True)

                df_data = df_data.reindex(columns=lst_order_col)

                if is_to_client:
                    df_data["Promo Details OFF"] = [
                        np.nan if pd.isnull(pct)
                        else f"OFF {'{:.0f}'.format(pct * 100)}% - reg. price: {'{:,.0f}'.format(pri)}"
                        for pct, pri in zip(df_data["Promo Details OFF"], df_data["Promo Details Reg Price"])]

                    df_data.rename(columns={"Promo Details OFF": "Promo Details"}, inplace=True)
                    df_data.drop(columns=["Promo Details Reg Price"], inplace=True)

            else:

                lst_order_col = ['STT', 'Competitor', 'Competitor Price', 'Promo Details OFF', 'Promo Details Reg Price', "Agency's comment"]
                df_sku_data = val.loc[:, lst_order_col]

                df_sku_data['sheet_stt'] = df_sku_data['Competitor'] + '_' + df_sku_data['STT'].astype(str)
                df_sku_data.set_index('sheet_stt', inplace=True)
                df_sku_data.drop(columns=['STT', 'Competitor'], inplace=True)

                if is_to_client:
                    df_sku_data["Promo Details OFF"] = [
                        np.nan if pd.isnull(pct)
                        else f"OFF {'{:.0f}'.format(pct * 100)}% - reg. price: {'{:,.0f}'.format(pri)}"
                        for pct, pri in zip(df_sku_data["Promo Details OFF"], df_sku_data["Promo Details Reg Price"])]

                    df_sku_data.rename(columns={
                        'Competitor Price': f'Price W{key}',
                        'Promo Details OFF': f'Promo W{key}'
                    }, inplace=True)

                    df_sku_data.drop(columns=["Promo Details Reg Price", "Agency's comment"], inplace=True)

                else:
                    df_sku_data.rename(columns={
                        'Competitor Price': f'Price W{key}',
                        'Promo Details OFF': f'Promo OFF W{key}',
                        'Promo Details Reg Price': f'Promo Reg Price W{key}',
                        "Agency's comment": f"Agency's comment W{key}"
                    }, inplace=True)

                df_data = pd.concat([df_data, df_sku_data], axis=1)

                if key == int_pre_week and not is_to_client:

                    for idx in df_data.index:
                        str_err = str()

                        str_cmt_cur = str(df_data.at[idx, f"Agency's comment"]).upper()
                        str_cmt_pre = str(df_data.at[idx, f"Agency's comment W{key}"]).upper()

                        if 'ONLINE' in str_cmt_cur and 'ONLINE' not in str_cmt_pre:
                            str_err = "QC check địa bàn sản phẩm này có bán, có giá không?"

                        # ----------------------------------------------------------------------------------------------
                        # price_cur_week
                        # 'Competitor Price' | 'Promo Details Reg Price'

                        # price_pre_week
                        # f'Price W{key}' | f'Promo Reg Price W{key}'

                        if df_data.at[idx, 'Competitor Price'] in [0, 99] or df_data.at[idx, f'Price W{key}'] in [0, 99]:
                            if str_err:
                                df_data.at[idx, "ERRORS_NOTE"] = str_err
                            continue

                        price_cur_week = df_data.at[idx, 'Promo Details Reg Price'] if df_data.at[idx, 'Promo Details Reg Price'] > 99 else df_data.at[idx, 'Competitor Price']
                        price_pre_week = df_data.at[idx, f'Promo Reg Price W{key}'] if df_data.at[idx, f'Promo Reg Price W{key}'] > 99 else df_data.at[idx, f'Price W{key}']

                        gap = self.cal_price_gap(price_cur_week, price_pre_week)

                        if abs(gap) > 10:
                            str_price_err = f"Check chênh lệch giá W{int_week}={format(price_cur_week, ',.0f')} {'TĂNG' if gap > 0 else 'GIẢM'} {format(abs(gap), '.2f')}% so với W{key}={format(price_pre_week, ',.0f')})"

                            if str_err:
                                str_err += f' | {str_price_err}'
                            else:
                                str_err = str_price_err

                        # ALWAYS CHECK 'Competitor Price' vs f'Price W{key}' WHEN BOTH have promotion
                        if not pd.isnull(df_data.at[idx, 'Promo Details OFF']) and not pd.isnull(df_data.at[idx, f'Promo OFF W{key}']):
                            price_cur_week = df_data.at[idx, 'Competitor Price']
                            price_pre_week = df_data.at[idx, f'Price W{key}']

                            gap = self.cal_price_gap(price_cur_week, price_pre_week)

                            if abs(gap) > 10:
                                str_price_err = f"Check chênh lệch giá W{int_week}={format(price_cur_week, ',.0f')} {'TĂNG' if gap > 0 else 'GIẢM'} {format(abs(gap), '.2f')}% so với W{key}={format(price_pre_week, ',.0f')})"

                                if str_err:
                                    str_err += f' | {str_price_err}'
                                else:
                                    str_err = str_price_err

                        if len(df_data.at[idx, "ERRORS_NOTE"]) == 0:
                            df_data.at[idx, "ERRORS_NOTE"] = str_err
                        else:
                            df_data.at[idx, "ERRORS_NOTE"] = ' | '.join([df_data.at[idx, "ERRORS_NOTE"], str_err])

        self.format_xlsx_data(df_data=df_data, is_to_client=is_to_client, file_name=f'{file_name}.xlsx', int_cur_week=int_week, int_pre_week=int_pre_week)

        full_file_name = str()
        if is_to_ggdrive:
            full_file_name = self.upload_xlsx_data_to_ggdrive(file_name=file_name,
                                                              str_folder_id='14hcTXJ8ffa5PGkHoBcg1rMHSp8nsleM_',
                                                              is_to_client=is_to_client)

        return full_file_name if is_to_ggdrive else f'{file_name}.xlsx'


    @staticmethod
    def cal_price_gap(price_cur_week: int, price_pre_week: int) -> float:
        gap = (price_cur_week - price_pre_week) * 100 / price_pre_week
        return gap


    @staticmethod
    def format_xlsx_data(df_data: pd.DataFrame, is_to_client: bool, file_name: str, int_cur_week: int, int_pre_week: int):
        if is_to_client:
            df_data.drop(columns=['Sheet', 'Promo Start Date', 'Promo End Date', 'ERRORS_NOTE'], inplace=True)
            df_data['Week Number'].replace({int_cur_week: int(str(int_cur_week)[-2:])}, inplace=True)


        dict_sheet = {
            'Watsons': df_data.query("Competitor == 'Watsons'").copy(),
            'BHX': df_data.query("Competitor == 'BHX'").copy()
        }

        if not is_to_client:
            if int_pre_week < 0:
                lst_col_to_show = [
                    'STT', 'Aricle Number (banner)', 'BARCODE', 'Product Description',
                    'Competitor Price', 'Promo Details Reg Price', 'Promo Details OFF',
                    'Promo Start Date', 'Promo End Date', "Agency's comment",
                    'ERRORS_NOTE'
                ]

                for ws_name in ['Watsons', 'BHX']:
                    dict_sheet[ws_name] = dict_sheet.get(ws_name).loc[:, lst_col_to_show]
                    dict_sheet[ws_name].rename(columns={
                        'Competitor Price': f'Price W{int_cur_week}',
                        'Promo Details Reg Price': f'Regular price W{int_cur_week}',
                        'Promo Details OFF': f'Promo OFF W{int_cur_week}',
                        'ERRORS_NOTE': 'ERRORS'
                    }, inplace=True)
            else:
                lst_col_to_show = [
                    'STT', 'Aricle Number (banner)', 'BARCODE', 'Product Description',
                    'Competitor Price', 'Promo Details Reg Price', 'Promo Details OFF',
                    'Promo Start Date', 'Promo End Date', "Agency's comment",
                    f'Price W{int_pre_week}', f'Promo Reg Price W{int_pre_week}', f'Promo OFF W{int_pre_week}', 'ERRORS_NOTE'
                ]

                for ws_name in ['Watsons', 'BHX']:
                    dict_sheet[ws_name] = dict_sheet.get(ws_name).loc[:, lst_col_to_show]
                    dict_sheet[ws_name].rename(columns={
                        'Competitor Price': f'Price W{int_cur_week}',
                        'Promo Details Reg Price': f'Regular price W{int_cur_week}',
                        'Promo Details OFF': f'Promo OFF W{int_cur_week}',

                        f'Promo Reg Price W{int_pre_week}': f'Regular price W{int_pre_week}',

                        'ERRORS_NOTE': 'ERRORS'
                    }, inplace=True)


        fil_color = {
            'gray': PatternFill(patternType='solid', fgColor='D9D9D9'),
            'blue': PatternFill(patternType='solid', fgColor='6D9EEB'),
            'green': PatternFill(patternType='solid', fgColor='B6D7A8'),
            'orange': PatternFill(patternType='solid', fgColor='FF9900'),
            'red': PatternFill(patternType='solid', fgColor='FF0000'),
        }

        date_style_1 = NamedStyle(name='date_style_1', number_format='MM/DD/YYYY HH:MM')
        date_style_2 = NamedStyle(name='date_style_2', number_format='MM/DD/YYYY')
        num_style_1 = NamedStyle(name='num_style_1', number_format='0')
        num_style_2 = NamedStyle(name='num_style_2', number_format='0,000')
        num_style_3 = NamedStyle(name='num_style_3', number_format='0%')

        with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
            for key, val in dict_sheet.items():

                val.to_excel(writer, sheet_name=key, index=False)

                # set index column width
                ws = writer.sheets[key]
                ws.row_dimensions[1].height = 60
                ws.freeze_panes = ws['B2']

                for col in range(1, ws.max_column + 1):

                    col_letter = get_column_letter(col)
                    cur_cell = ws[f"{col_letter}1"]

                    ws.column_dimensions[col_letter].best_fit = True

                    cur_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

                    if cur_cell.value in ['STT', 'Link', 'Promo Start Date', 'Promo End Date']:
                        cur_cell.fill = fil_color['gray']
                    elif cur_cell.value in ["Market", "Banner", "Article Number", "Aricle Number (banner)", "BARCODE", "Product Description", "DF Product Description", "Division (Fresh/Grocery)", "Store Name", "Competitor no.", "Competitor product information", "Checked Product Size", "Checked Product UOM"]:
                        cur_cell.fill = fil_color['green']
                    elif cur_cell.value in ["Week Number", "Data Collection Start Date (Fri)", "Field Work Date", "Competitor", "Competitor Price", "Promo Indicator", "Promo Details", "Promo Details OFF", "Promo Details Reg Price", "Competitor Promo Type", "Others Remarks", "Agency's comment"]:
                        cur_cell.fill = fil_color['blue']
                    elif cur_cell.value in ["ERRORS"]:
                        cur_cell.fill = fil_color['red']
                    else:
                        cur_cell.fill = fil_color['orange']

                    for row in range(2, ws.max_row + 1):
                        cur_cell_row = ws[f"{col_letter}{row}"]

                        if cur_cell.value in ["Data Collection Start Date (Fri)"]:
                            cur_cell_row.style = date_style_1
                            ws.column_dimensions[col_letter].width = 16

                        elif cur_cell.value in ["Field Work Date", "Promo Start Date", "Promo End Date"]:
                            cur_cell_row.style = date_style_2
                            ws.column_dimensions[col_letter].width = 11

                        elif cur_cell.value in ['Article Number']:
                            ws.column_dimensions[col_letter].width = 10

                        elif cur_cell.value in ['Aricle Number (banner)']:
                            ws.column_dimensions[col_letter].width = 15

                        elif cur_cell.value in ["BARCODE"]:
                            cur_cell_row.style = num_style_1
                            ws.column_dimensions[col_letter].width = 14

                        elif cur_cell.value in ['Product Description']:
                            ws.column_dimensions[col_letter].width = 55

                        elif 'Price' in cur_cell.value or 'price' in cur_cell.value:
                            cur_cell_row.style = num_style_2

                            if 'Reg Price' not in cur_cell.value and 'Regular price' not in cur_cell.value:

                                if cur_cell_row.value in [0, '']:
                                    cur_cell_row.value = "NA"

                                if cur_cell_row.value == 99:
                                    cur_cell_row.value = "OOS"

                        elif cur_cell.value in ['Promo Details']:
                            ws.column_dimensions[col_letter].width = 28

                        elif 'Promo W' in cur_cell.value:
                            if is_to_client:
                                ws.column_dimensions[col_letter].width = 28

                        elif cur_cell.value in ['Promo Details OFF'] or 'Promo OFF W' in cur_cell.value:
                            cur_cell_row.style = num_style_3


    def upload_xlsx_data_to_ggdrive(self, file_name: str, str_folder_id: str, is_to_client: bool) -> str:

        self.logger.info(f"Upload '{file_name}' to Google Drive")

        gauth = GoogleAuth()
        drive = GoogleDrive(gauth)

        ver = 0
        file_list = drive.ListFile({'q': f"'{str_folder_id}' in parents and trashed=False"}).GetList()

        for file in file_list:
            if is_to_client:
                if file['title'] == f'{file_name}.xlsx':
                    file.Delete()
            else:
                if 'INTERNAL CHECKING' in file['title'].upper():
                    pre_ver = int(file['title'].rsplit('_', 1)[1].replace('v', '').replace('.xlsx', ''))
                    if pre_ver > ver:
                        ver = pre_ver

        ver += 1
        full_file_name = f'{file_name}.xlsx' if is_to_client else f'{file_name}_v{ver}.xlsx'

        if full_file_name != f'{file_name}.xlsx':
            os.rename(f'{file_name}.xlsx', full_file_name)

        gfile = drive.CreateFile({'parents': [{'id': f'{str_folder_id}'}]})
        gfile.SetContentFile(full_file_name)
        gfile.Upload()

        return full_file_name

