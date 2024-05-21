from scipy import stats
import pandas as pd
import numpy as np
import openpyxl
import traceback
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from scipy.stats import pearsonr
import time
from app.routers.MSN.Old.MSN_Addin_Variables import AddinVariables
# from .MSN_Data_Table import DataTable


class ToplineExporter:

    def __init__(self, prj: dict, df: pd.DataFrame, dictVarName: dict, dictValLbl: dict, lstSPCodes: list, export_section: str):
        self.prj = prj
        self.df, self.dictVarName, self.dictValLbl = df, dictVarName, dictValLbl
        self.export_section_name = self.prj['detail']['sections'][export_section]['name']

        self.toplineName = f"{self.prj['internal_id']}_{self.prj['name']}_{self.export_section_name}_Topline.xlsx"
        self.toplineTitle = f"{self.prj['name']}_{self.export_section_name}"
        self.toplineFile = None

        self.productCode = lstSPCodes
        self.isDisplayPctSign = self.prj['detail']['topline_design']['is_display_pct_sign']
        self.isJR3Factors = self.prj['detail']['topline_design']['is_jar_scale_3']

        self.dictSide, self.dictHeader = dict(), dict()

        # Final Result container
        self.dictTtest = dict()
        self.dictUA = dict()

        self.dfCorr = pd.DataFrame
        self.dictCorrCols = dict()



    def getInfo_RunSig(self):

        try:
            isSuccess = self.getInfo()
            if not isSuccess[0]:
                return isSuccess

            # # New way to process data tables
            # dTable = DataTable(self.dictSide, self.dictHeader, self.export_section_name, self.productCode,
            #                    self.dictValLbl, self.isJR3Factors, self.isDisplayPctSign, self.df)
            # dTable.create_tbl()

            # Don't delete
            st = time.process_time()
            self.run_topline_format()
            et = time.process_time()
            print('CPU Execution time:', et - st, 'seconds')

            return True, None

        except Exception:
            print(traceback.format_exc())
            return False, traceback.format_exc()


    def toExcel(self, lstSheet):

        try:

            wb = openpyxl.Workbook()
            fileName = self.toplineName

            if 'Handcount' in lstSheet:
                self.toSummary(wb, self.dictTtest, lstQreType=['OL', 'FC'])

            if 'Summary' in lstSheet:
                self.toSummary(wb, self.dictTtest)

            if 'Tabulation' in lstSheet:
                self.toTabulation(wb, self.dictTtest)

            if 'OL_Summary' in lstSheet:
                self.toOlJrSummary(wb, self.dictTtest, 'OL')

            if 'JR_Summary' in lstSheet:
                self.toOlJrSummary(wb, self.dictTtest, 'JR')

            if 'UA' in lstSheet:
                self.toUandA(wb, self.dictUA)

            if 'Correlation' in lstSheet:
                self.toCorr(wb)

            wb.remove(wb['Sheet'])

            wb.save(fileName)
            print('Workbook saved')

            wb.close()
            print('Workbook closed')

            print('----------------------------Topline exporting completed----------------------------')

            return True, None

        except Exception:
            print(traceback.format_exc())
            return False, traceback.format_exc()


    def getInfo(self):

        try:
            # Compute new vars
            addVars = AddinVariables(self.prj['detail']['addin_vars'], self.df, self.dictVarName, self.dictValLbl, self.productCode)
            isSuccess = addVars.addin()

            if not isSuccess[0]:
                return isSuccess
            # End Compute new vars

            self.df, self.dictVarName, self.dictValLbl = addVars.df, addVars.dictVarName, addVars.dictValLbl

            # Header
            for key, val in self.prj['detail']['topline_design']['header'].items():

                excl_hder = str(val['hidden_cats']).split('|') if len(val['hidden_cats']) > 0 else []
                excl_hder = list(map(int, excl_hder))

                # self.dictHeader[int(key)] = {
                self.dictHeader[str(key)] = {
                    'qre': val['name'],
                    'lbl': val['lbl'],
                    'excl': excl_hder,
                    'run_secs': val['run_secs'].split(',') if '|' not in val['run_secs'] else val['run_secs'].split('|'),
                }

            # Side Axis
            for key, val in self.prj['detail']['topline_design']['side'].items():
                # stt = int(key)
                stt = str(key)

                self.dictSide[stt] = {
                    'groupLbl': val['group_lbl'],
                    'qre': val['name'],
                    'qreLbl': val['lbl'],
                    'type': val['type'],
                    'atts': list(),
                    'MACats': list(),
                    'excl': list(),  # hidden cats
                    'isCount': val['is_count'],
                    'isCorr': val['is_corr'],
                    'isUA': val['is_ua'],
                    'qres': list(),
                }

                if val['type'] in ['MA']:
                    lst_ma_cats = str(val['ma_cats']).split('|') if len(val['ma_cats']) > 0 else []
                    lst_ma_cats = list(map(int, lst_ma_cats))
                    self.dictSide[stt]['MACats'] = lst_ma_cats


                excl = str(val['hidden_cats']).split('|') if len(val['hidden_cats']) > 0 else []
                excl = list(map(int, excl))
                self.dictSide[stt]['excl'] = excl

                for item in ['T2B', 'B2B', 'Mean']:
                    if val[str(item).lower()]:
                        self.dictSide[stt]['atts'].append(item)

                if self.dictSide[stt]['isUA']:
                    self.dictSide[stt]['qres'] = [self.dictSide[stt]['qre']]
                else:
                    self.dictSide[stt]['qres'] = [
                        f"{self.dictSide[stt]['qre']}_{self.productCode[0]}",
                        f"{self.dictSide[stt]['qre']}_{self.productCode[1]}"
                    ]

                if val['type'] == 'OL' and val['is_corr']:
                    self.dictCorrCols[val['name']] = val['lbl']


            # Create corr df
            lst_Qre_Corr = list()
            for prod in self.productCode:
                lst_Qre_Corr.append([f"{qre}_{prod}" for qre in self.dictCorrCols.keys()])

            df1 = self.df.loc[:, lst_Qre_Corr[0]]
            df1.columns = list(self.dictCorrCols.keys())

            df2 = self.df.loc[:, lst_Qre_Corr[1]]
            df2.columns = list(self.dictCorrCols.keys())

            self.dfCorr = pd.concat([df1, df2], axis=0, ignore_index=True)
            self.dfCorr.replace({None: np.nan}, inplace=True)

            return True, None

        except Exception:
            print(traceback.format_exc())
            return False, traceback.format_exc()


    def run_topline_format(self):

        print('----------------------------Topline formatting----------------------------')

        df = self.df
        dictTtest = dict()
        dictUA = dict()

        for hd_key, hd_val in self.dictHeader.items():

            print(f"Formatting - {hd_val['qre']}")

            if hd_val['qre'] == 'Total':

                strSubGroupName = hd_val['qre']
                dfFil = df.copy()

                dictTtest[strSubGroupName] = {'subGroupLbl': hd_val['lbl'], 'subCodeLbl': hd_val['lbl'], 'sideQres': {}}
                # dictTtest = self.add_Ttest_SideQres(dictTtest, strSubGroupName, dfFil)

                dictUA[strSubGroupName] = {'subGroupLbl': hd_val['lbl'], 'subCodeLbl': hd_val['lbl'], 'sideQres': {}}
                # dictUA = self.add_UA_SideQres(dictUA, strSubGroupName, dfFil)

                dictTtest, dictUA = self.add_Ttest_UA_SideQres(dictTtest, dictUA, strSubGroupName, dfFil)

            else:

                if self.export_section_name in hd_val['run_secs']:

                    for cat, catLbl in self.dictValLbl[hd_val['qre']].items():

                        if cat in hd_val['excl']:
                            continue

                        strSubGroupName = f"{hd_val['qre']}_{int(cat)}"
                        dfFil = df.loc[df[hd_val['qre']] == cat, :].copy()

                        dictTtest[strSubGroupName] = {'subGroupLbl': hd_val['lbl'], 'subCodeLbl': catLbl, 'sideQres': {}}
                        # dictTtest = self.add_Ttest_SideQres(dictTtest, strSubGroupName, dfFil)

                        dictUA[strSubGroupName] = {'subGroupLbl': hd_val['lbl'], 'subCodeLbl': catLbl, 'sideQres': {}}
                        # dictUA = self.add_UA_SideQres(dictUA, strSubGroupName, dfFil)

                        dictTtest, dictUA = self.add_Ttest_UA_SideQres(dictTtest, dictUA, strSubGroupName, dfFil)


        self.dictTtest = dictTtest
        self.dictUA = dictUA

        print('----------------------------Topline formatted-----------------------------')








    def add_Ttest_UA_SideQres(self, dictTtest: dict, dictUA: dict, strSubGroupName: str, dfFil: pd.DataFrame):

        for key, val in self.dictSide.items():

            if not val['isUA']:

                dictTtest[strSubGroupName]['sideQres'][key] = {
                    'qreLbl': val['qreLbl'],
                    'type': val['type'],
                    'isCount': val['isCount'],
                    'productCodes': [val['qres'][0].rsplit('_', 1)[-1], val['qres'][1].rsplit('_', 1)[-1]],
                    'groupLbl': val['groupLbl'],
                    'sigResult': self.run_ttest(dfFil, val),
                }


            else:

                dictUA[strSubGroupName]['sideQres'][key] = {
                    'qre': val['qre'],
                    'qreLbl': val['qreLbl'],
                    'type': val['type'],
                    'isCount': val['isCount'],
                    'groupLbl': val['groupLbl'],
                    'result': self.run_UA(dfFil, val),
                }


        return dictTtest, dictUA



    def add_Ttest_SideQres(self, dictTtest: dict, strSubGroupName: str, dfFil: pd.DataFrame):

        for key, val in self.dictSide.items():

            if not val['isUA']:

                dictTtest[strSubGroupName]['sideQres'][key] = {
                    'qreLbl': val['qreLbl'],
                    'type': val['type'],
                    'isCount': val['isCount'],
                    'productCodes': [val['qres'][0].rsplit('_', 1)[-1], val['qres'][1].rsplit('_', 1)[-1]],
                    'groupLbl': val['groupLbl'],
                    'sigResult': self.run_ttest(dfFil, val),
                }

        return dictTtest



    def run_ttest(self, df: pd.DataFrame, val: dict):

        df = df.copy()
        qres, atts, excl, qType, isCount = val['qres'], val['atts'], val['excl'], val['type'], val['isCount']

        # if 'C4a' in qres[0] and qType == "MA":
        #     a = 1

        dictSideQreFormat = dict()

        if qType == 'MA':
            maCats = val['MACats']

            maQre0 = qres[0].rsplit('_', 1)
            maQre1 = qres[1].rsplit('_', 1)

            qresName0 = [f'{maQre0[0]}_{cat}_{maQre0[1]}' for cat in maCats]
            qresName1 = [f'{maQre1[0]}_{cat}_{maQre1[1]}' for cat in maCats]

            df['Sum0'] = df[qresName0].sum(axis=1)
            df['Sum0'].replace({0: np.nan}, inplace=True)

            df['Sum1'] = df[qresName1].sum(axis=1)
            df['Sum1'].replace({0: np.nan}, inplace=True)


            total0 = int(df['Sum0'].count())
            total1 = int(df['Sum1'].count())

            # Base
            if 'Base' not in excl:
                dictSideQreFormat.update({'base': {
                    'catLbl': 'Base',
                    'val0': total0,
                    'sig0': 0,
                    'val1': total1,
                    'sig1': 0,
                }})

            # Categories MA
            for cat in maCats:
                qName0 = f'{maQre0[0]}_{cat}_{maQre0[1]}'
                qName1 = f'{maQre1[0]}_{cat}_{maQre1[1]}'

                for (key0, val_0), (key1, val_1) in zip(self.dictValLbl[qName0].items(), self.dictValLbl[qName1].items()):

                    if key0 in excl or key1 in excl:
                        continue

                    count0 = int(df.loc[df[qName0] == key0, qName0].count())
                    count1 = int(df.loc[df[qName1] == key1, qName1].count())

                    if isCount:
                        val0 = count0
                        val1 = count1
                    else:
                        val0 = count0 / total0 if total0 > 0 else 0
                        val1 = count1 / total1 if total1 > 0 else 0

                    # dictSideQreFormat.update({int(cat): {
                    dictSideQreFormat.update({str(cat): {
                        'catLbl': val_0,
                        'val0': val0,
                        'sig0': 0,
                        'val1': val1,
                        'sig1': 0,
                    }})

        elif qType == 'NUM':

            total0 = int(df[qres[0]].count())
            total1 = int(df[qres[1]].count())

            # Base
            if 'Base' not in excl:
                dictSideQreFormat.update({'base': {
                    'catLbl': 'Base',
                    'val0': int(total0),
                    'sig0': 0,
                    'val1': int(total1),
                    'sig1': 0,
                }})

            # a1 = qres[0]
            # a2 = qres[1]

            dictSideQreFormat.update({'mean': {
                'catLbl': 'Mean',
                'val0': df[qres[0]].mean() if total0 else 0,
                'sig0': 0,
                'val1': df[qres[1]].mean() if total1 else 0,
                'sig1': 0,
            }})

        else:

            total0 = df[qres[0]].count()
            total1 = df[qres[1]].count()

            # Base
            if 'Base' not in excl:
                dictSideQreFormat.update({'base': {
                    'catLbl': 'Base',
                    'val0': int(total0),
                    'sig0': 0,
                    'val1': int(total1),
                    'sig1': 0,
                }})

            # Scale / Categories
            for key, val in self.dictValLbl[qres[0]].items():

                if key in excl:
                    continue

                count0 = df.loc[df[qres[0]] == key, qres[0]].count()
                count1 = df.loc[df[qres[1]] == key, qres[1]].count()


                df['temp0'] = [1 if a == key else 0 for a in df[qres[0]]]
                df['temp1'] = [1 if a == key else 0 for a in df[qres[1]]]

                if isCount:
                    val0 = int(count0)
                    val1 = int(count1)
                else:
                    val0 = count0 / total0 if total0 > 0 else 0
                    val1 = count1 / total1 if total1 > 0 else 0

                # dictSideQreFormat.update({key: {
                dictSideQreFormat.update({str(key): {
                    'catLbl': val,
                    'val0': val0,
                    'sig0': 0,
                    'val1': val1,
                    'sig1': 0,
                }})

                if not isCount and total0 == total1 >= 30 and qType in ['OL', 'FC']:
                    # dictSideQreFormat[key] = self.run_ttest_rel(dictSideQreFormat[key], df['temp0'], df['temp1'])
                    dictSideQreFormat[str(key)] = self.run_ttest_rel(dictSideQreFormat[str(key)], df['temp0'], df['temp1'])


            lstCats = list(self.dictValLbl[qres[0]].keys())
            lstCats.sort()

            # T2B
            if 'T2B' in atts:

                if qType == 'JR' and len(lstCats) == 6:
                    minVal = lstCats[-3]
                    maxVal = lstCats[-2]
                else:
                    minVal = lstCats[-2]
                    maxVal = lstCats[-1]

                count0 = df.loc[(df[qres[0]] >= minVal) & (df[qres[0]] <= maxVal), qres[0]].count()
                count1 = df.loc[(df[qres[1]] >= minVal) & (df[qres[1]] <= maxVal), qres[1]].count()

                df['temp0'] = [np.nan if pd.isna(a) else (1 if maxVal >= a >= minVal else 0) for a in df[qres[0]]]
                df['temp1'] = [np.nan if pd.isna(a) else (1 if maxVal >= a >= minVal else 0) for a in df[qres[1]]]

                dictSideQreFormat.update({'t2b': {
                    'catLbl': 'T2B',
                    'val0': count0 / total0 if total0 > 0 else 0,
                    'sig0': 0,
                    'val1': count1 / total1 if total1 > 0 else 0,
                    'sig1': 0,
                }})

                if not isCount and total0 == total1 >= 30 and qType == 'OL':
                    dictSideQreFormat['t2b'] = self.run_ttest_rel(dictSideQreFormat['t2b'], df['temp0'], df['temp1'])

            # B2B
            if 'B2B' in atts:
                count0 = df.loc[df[qres[0]] <= lstCats[1], qres[0]].count()
                count1 = df.loc[df[qres[1]] <= lstCats[1], qres[1]].count()

                df['temp0'] = [np.nan if pd.isna(a) else (1 if a <= lstCats[1] else 0) for a in df[qres[0]]]
                df['temp1'] = [np.nan if pd.isna(a) else (1 if a <= lstCats[1] else 0) for a in df[qres[1]]]

                dictSideQreFormat.update({'b2b': {
                    'catLbl': 'B2B',
                    'val0': count0 / total0 if total0 > 0 else 0,
                    'sig0': 0,
                    'val1': count1 / total1 if total1 > 0 else 0,
                    'sig1': 0,
                }})

                if not isCount and total0 == total1 >= 30 and qType == 'OL':
                    dictSideQreFormat['b2b'] = self.run_ttest_rel(dictSideQreFormat['b2b'], df['temp0'], df['temp1'])

            # Mean
            if 'Mean' in atts:

                if qType == 'JR':

                    if self.isJR3Factors:
                        df[qres[0]].replace({4: 2, 5: 1}, inplace=True)
                        df[qres[1]].replace({4: 2, 5: 1}, inplace=True)

                    if len(lstCats) == 6:
                        df[qres[0]].replace({6: np.nan}, inplace=True)
                        df[qres[1]].replace({6: np.nan}, inplace=True)

                dictSideQreFormat.update({'mean': {
                    'catLbl': 'Mean',
                    'val0': df[qres[0]].mean() if df[qres[0]].count() else 0,
                    'sig0': 0,
                    'val1': df[qres[1]].mean() if df[qres[1]].count() else 0,
                    'sig1': 0,
                }})

                if total0 == total1 >= 30 and qType == 'OL':
                    dictSideQreFormat['mean'] = self.run_ttest_rel(dictSideQreFormat['mean'], df[qres[0]], df[qres[1]])


            # Std
            if qType in ['OL', 'JR']:

                if qType == 'JR':
                    df[qres[0]].replace({4: 2, 5: 1}, inplace=True)
                    df[qres[1]].replace({4: 2, 5: 1}, inplace=True)

                    if len(lstCats) == 6:
                        df[qres[0]].replace({6: np.nan}, inplace=True)
                        df[qres[1]].replace({6: np.nan}, inplace=True)

                dictSideQreFormat.update({'std': {
                    'catLbl': 'Standard Deviation',
                    'val0': df[qres[0]].std() if df[qres[0]].count() else 0,
                    'sig0': 0,
                    'val1': df[qres[1]].std() if df[qres[1]].count() else 0,
                    'sig1': 0,
                }})


        return dictSideQreFormat



    def run_ttest_rel(self, dictToSig: dict, arr1, arr2):

        if len(arr1) >= 30 <= len(arr2):
            sigResult = stats.ttest_rel(arr1, arr2)

            if sigResult.statistic > 0:
                dictToSig['sig0'] = self.convertSigValue(sigResult.pvalue)
            else:
                dictToSig['sig1'] = self.convertSigValue(sigResult.pvalue)

        return dictToSig


    @staticmethod
    def convertSigValue(pval):
        if pval <= .05:
            return 95
        elif pval <= .1:
            return 90
        elif pval <= .2:
            return 80
        else:
            return 0



    def add_UA_SideQres(self, dictUA: dict, strSubGroupName: str, dfFil: pd.DataFrame):

        for key, val in self.dictSide.items():

            if val['isUA']:

                dictUA[strSubGroupName]['sideQres'][key] = {
                    'qre': val['qre'],
                    'qreLbl': val['qreLbl'],
                    'type': val['type'],
                    'isCount': val['isCount'],
                    'groupLbl': val['groupLbl'],
                    'result': self.run_UA(dfFil, val),
                }

        return dictUA



    def run_UA(self, df: pd.DataFrame, val: dict):

        df = df.copy()
        qres, atts, excl, qType, isCount, maCats = val['qres'], val['atts'], val['excl'], val['type'], val['isCount'], val['MACats']

        dictSideQreFormat = dict()


        if qType == 'MA':

            qresName = [f'{qres[0]}_{cat}' for cat in maCats]
            df['Sum'] = df[qresName].sum(axis=1)
            df['Sum'].replace({0: np.nan}, inplace=True)

            total0 = int(df['Sum'].count())

            # Base
            if 'Base' not in excl:
                dictSideQreFormat.update({'base': {
                    'catLbl': 'Base',
                    'val0': total0,
                }})

            # Categories MA
            for cat in maCats:
                qName = f'{qres[0]}_{cat}'

                for key, val in self.dictValLbl[qName].items():

                    if key in excl:
                        continue

                    count0 = int(df.loc[df[qName] == key, qName].count())

                    if isCount:
                        val0 = count0
                    else:
                        val0 = count0 / total0 if total0 > 0 else 0

                    # dictSideQreFormat.update({int(cat): {
                    dictSideQreFormat.update({str(cat): {
                        'catLbl': val,
                        'val0': val0
                    }})

        elif qType == 'NUM':

            total0 = int(df[qres[0]].count())

            # Base
            if 'Base' not in excl:
                dictSideQreFormat.update({'base': {
                    'catLbl': 'Base',
                    'val0': total0,
                }})

            # Mean
            dictSideQreFormat.update({'mean': {
                'catLbl': 'Mean',
                'val0': df[qres[0]].mean() if df[qres[0]].count() else 0,
            }})

        else:

            total0 = int(df[qres[0]].count())

            # Base
            if 'Base' not in excl:
                dictSideQreFormat.update({'base': {
                    'catLbl': 'Base',
                    'val0': total0,
                }})

            # Scale / Categories
            for key, val in self.dictValLbl[qres[0]].items():

                if key in excl:
                    continue

                count0 = int(df.loc[df[qres[0]] == key, qres[0]].count())

                if isCount:
                    val0 = count0
                else:
                    val0 = count0 / total0 if total0 > 0 else 0

                # dictSideQreFormat.update({int(key): {
                dictSideQreFormat.update({str(key): {
                    'catLbl': val,
                    'val0': val0,
                }})

            lstCats = list(self.dictValLbl[qres[0]].keys())
            lstCats.sort()

            # T2B
            if 'T2B' in atts:
                count0 = int(df.loc[df[qres[0]] >= lstCats[-2], qres[0]].count())

                dictSideQreFormat.update({'t2b': {
                    'catLbl': 'T2B',
                    'val0': count0 / total0 if total0 > 0 else 0,
                }})


            # B2B
            if 'B2B' in atts:
                count0 = int(df.loc[df[qres[0]] <= lstCats[1], qres[0]].count())

                dictSideQreFormat.update({'b2b': {
                    'catLbl': 'B2B',
                    'val0': count0 / total0 if total0 > 0 else 0,
                }})


            # Mean
            if 'Mean' in atts:

                if qType == 'JR':

                    if self.isJR3Factors:
                        df[qres[0]].replace({4: 2, 5: 1}, inplace=True)

                dictSideQreFormat.update({'mean': {
                    'catLbl': 'Mean',
                    'val0': df[qres[0]].mean() if df[qres[0]].count() else 0,
                }})



            # Std
            if qType in ['OL', 'JR']:

                if qType == 'JR':
                    df[qres[0]].replace({4: 2, 5: 1}, inplace=True)

                dictSideQreFormat.update({'std': {
                    'catLbl': 'Standard Deviation',
                    'val0': df[qres[0]].std() if df[qres[0]].count() else 0,
                }})


        return dictSideQreFormat



    @staticmethod
    def fillSigColor(sigVal):

        cellFont = Font(color='00000000')

        if sigVal == 80:
            cellFont = Font(color='0000FF00')  # green
        elif sigVal == 90:
            cellFont = Font(color='000000FF')  # blue
        elif sigVal == 95:
            cellFont = Font(color='00FF0000')  # red

        return cellFont


    def toSignificant(self, wb: openpyxl.Workbook, dictTtest: dict):

        print('Export Significant')

        ws = wb['Sheet']
        ws.title = 'significant'

        thin = Side(border_style='thin', color='000000')
        medium = Side(border_style='medium', color='000000')

        headerStartCol = 3
        sideStartRow = -1

        for key, val in dictTtest.items():

            subGroupCellQre = ws.cell(row=1, column=headerStartCol)
            subGroupCellQre.value = val['subGroupLbl']
            subGroupCellQre.font = Font(bold=True)
            subGroupCellQre.fill = PatternFill('solid', fgColor='FDE9D9')
            subGroupCellQre.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=1, column=headerStartCol+1).value = val['subGroupLbl']

            ws.merge_cells(start_row=2, start_column=headerStartCol, end_row=2, end_column=headerStartCol + 1)
            subGroupCellVal = ws.cell(row=2, column=headerStartCol)
            subGroupCellVal.value = val['subCodeLbl']
            subGroupCellVal.font = Font(bold=True)
            subGroupCellVal.fill = PatternFill('solid', fgColor='FFC000')
            subGroupCellVal.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=2, column=headerStartCol).border = Border(left=medium, top=thin)
            ws.cell(row=2, column=headerStartCol + 1).border = Border(right=medium, top=thin)


            sideStartRow = 4

            for key1, val1 in val['sideQres'].items():

                if key == 'Total':
                    ws.merge_cells(start_row=sideStartRow, start_column=1, end_row=sideStartRow + len(val1['sigResult']) - 2, end_column=1)
                    cellQreLbl = ws.cell(row=sideStartRow, column=1)
                    cellQreLbl.value = val1["qreLbl"]
                    cellQreLbl.font = Font(bold=True)
                    cellQreLbl.fill = PatternFill('solid', fgColor='DDEBF7')
                    cellQreLbl.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    cellQreLbl.border = Border(top=medium)


                cellProd1 = ws.cell(row=3, column=headerStartCol)
                cellProd1.value = int(val1['productCodes'][0])
                cellProd1.font = Font(bold=True, color='FFFF00')
                cellProd1.fill = PatternFill('solid', fgColor='002060')
                cellProd1.alignment = Alignment(horizontal='center', vertical='center')
                cellProd1.border = Border(left=medium, right=thin, top=thin)

                cellProd2 = ws.cell(row=3, column=headerStartCol + 1)
                cellProd2.value = int(val1['productCodes'][1])
                cellProd2.font = Font(bold=True, color='FFFF00')
                cellProd2.fill = PatternFill('solid', fgColor='002060')
                cellProd2.alignment = Alignment(horizontal='center', vertical='center')
                cellProd2.border = Border(right=medium, top=thin)

                for key2, val2 in val1['sigResult'].items():

                    if key2 not in ['std']:

                        cellValLbl = ws.cell(row=sideStartRow, column=2)
                        cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                        cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                        cellValLbl.value = val2['catLbl']
                        cellVal0.value = val2['val0']
                        cellVal1.value = val2['val1']

                        cellValLbl.border = Border(left=thin, right=medium)
                        cellVal0.border = Border(right=thin)
                        cellVal1.border = Border(right=medium)

                        if key2 == 'base':

                            cellValLbl.font, cellValLbl.fill = Font(bold=True), PatternFill('solid', fgColor='C6E0B4')
                            cellVal0.font, cellVal0.fill = Font(bold=True), PatternFill('solid', fgColor='C6E0B4')
                            cellVal1.font, cellVal1.fill = Font(bold=True), PatternFill('solid', fgColor='C6E0B4')

                            cellValLbl.border = Border(left=thin, top=medium, right=medium, bottom=thin)
                            cellVal0.border = Border(left=thin, top=medium, right=thin, bottom=thin)
                            cellVal1.border = Border(left=thin, top=medium, right=medium, bottom=thin)

                        else:
                            if key2 in ['mean', 'std']:
                                cellVal0.number_format = '0.00'
                                cellVal1.number_format = '0.00'
                            else:
                                if not val1['isCount']:
                                    if self.isDisplayPctSign:
                                        cellVal0.number_format = '0%'
                                        cellVal1.number_format = '0%'
                                    else:
                                        cellVal0.number_format = '0'
                                        cellVal1.number_format = '0'

                                        cellVal0.value = cellVal0.value * 100
                                        cellVal1.value = cellVal1.value * 100

                            cellVal0.font = self.fillSigColor(val2['sig0'])
                            cellVal1.font = self.fillSigColor(val2['sig1'])


                        sideStartRow += 1

            headerStartCol += 2

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.freeze_panes = 'C4'


        icol1 = 3
        while ws.cell(row=1, column=icol1).value is not None:

            icol2 = icol1 + 1
            while ws.cell(row=1, column=icol1).value == ws.cell(row=1, column=icol2).value:
                icol2 += 1

            ws.merge_cells(start_row=1, start_column=icol1, end_row=1, end_column=icol2 - 1)

            ws.cell(row=1, column=icol1).border = Border(left=medium, top=medium, right=medium, bottom=thin)
            ws.cell(row=1, column=icol2 - 1).border = Border(left=medium, top=medium, right=medium, bottom=thin)

            icol1 = icol2

        for icol in range(1, headerStartCol):
            ws.cell(sideStartRow, icol).border = Border(top=medium)



    def toSummary(self, wb: openpyxl.workbook, dictTtest: dict, lstQreType=None):

        if lstQreType is None:
            print('Export Topline Summary')
            lstQreType = ['OL', 'JR', 'FC']
            ws = wb.create_sheet('1. Summary')
        else:
            print('Export Topline Handcount')
            ws = wb.create_sheet('0. Handcount')

        ws.sheet_properties.tabColor = '002060'

        thin = Side(border_style='thin', color='000000')
        medium = Side(border_style='medium', color='000000')
        dot = Side(border_style='dotted', color='000000')

        cellTitle = ws.cell(row=1, column=1)
        cellTitle.value = self.toplineTitle
        cellTitle.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cellTitle.font = Font(bold=True, color='0070C0', size=24)
        ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=5)

        headerStartCol = 6
        sideStartRow = -1

        for key, val in dictTtest.items():

            subGroupCellQre = ws.cell(row=1, column=headerStartCol)
            subGroupCellQre.value = val['subGroupLbl']
            subGroupCellQre.font = Font(bold=True)
            subGroupCellQre.fill = PatternFill('solid', fgColor='FDE9D9')
            subGroupCellQre.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=1, column=headerStartCol + 1).value = val['subGroupLbl']

            ws.merge_cells(start_row=2, start_column=headerStartCol, end_row=2, end_column=headerStartCol + 1)
            subGroupCellVal = ws.cell(row=2, column=headerStartCol)
            subGroupCellVal.value = val['subCodeLbl']
            subGroupCellVal.font = Font(bold=True)
            subGroupCellVal.fill = PatternFill('solid', fgColor='FFC000')
            subGroupCellVal.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=2, column=headerStartCol).border = Border(left=medium, top=thin)
            ws.cell(row=2, column=headerStartCol + 1).border = Border(right=medium, top=thin)

            sideStartRow = 5

            # mean OL + JR & others Qre
            for key1, val1 in val['sideQres'].items():

                if val1['type'] in lstQreType:  # ['OL', 'JR', 'FC']

                    if key == 'Total':
                        ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                        ws.cell(row=sideStartRow, column=5).value = val1["qreLbl"].split('. ')[1]

                        if val1['type'] == 'OL':
                            ws.cell(row=sideStartRow, column=4).value = '5pt'
                        else:
                            ws.cell(row=sideStartRow, column=4).value = val1['type']

                        if val1['type'] == 'JR':
                            ws.cell(row=sideStartRow, column=3).fill = PatternFill('solid', fgColor='FCE4D6')
                            ws.cell(row=sideStartRow, column=4).fill = PatternFill('solid', fgColor='FCE4D6')
                            ws.cell(row=sideStartRow, column=5).fill = PatternFill('solid', fgColor='FCE4D6')

                        ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                        ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                        ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)


                    cellProd1 = ws.cell(row=3, column=headerStartCol)
                    cellProd1.value = int(val1['productCodes'][0])
                    cellProd1.font = Font(bold=True, color='FFFF00')
                    cellProd1.fill = PatternFill('solid', fgColor='002060')
                    cellProd1.alignment = Alignment(horizontal='center', vertical='center')
                    cellProd1.border = Border(left=medium, right=thin, top=thin)

                    cellProd2 = ws.cell(row=3, column=headerStartCol + 1)
                    cellProd2.value = int(val1['productCodes'][1])
                    cellProd2.font = Font(bold=True, color='FFFF00')
                    cellProd2.fill = PatternFill('solid', fgColor='002060')
                    cellProd2.alignment = Alignment(horizontal='center', vertical='center')
                    cellProd2.border = Border(right=medium, top=thin)

                    for key2, val2 in val1['sigResult'].items():

                        if key2 == 'base':
                            cellValLbl = ws.cell(row=4, column=1)
                            cellVal0 = ws.cell(row=4, column=headerStartCol)
                            cellVal1 = ws.cell(row=4, column=headerStartCol + 1)

                            if cellValLbl.value is None:
                                ws.cell(4, 1).border = Border(top=medium, bottom=medium)
                                ws.cell(4, 2).border = Border(top=medium, bottom=medium)
                                ws.cell(4, 3).border = Border(top=medium, bottom=medium)
                                ws.cell(4, 4).border = Border(top=medium, bottom=medium)
                                ws.cell(4, 5).border = Border(top=medium, bottom=medium, right=medium)

                                ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)
                                cellValLbl.value = 'N='  # val2['catLbl']
                                cellValLbl.font, cellValLbl.fill = Font(bold=True, color='FF0000'), PatternFill('solid', fgColor='C6E0B4')
                                cellValLbl.border = Border(left=medium, top=medium, right=medium, bottom=thin)
                                cellValLbl.alignment = Alignment(horizontal='right')


                            if cellVal0.value is None:
                                cellVal0.value = val2['val0']
                                cellVal0.font, cellVal0.fill = Font(bold=True, color='FF0000'), PatternFill('solid', fgColor='C6E0B4')
                                cellVal0.border = Border(left=thin, top=medium, right=thin, bottom=medium)

                            if cellVal1.value is None:
                                cellVal1.value = val2['val1']
                                cellVal1.font, cellVal1.fill = Font(bold=True, color='FF0000'), PatternFill('solid', fgColor='C6E0B4')
                                cellVal1.border = Border(left=thin, top=medium, right=medium, bottom=medium)

                        else:
                            if key2 == 'mean' or (key2 != 'mean' and val1['type'] not in ['OL', 'JR']):
                                cellValLbl = ws.cell(row=sideStartRow, column=2)
                                cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                                cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)


                                if val1['type'] in ['FC']:
                                    ws.cell(row=sideStartRow, column=1).value = 'CÁC CÂU SO SÁNH'
                                elif val1['type'] in ['SA', 'MA']:
                                    ws.cell(row=sideStartRow, column=1).value = 'CÁC CÂU KHÁC'
                                    ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                                    ws.cell(row=sideStartRow, column=4).value = val1["qreLbl"].split('. ')[1]
                                    ws.cell(row=sideStartRow, column=5).value = val2['catLbl']

                                    ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                                    ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                                    ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)
                                else:
                                    ws.cell(row=sideStartRow, column=1).value = 'CÁC CÂU OL(Mean) & JR (Mean)'  # val2['catLbl']

                                cellValLbl.value = val1['groupLbl']


                                cellVal0.value = val2['val0']
                                cellVal1.value = val2['val1']

                                cellValLbl.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                                cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                                cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                                if key2 == 'mean':
                                    cellVal0.number_format = '0.00'
                                    cellVal1.number_format = '0.00'
                                else:
                                    if not val1['isCount']:
                                        if self.isDisplayPctSign:
                                            cellVal0.number_format = '0%'
                                            cellVal1.number_format = '0%'
                                        else:
                                            cellVal0.number_format = '0'
                                            cellVal1.number_format = '0'

                                            cellVal0.value = cellVal0.value * 100
                                            cellVal1.value = cellVal1.value * 100

                                if val1['type'] in ['OL', 'FC']:
                                    cellVal0.font = self.fillSigColor(val2['sig0'])
                                    cellVal1.font = self.fillSigColor(val2['sig1'])

                                sideStartRow += 1

            # mean JR
            for key1, val1 in val['sideQres'].items():

                if val1['type'] == 'JR' and val1['type'] in lstQreType:

                    if key == 'Total':
                        ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                        ws.cell(row=sideStartRow, column=5).value = val1["qreLbl"].split('. ')[1]
                        ws.cell(row=sideStartRow, column=4).value = val1['type']

                        ws.cell(row=sideStartRow, column=3).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=4).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=5).fill = PatternFill('solid', fgColor='FCE4D6')

                        ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                        ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                        ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)


                    for key2, val2 in val1['sigResult'].items():

                        if key2 == 'mean':
                            cellValLbl = ws.cell(row=sideStartRow, column=2)
                            cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                            cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                            ws.cell(row=sideStartRow, column=1).value = 'CÁC CÂU JR (Mean) - Sig vs 3'  # val2['catLbl']

                            cellValLbl.value = val1['groupLbl']

                            cellVal0.value = val2['val0']
                            cellVal1.value = val2['val1']

                            cellValLbl.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                            cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                            cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                            cellVal0.number_format = '0.00'
                            cellVal1.number_format = '0.00'

                            sideStartRow += 1

            # % JR - Code 3
            for key1, val1 in val['sideQres'].items():

                if val1['type'] == 'JR' and val1['type'] in lstQreType:

                    if key == 'Total':
                        ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                        ws.cell(row=sideStartRow, column=5).value = val1["qreLbl"].split('. ')[1]
                        ws.cell(row=sideStartRow, column=4).value = val1['type']

                        ws.cell(row=sideStartRow, column=3).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=4).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=5).fill = PatternFill('solid', fgColor='FCE4D6')

                        ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                        ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                        ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)

                    for key2, val2 in val1['sigResult'].items():

                        if key2 == '3':
                            cellValLbl = ws.cell(row=sideStartRow, column=2)
                            cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                            cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                            ws.cell(row=sideStartRow, column=1).value = 'CÁC CÂU JR (%JR)'  # val2['catLbl']

                            cellValLbl.value = val1['groupLbl']

                            cellVal0.value = val2['val0']
                            cellVal1.value = val2['val1']

                            cellValLbl.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                            cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                            cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                            if self.isDisplayPctSign:
                                cellVal0.number_format = '0%'
                                cellVal1.number_format = '0%'
                            else:
                                cellVal0.number_format = '0'
                                cellVal1.number_format = '0'

                                cellVal0.value = cellVal0.value * 100
                                cellVal1.value = cellVal1.value * 100

                            sideStartRow += 1

            # % T2B - OL
            for key1, val1 in val['sideQres'].items():

                if val1['type'] == 'OL':

                    if key == 'Total':
                        ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                        ws.cell(row=sideStartRow, column=5).value = val1["qreLbl"].split('. ')[1]
                        ws.cell(row=sideStartRow, column=4).value = '5pt'  # val1['type']

                        ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                        ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                        ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)

                    for key2, val2 in val1['sigResult'].items():

                        if key2 == 't2b':
                            cellValLbl = ws.cell(row=sideStartRow, column=2)
                            cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                            cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                            ws.cell(row=sideStartRow, column=1).value = 'Mức độ thích (%T2B)'  # val2['catLbl']

                            cellValLbl.value = val1['groupLbl']

                            cellVal0.value = val2['val0']
                            cellVal1.value = val2['val1']

                            cellValLbl.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                            cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                            cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                            if self.isDisplayPctSign:
                                cellVal0.number_format = '0%'
                                cellVal1.number_format = '0%'
                            else:
                                cellVal0.number_format = '0'
                                cellVal1.number_format = '0'

                                cellVal0.value = cellVal0.value * 100
                                cellVal1.value = cellVal1.value * 100

                            cellVal0.font = self.fillSigColor(val2['sig0'])
                            cellVal1.font = self.fillSigColor(val2['sig1'])

                            sideStartRow += 1

            # % T2B - JR
            for key1, val1 in val['sideQres'].items():

                if val1['type'] == 'JR' and val1['type'] in lstQreType:

                    if key == 'Total':
                        ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                        ws.cell(row=sideStartRow, column=5).value = val1["qreLbl"].split('. ')[1]
                        ws.cell(row=sideStartRow, column=4).value = val1['type']

                        ws.cell(row=sideStartRow, column=3).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=4).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=5).fill = PatternFill('solid', fgColor='FCE4D6')

                        ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                        ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                        ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)

                    for key2, val2 in val1['sigResult'].items():

                        if key2 == 't2b':
                            cellValLbl = ws.cell(row=sideStartRow, column=2)
                            cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                            cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                            ws.cell(row=sideStartRow, column=1).value = 'CÁC CÂU JR (%T2B)'  # val2['catLbl']

                            cellValLbl.value = val1['groupLbl']

                            cellVal0.value = val2['val0']
                            cellVal1.value = val2['val1']

                            cellValLbl.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                            cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                            cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                            if self.isDisplayPctSign:
                                cellVal0.number_format = '0%'
                                cellVal1.number_format = '0%'
                            else:
                                cellVal0.number_format = '0'
                                cellVal1.number_format = '0'

                                cellVal0.value = cellVal0.value * 100
                                cellVal1.value = cellVal1.value * 100

                            sideStartRow += 1

            # % B2B - OL
            for key1, val1 in val['sideQres'].items():

                if val1['type'] == 'OL':

                    if key == 'Total':
                        ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                        ws.cell(row=sideStartRow, column=5).value = val1["qreLbl"].split('. ')[1]
                        ws.cell(row=sideStartRow, column=4).value = '5pt'  # val1['type']

                        ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                        ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                        ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)

                    for key2, val2 in val1['sigResult'].items():

                        if key2 == 'b2b':
                            cellValLbl = ws.cell(row=sideStartRow, column=2)
                            cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                            cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                            ws.cell(row=sideStartRow, column=1).value = 'Mức độ thích (%B2B)'  # val2['catLbl']

                            cellValLbl.value = val1['groupLbl']

                            cellVal0.value = val2['val0']
                            cellVal1.value = val2['val1']

                            cellValLbl.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                            cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                            cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                            if self.isDisplayPctSign:
                                cellVal0.number_format = '0%'
                                cellVal1.number_format = '0%'
                            else:
                                cellVal0.number_format = '0'
                                cellVal1.number_format = '0'

                                cellVal0.value = cellVal0.value * 100
                                cellVal1.value = cellVal1.value * 100

                            cellVal0.font = self.fillSigColor(val2['sig0'])
                            cellVal1.font = self.fillSigColor(val2['sig1'])

                            sideStartRow += 1

            # % B2B - JR
            for key1, val1 in val['sideQres'].items():

                if val1['type'] == 'JR' and val1['type'] in lstQreType:

                    if key == 'Total':
                        ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                        ws.cell(row=sideStartRow, column=5).value = val1["qreLbl"].split('. ')[1]
                        ws.cell(row=sideStartRow, column=4).value = val1['type']

                        ws.cell(row=sideStartRow, column=3).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=4).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=5).fill = PatternFill('solid', fgColor='FCE4D6')

                        ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                        ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                        ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)

                    for key2, val2 in val1['sigResult'].items():

                        if key2 == 'b2b':
                            cellValLbl = ws.cell(row=sideStartRow, column=2)
                            cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                            cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                            ws.cell(row=sideStartRow, column=1).value = 'CÁC CÂU JR (%B2B)'  # val2['catLbl']

                            cellValLbl.value = val1['groupLbl']

                            cellVal0.value = val2['val0']
                            cellVal1.value = val2['val1']

                            cellValLbl.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                            cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                            cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                            if self.isDisplayPctSign:
                                cellVal0.number_format = '0%'
                                cellVal1.number_format = '0%'
                            else:
                                cellVal0.number_format = '0'
                                cellVal1.number_format = '0'

                                cellVal0.value = cellVal0.value * 100
                                cellVal1.value = cellVal1.value * 100

                            sideStartRow += 1


            headerStartCol += 2

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 50
        ws.freeze_panes = 'F5'

        icol1 = 6
        while ws.cell(row=1, column=icol1).value is not None:

            icol2 = icol1 + 1
            while ws.cell(row=1, column=icol1).value == ws.cell(row=1, column=icol2).value:
                icol2 += 1

            ws.merge_cells(start_row=1, start_column=icol1, end_row=1, end_column=icol2 - 1)

            ws.cell(row=1, column=icol1).border = Border(left=medium, top=medium, right=medium, bottom=thin)
            ws.cell(row=1, column=icol2 - 1).border = Border(left=medium, top=medium, right=medium, bottom=thin)

            icol1 = icol2


        icol = 3
        irow1 = 5
        while ws.cell(row=irow1, column=icol).value is not None:

            irow2 = irow1 + 1
            while ws.cell(row=irow1, column=icol).value == ws.cell(row=irow2, column=icol).value:
                irow2 += 1

            if irow1 != irow2 - 1:
                ws.merge_cells(start_row=irow1, start_column=icol, end_row=irow2 - 1, end_column=icol)
                ws.cell(row=irow1, column=icol).alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

                ws.merge_cells(start_row=irow1, start_column=icol + 1, end_row=irow2 - 1, end_column=icol + 1)
                ws.cell(row=irow1, column=icol + 1).alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

                ws.cell(row=irow1, column=icol).border = Border(left=medium, top=thin, right=thin, bottom=thin)

                ws.cell(row=irow2 - 1, column=3).border = Border(left=medium, right=thin, bottom=thin)
                ws.cell(row=irow2 - 1, column=4).border = Border(left=thin, right=thin, bottom=thin)
                ws.cell(row=irow2 - 1, column=5).border = Border(left=thin, right=medium, bottom=thin)

                for i in range(6, headerStartCol):

                    if i % 2 == 0:
                        ws.cell(row=irow2 - 1, column=i).border = Border(left=medium, right=thin, bottom=thin, top=dot)
                    else:
                        ws.cell(row=irow2 - 1, column=i).border = Border(left=thin, right=medium, bottom=thin, top=dot)

            irow1 = irow2


        icol = 2
        irow1 = 5
        while ws.cell(row=irow1, column=icol).value is not None:

            irow2 = irow1 + 1
            while ws.cell(row=irow1, column=icol).value == ws.cell(row=irow2, column=icol).value:
                irow2 += 1

            ws.merge_cells(start_row=irow1, start_column=icol, end_row=irow2 - 1, end_column=icol)
            ws.cell(row=irow1, column=icol).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            ws.cell(row=irow1, column=icol).font = Font(bold=True)

            if icol == 1:
                ws.cell(row=irow1, column=icol).fill = PatternFill('solid', fgColor='FFE699')

            else:
                ws.cell(row=irow1, column=icol).fill = PatternFill('solid', fgColor='DDEBF7')

            ws.cell(row=irow1, column=icol).border = Border(left=medium, top=thin, right=medium, bottom=thin)

            ws.cell(row=irow2 - 1, column=1).border = Border(left=medium, bottom=thin)
            ws.cell(row=irow2 - 1, column=2).border = Border(left=medium, bottom=thin)
            ws.cell(row=irow2 - 1, column=3).border = Border(left=medium, bottom=thin)
            ws.cell(row=irow2 - 1, column=4).border = Border(left=thin, bottom=thin)
            ws.cell(row=irow2 - 1, column=5).border = Border(left=thin, bottom=thin, right=medium)

            for i in range(6, headerStartCol):

                if i % 2 == 0:
                    ws.cell(row=irow2 - 1, column=i).border = Border(bottom=thin, right=thin)
                else:
                    ws.cell(row=irow2 - 1, column=i).border = Border(bottom=thin, right=medium)


            irow1 = irow2


        icol = 1
        irow1 = 5
        while ws.cell(row=irow1, column=icol).value is not None:

            irow2 = irow1 + 1
            while ws.cell(row=irow1, column=icol).value == ws.cell(row=irow2, column=icol).value:
                irow2 += 1

            ws.merge_cells(start_row=irow1, start_column=icol, end_row=irow2 - 1, end_column=icol)
            ws.cell(row=irow1, column=icol).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            ws.cell(row=irow1, column=icol).font = Font(bold=True)

            if icol == 1:
                ws.cell(row=irow1, column=icol).fill = PatternFill('solid', fgColor='FFE699')

            else:
                ws.cell(row=irow1, column=icol).fill = PatternFill('solid', fgColor='DDEBF7')

            ws.cell(row=irow1, column=icol).border = Border(left=medium, top=medium, right=medium, bottom=thin)

            ws.cell(row=irow2 - 1, column=1).border = Border(left=medium, bottom=medium)
            ws.cell(row=irow2 - 1, column=2).border = Border(left=medium, bottom=medium)
            ws.cell(row=irow2 - 1, column=3).border = Border(left=medium, bottom=medium)
            ws.cell(row=irow2 - 1, column=4).border = Border(left=thin, bottom=medium)
            ws.cell(row=irow2 - 1, column=5).border = Border(left=thin, bottom=medium, right=medium)

            for i in range(6, headerStartCol):

                if i % 2 == 0:
                    ws.cell(row=irow2 - 1, column=i).border = Border(bottom=medium, right=thin)
                else:
                    ws.cell(row=irow2 - 1, column=i).border = Border(bottom=medium, right=medium)


            irow1 = irow2


        ws.cell(row=sideStartRow + 1, column=4).value = 'Lưu ý:'
        ws.cell(row=sideStartRow + 1, column=4).font = Font(bold=True, color='00FF0000')

        ws.cell(row=sideStartRow + 2, column=4).value = 'Chỉ chạy Sig. Test cho các câu OL, các câu JAR không chạy Sig. Test'

        ws.cell(row=sideStartRow + 3, column=4).value = 'Ký hiệu Sig. Test'
        ws.cell(row=sideStartRow + 3, column=4).font = Font(bold=True)

        ws.cell(row=sideStartRow + 4, column=4).value = 'Đỏ: Sig. ở 95% trở lên'
        ws.cell(row=sideStartRow + 4, column=4).font = Font(bold=True, color='00FF0000')

        ws.cell(row=sideStartRow + 5, column=4).value = 'Xanh dương: Sig. ở 90% đến dưới 95%'
        ws.cell(row=sideStartRow + 5, column=4).font = Font(bold=True, color='000000FF')

        ws.cell(row=sideStartRow + 6, column=4).value = 'Xanh lá: Sig. ở 80% đến dưới 90%'
        ws.cell(row=sideStartRow + 6, column=4).font = Font(bold=True, color='0000FF00')



    def toTabulation(self, wb: openpyxl.workbook, dictTtest: dict):

        print('Export Topline Tabulation')

        ws = wb.create_sheet('2. Tabulation')
        ws.sheet_properties.tabColor = '002060'

        thin = Side(border_style='thin', color='000000')
        medium = Side(border_style='medium', color='000000')
        dot = Side(border_style='dotted', color='000000')

        cellTitle = ws.cell(row=1, column=1)
        cellTitle.value = self.toplineTitle
        cellTitle.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cellTitle.font = Font(bold=True, color='0070C0', size=24)
        ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=2)

        headerStartCol = 3
        sideStartRow = -1

        for key, val in dictTtest.items():

            subGroupCellQre = ws.cell(row=1, column=headerStartCol)
            subGroupCellQre.value = val['subGroupLbl']
            subGroupCellQre.font = Font(bold=True)
            subGroupCellQre.fill = PatternFill('solid', fgColor='FDE9D9')
            subGroupCellQre.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=1, column=headerStartCol + 1).value = val['subGroupLbl']

            ws.merge_cells(start_row=2, start_column=headerStartCol, end_row=2, end_column=headerStartCol + 1)
            subGroupCellVal = ws.cell(row=2, column=headerStartCol)
            subGroupCellVal.value = val['subCodeLbl']
            subGroupCellVal.font = Font(bold=True)
            subGroupCellVal.fill = PatternFill('solid', fgColor='FFC000')
            subGroupCellVal.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=2, column=headerStartCol).border = Border(left=medium, top=thin)
            ws.cell(row=2, column=headerStartCol + 1).border = Border(right=medium, top=thin)

            sideStartRow = 5

            for key1, val1 in val['sideQres'].items():

                if key == 'Total':
                    cellQreLbl = ws.cell(row=sideStartRow, column=1)
                    cellQreLbl.value = val1["qreLbl"].split('. ')[0]
                    cellQreLbl.font = Font(bold=True)
                    cellQreLbl.fill = PatternFill('solid', fgColor='FFF2CC')
                    cellQreLbl.alignment = Alignment(horizontal='center', vertical='center')
                    cellQreLbl.border = Border(top=medium, bottom=thin)

                cellProd1 = ws.cell(row=3, column=headerStartCol)
                cellProd1.value = int(val1['productCodes'][0])
                cellProd1.font = Font(bold=True, color='FFFF00')
                cellProd1.fill = PatternFill('solid', fgColor='002060')
                cellProd1.alignment = Alignment(horizontal='center', vertical='center')
                cellProd1.border = Border(left=medium, right=thin, top=thin)

                cellProd2 = ws.cell(row=3, column=headerStartCol + 1)
                cellProd2.value = int(val1['productCodes'][1])
                cellProd2.font = Font(bold=True, color='FFFF00')
                cellProd2.fill = PatternFill('solid', fgColor='002060')
                cellProd2.alignment = Alignment(horizontal='center', vertical='center')
                cellProd2.border = Border(right=medium, top=thin)

                for key2, val2 in val1['sigResult'].items():

                    if key2 not in ['t2b', 'b2b', 'mean', 'std']:
                        cellValLbl = ws.cell(row=sideStartRow, column=2)
                        cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                        cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                        cellValLbl.value = val2['catLbl']
                        cellValLbl.alignment = Alignment(horizontal='right')

                        cellVal0.value = val2['val0']
                        cellVal1.value = val2['val1']

                        cellValLbl.border = Border(left=thin, right=medium, top=dot, bottom=dot)
                        cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                        cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                        if key2 == 'base':

                            cellValLbl.value = val1["qreLbl"].split('. ')[1]
                            cellValLbl.alignment = Alignment(horizontal='left')

                            cellValLbl.font, cellValLbl.fill = Font(bold=True), PatternFill('solid', fgColor='FFF2CC')
                            cellVal0.font, cellVal0.fill = Font(bold=True), PatternFill('solid', fgColor='FFF2CC')
                            cellVal1.font, cellVal1.fill = Font(bold=True), PatternFill('solid', fgColor='FFF2CC')

                            cellValLbl.border = Border(left=thin, top=medium, right=medium, bottom=thin)
                            cellVal0.border = Border(left=thin, top=medium, right=thin, bottom=thin)
                            cellVal1.border = Border(left=thin, top=medium, right=medium, bottom=thin)

                            if ws.cell(4, headerStartCol).value is None:
                                if ws.cell(4, 1).value is None:
                                    ws.cell(4, 1).value = 'N='
                                    ws.cell(4, 1).alignment = Alignment(horizontal='right')

                                    ws.cell(4, 1).border = Border(top=medium, bottom=medium, left=medium, right=medium)
                                    ws.cell(4, 2).border = Border(top=medium, bottom=medium, left=medium, right=medium)

                                    ws.cell(4, 1).font = Font(bold=True, color='FF0000')
                                    ws.cell(4, 1).fill = PatternFill('solid', fgColor='C6E0B4')

                                    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)


                                ws.cell(row=4, column=headerStartCol).value = val2['val0']
                                ws.cell(row=4, column=headerStartCol + 1).value = val2['val1']

                                ws.cell(4, headerStartCol).border = Border(top=medium, bottom=medium, left=medium, right=thin)
                                ws.cell(4, headerStartCol + 1).border = Border(top=medium, bottom=medium, left=medium, right=medium)

                                ws.cell(4, headerStartCol).font = Font(bold=True, color='FF0000')
                                ws.cell(4, headerStartCol).fill = PatternFill('solid', fgColor='C6E0B4')

                                ws.cell(4, headerStartCol + 1).font = Font(bold=True, color='FF0000')
                                ws.cell(4, headerStartCol + 1).fill = PatternFill('solid', fgColor='C6E0B4')

                            if val2['val0'] == ws.cell(4, headerStartCol).value:
                                cellVal0.value = None

                            if val2['val1'] == ws.cell(4, headerStartCol + 1).value:
                                cellVal1.value = None

                        else:

                            if key2 in ['mean', 'std']:
                                cellVal0.number_format = '0.00'
                                cellVal1.number_format = '0.00'
                            else:
                                if not val1['isCount']:
                                    if self.isDisplayPctSign:
                                        cellVal0.number_format = '0%'
                                        cellVal1.number_format = '0%'
                                    else:
                                        cellVal0.number_format = '0'
                                        cellVal1.number_format = '0'

                                        cellVal0.value = cellVal0.value * 100
                                        cellVal1.value = cellVal1.value * 100



                        sideStartRow += 1

                if val1['type'] in ['OL', 'JR', 'NUM']:

                    if val1['type'] == 'NUM':
                        lstAtt = ['mean']
                    else:
                        lstAtt = ['b2b', '3', 't2b', 'mean', 'std']

                    for item in lstAtt:

                        key2 = item
                        val2 = val1['sigResult'][key2]

                        cellValLbl = ws.cell(row=sideStartRow, column=2)
                        cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                        cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                        cellValLbl.alignment = Alignment(horizontal='right')

                        if key2 == '3':
                            cellValLbl.value = 'Medium' if val1['type'] == 'OL' else 'JR'
                        else:
                            cellValLbl.value = val2['catLbl']

                        cellVal0.value = val2['val0']
                        cellVal1.value = val2['val1']

                        cellValLbl.border = Border(left=thin, right=medium, top=dot, bottom=dot)
                        cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                        cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                        cellValLbl.font = Font(bold=True)
                        cellVal0.font = Font(bold=True)
                        cellVal1.font = Font(bold=True)

                        if key2 in ['mean', 'std']:
                            cellVal0.number_format = '0.00'
                            cellVal1.number_format = '0.00'

                            cellValLbl.font = Font(bold=True, color='4472C4')
                            cellVal0.font = Font(bold=True, color='4472C4')
                            cellVal1.font = Font(bold=True, color='4472C4')

                        else:
                            if not val1['isCount']:
                                if self.isDisplayPctSign:
                                    cellVal0.number_format = '0%'
                                    cellVal1.number_format = '0%'
                                else:
                                    cellVal0.number_format = '0'
                                    cellVal1.number_format = '0'

                                    cellVal0.value = cellVal0.value * 100
                                    cellVal1.value = cellVal1.value * 100

                        sideStartRow += 1

            headerStartCol += 2

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 46
        ws.freeze_panes = 'C5'

        icol1 = 3
        while ws.cell(row=1, column=icol1).value is not None:

            icol2 = icol1 + 1
            while ws.cell(row=1, column=icol1).value == ws.cell(row=1, column=icol2).value:
                icol2 += 1

            ws.merge_cells(start_row=1, start_column=icol1, end_row=1, end_column=icol2 - 1)

            ws.cell(row=1, column=icol1).border = Border(left=medium, top=medium, right=medium, bottom=thin)
            ws.cell(row=1, column=icol2 - 1).border = Border(left=medium, top=medium, right=medium, bottom=thin)

            icol1 = icol2

        for icol in range(1, headerStartCol):
            ws.cell(sideStartRow, icol).border = Border(top=medium)



    def toOlJrSummary(self, wb: openpyxl.workbook, dictTtest: dict, qType: str):

        print(f'Export Topline {qType} Summary')

        stt = 3 if qType == 'OL' else 4

        ws = wb.create_sheet(f'{stt}. {qType} Summary')
        ws.sheet_properties.tabColor = '002060'

        thin = Side(border_style='thin', color='000000')
        medium = Side(border_style='medium', color='000000')
        dot = Side(border_style='dotted', color='000000')

        cellTitle = ws.cell(row=1, column=1)
        cellTitle.value = self.toplineTitle
        cellTitle.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cellTitle.font = Font(bold=True, color='0070C0', size=24)
        ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=4)

        headerStartCol = 5
        sideStartRow = -1

        for key, val in dictTtest.items():

            subGroupCellQre = ws.cell(row=1, column=headerStartCol)
            subGroupCellQre.value = val['subGroupLbl']
            subGroupCellQre.font = Font(bold=True)
            subGroupCellQre.fill = PatternFill('solid', fgColor='FDE9D9')
            subGroupCellQre.alignment = Alignment(horizontal='center', vertical='center')

            for i in range(1, 8):
                ws.cell(row=1, column=headerStartCol + i).value = val['subGroupLbl']

            ws.merge_cells(start_row=2, start_column=headerStartCol, end_row=2, end_column=headerStartCol + 7)
            subGroupCellVal = ws.cell(row=2, column=headerStartCol)
            subGroupCellVal.value = val['subCodeLbl']
            subGroupCellVal.font = Font(bold=True)
            subGroupCellVal.fill = PatternFill('solid', fgColor='FFC000')
            subGroupCellVal.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=2, column=headerStartCol).border = Border(left=medium, top=thin)
            for i in range(1, 8):
                ws.cell(row=2, column=headerStartCol + i).border = Border(right=medium, top=thin)

            sideStartRow = 6

            for key1, val1 in val['sideQres'].items():

                if val1['type'] == qType:

                    if key == 'Total':
                        ws.cell(row=sideStartRow, column=1).value = val1['groupLbl']
                        ws.cell(row=sideStartRow, column=2).value = val1["qreLbl"].split('. ')[0]

                        if qType == 'OL':
                            ws.cell(row=sideStartRow, column=3).value = '5pt'
                        else:
                            ws.cell(row=sideStartRow, column=3).value = 'JR'

                        ws.cell(row=sideStartRow, column=4).value = val1["qreLbl"].split('. ')[1]

                        ws.cell(row=sideStartRow, column=1).border = Border(left=medium, right=medium)
                        ws.cell(row=sideStartRow, column=2).border = Border(left=medium, right=thin, top=dot, bottom=dot)
                        ws.cell(row=sideStartRow, column=3).border = Border(left=thin, right=thin, top=dot, bottom=dot)
                        ws.cell(row=sideStartRow, column=4).border = Border(left=thin, right=medium, top=dot, bottom=dot)


                    ws.merge_cells(start_row=3, start_column=headerStartCol, end_row=3, end_column=headerStartCol + 3)
                    cellProd1 = ws.cell(row=3, column=headerStartCol)
                    cellProd1.value = int(val1['productCodes'][0])
                    cellProd1.font = Font(bold=True, color='FFFF00')
                    cellProd1.fill = PatternFill('solid', fgColor='002060')
                    cellProd1.alignment = Alignment(horizontal='center', vertical='center')
                    cellProd1.border = Border(left=medium, right=medium, top=thin)

                    ws.merge_cells(start_row=3, start_column=headerStartCol + 4, end_row=3, end_column=headerStartCol + 7)
                    cellProd2 = ws.cell(row=3, column=headerStartCol + 4)
                    cellProd2.value = int(val1['productCodes'][1])
                    cellProd2.font = Font(bold=True, color='FFFF00')
                    cellProd2.fill = PatternFill('solid', fgColor='002060')
                    cellProd2.alignment = Alignment(horizontal='center', vertical='center')
                    cellProd2.border = Border(left=medium, right=medium, top=thin)


                    for i in range(4):

                        cellSP1 = ws.cell(4, headerStartCol + i)
                        cellSP2 = ws.cell(4, headerStartCol + i + 4)

                        cellSP1.alignment = Alignment(horizontal='center', vertical='center')
                        cellSP2.alignment = Alignment(horizontal='center', vertical='center')

                        if i == 0:
                            cellSP1.value = 'Mean'
                            cellSP2.value = cellSP1.value

                        elif i == 1:
                            cellSP1.value = 'B2B%'
                            cellSP2.value = cellSP1.value

                        elif i == 2:
                            cellSP1.value = 'Medium' if qType == 'OL' else 'JR%'
                            cellSP2.value = cellSP1.value

                        elif i == 3:
                            cellSP1.value = 'T2B%'
                            cellSP2.value = cellSP1.value

                        if i == 0:
                            cellSP1.border = Border(left=medium, right=thin, top=thin, bottom=thin)
                            cellSP2.border = Border(left=medium, right=thin, top=thin, bottom=thin)
                        elif i == 3:
                            cellSP1.border = Border(left=thin, right=medium, top=thin, bottom=thin)
                            cellSP2.border = Border(left=thin, right=medium, top=thin, bottom=thin)
                        else:
                            cellSP1.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                            cellSP2.border = Border(left=thin, right=thin, top=thin, bottom=thin)


                        cellSP1.font = Font(bold=True)
                        cellSP2.font = Font(bold=True)

                        cellSP1.fill = PatternFill('solid', fgColor='D9E1F2')
                        cellSP2.fill = PatternFill('solid', fgColor='D9E1F2')

                    for key2, val2 in val1['sigResult'].items():

                        if key2 == 'base':
                            cellValLbl = ws.cell(row=5, column=1)
                            cellVal0 = ws.cell(row=5, column=headerStartCol)
                            cellVal1 = ws.cell(row=5, column=headerStartCol + 4)

                            if cellValLbl.value is None:
                                ws.cell(5, 1).border = Border(top=medium, bottom=medium)
                                ws.cell(5, 2).border = Border(top=medium, bottom=medium)
                                ws.cell(5, 3).border = Border(top=medium, bottom=medium)
                                ws.cell(5, 4).border = Border(top=medium, bottom=medium, right=medium)

                                ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=4)
                                cellValLbl.value = 'N='
                                cellValLbl.font, cellValLbl.fill = Font(bold=True, color='FF0000'), PatternFill('solid', fgColor='C6E0B4')
                                cellValLbl.border = Border(left=medium, top=medium, right=medium, bottom=medium)
                                cellValLbl.alignment = Alignment(horizontal='right')


                            if cellVal0.value is None:
                                cellVal0.value = val2['val0']
                                ws.merge_cells(start_row=5, start_column=headerStartCol, end_row=5, end_column=headerStartCol + 3)
                                cellVal0.font, cellVal0.fill = Font(bold=True, color='FF0000'), PatternFill('solid', fgColor='C6E0B4')
                                # cellVal0.border = Border(left=thin, top=medium, right=thin, bottom=medium)
                                cellVal0.alignment = Alignment(horizontal='center')

                            if cellVal1.value is None:
                                cellVal1.value = val2['val1']
                                ws.merge_cells(start_row=5, start_column=headerStartCol + 4, end_row=5, end_column=headerStartCol + 7)
                                cellVal1.font, cellVal1.fill = Font(bold=True, color='FF0000'), PatternFill('solid', fgColor='C6E0B4')
                                # cellVal1.border = Border(left=thin, top=medium, right=medium, bottom=medium)
                                cellVal1.alignment = Alignment(horizontal='center')


                            for i in range(4):
                                cellSP1 = ws.cell(5, headerStartCol + i)
                                cellSP2 = ws.cell(5, headerStartCol + i + 4)

                                if i == 0:
                                    cellSP1.border = Border(left=medium, right=thin, top=dot, bottom=medium)
                                    cellSP2.border = Border(left=medium, right=thin, top=dot, bottom=medium)
                                elif i == 3:
                                    cellSP1.border = Border(left=thin, right=medium, top=dot, bottom=medium)
                                    cellSP2.border = Border(left=thin, right=medium, top=dot, bottom=medium)
                                else:
                                    cellSP1.border = Border(left=thin, right=thin, top=dot, bottom=medium)
                                    cellSP2.border = Border(left=thin, right=thin, top=dot, bottom=medium)

                        else:
                            if key2 in ['mean', 'b2b', '3', 't2b']:

                                stepSP1 = -1
                                cellBorder = Border()

                                if key2 == 'mean':
                                    stepSP1 = 0
                                    cellBorder = Border(left=medium, right=thin, top=dot, bottom=dot)

                                elif key2 == 'b2b':
                                    stepSP1 = 1
                                    cellBorder = Border(left=thin, right=thin, top=dot, bottom=dot)

                                elif key2 == '3':
                                    stepSP1 = 2
                                    cellBorder = Border(left=thin, right=thin, top=dot, bottom=dot)

                                elif key2 == 't2b':
                                    stepSP1 = 3
                                    cellBorder = Border(left=thin, right=medium, top=dot, bottom=dot)

                                stepSP2 = stepSP1 + 4

                                cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol + stepSP1)
                                cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + stepSP2)

                                cellVal0.border = cellBorder
                                cellVal1.border = cellBorder

                                cellVal0.value = val2['val0']
                                cellVal1.value = val2['val1']

                                if key2 == 'mean':
                                    cellVal0.number_format = '0.00'
                                    cellVal1.number_format = '0.00'
                                else:
                                    if not val1['isCount']:
                                        if self.isDisplayPctSign:
                                            cellVal0.number_format = '0%'
                                            cellVal1.number_format = '0%'
                                        else:
                                            cellVal0.number_format = '0'
                                            cellVal1.number_format = '0'

                                            cellVal0.value = cellVal0.value * 100
                                            cellVal1.value = cellVal1.value * 100

                    sideStartRow += 1


            headerStartCol += 8

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 4
        ws.column_dimensions['D'].width = 45
        ws.freeze_panes = 'E6'

        icol1 = 5
        while ws.cell(row=1, column=icol1).value is not None:

            icol2 = icol1 + 1
            while ws.cell(row=1, column=icol1).value == ws.cell(row=1, column=icol2).value:
                icol2 += 1

            ws.merge_cells(start_row=1, start_column=icol1, end_row=1, end_column=icol2 - 1)

            ws.cell(row=1, column=icol1).border = Border(left=medium, top=medium, right=medium, bottom=thin)
            ws.cell(row=1, column=icol2 - 1).border = Border(left=medium, top=medium, right=medium, bottom=thin)

            icol1 = icol2


        icol = 1
        irow1 = 6
        while ws.cell(row=irow1, column=icol).value is not None:

            irow2 = irow1 + 1
            while ws.cell(row=irow1, column=icol).value == ws.cell(row=irow2, column=icol).value:
                irow2 += 1

            ws.merge_cells(start_row=irow1, start_column=icol, end_row=irow2 - 1, end_column=icol)
            ws.cell(row=irow1, column=icol).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            ws.cell(row=irow1, column=icol).font = Font(bold=True)
            ws.cell(row=irow1, column=icol).fill = PatternFill('solid', fgColor='DDEBF7')

            ws.cell(row=irow1, column=icol).border = Border(left=medium, top=medium, right=medium, bottom=thin)

            ws.cell(row=irow2 - 1, column=2).border = Border(left=medium, bottom=medium, right=thin)
            ws.cell(row=irow2 - 1, column=3).border = Border(left=thin, bottom=medium, right=thin)
            ws.cell(row=irow2 - 1, column=4).border = Border(left=thin, bottom=medium, right=medium)

            for i in range(5, headerStartCol):

                if i % 2 != 0:
                    ws.cell(row=irow2 - 1, column=i).border = Border(bottom=medium, right=thin)
                else:
                    if i % 2 == 0 and i % 4 == 0:
                        ws.cell(row=irow2 - 1, column=i).border = Border(bottom=medium, right=medium)
                    else:
                        ws.cell(row=irow2 - 1, column=i).border = Border(bottom=medium, right=thin)

            irow1 = irow2


        for icol in range(1, headerStartCol):
            ws.cell(sideStartRow, icol).border = Border(top=medium)



    def toUandA(self, wb: openpyxl.workbook, dictUA: dict):

        print(f'Export Topline U&A')
        ws = wb.create_sheet(f'5. U&A')
        ws.sheet_properties.tabColor = '002060'

        thin = Side(border_style='thin', color='000000')
        medium = Side(border_style='medium', color='000000')
        dot = Side(border_style='dotted', color='000000')

        cellTitle = ws.cell(row=1, column=1)
        cellTitle.value = self.toplineTitle
        cellTitle.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cellTitle.font = Font(bold=True, color='0070C0', size=24)
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)

        headerStartCol = 3
        sideStartRow = -1

        for key, val in dictUA.items():

            subGroupCellQre = ws.cell(row=1, column=headerStartCol)
            subGroupCellQre.value = val['subGroupLbl']
            subGroupCellQre.font = Font(bold=True)
            subGroupCellQre.fill = PatternFill('solid', fgColor='FDE9D9')
            subGroupCellQre.alignment = Alignment(horizontal='center', vertical='center')

            subGroupCellVal = ws.cell(row=2, column=headerStartCol)
            subGroupCellVal.value = val['subCodeLbl']
            subGroupCellVal.font = Font(bold=True)
            subGroupCellVal.fill = PatternFill('solid', fgColor='FFC000')
            subGroupCellVal.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=2, column=headerStartCol).border = Border(left=medium, top=thin, right=medium)

            sideStartRow = 4
            groupLbl = ''
            for key1, val1 in val['sideQres'].items():

                if groupLbl != val1['groupLbl']:
                    groupLbl = val1['groupLbl']
                    ws.cell(row=sideStartRow, column=1).value = groupLbl
                    ws.cell(row=sideStartRow, column=1).font = Font(bold=True)
                    ws.cell(row=sideStartRow, column=1).fill = PatternFill('solid', fgColor='F4B084')
                    ws.cell(row=sideStartRow, column=1).alignment = Alignment(horizontal='center', vertical='center')

                    ws.cell(row=sideStartRow, column=1).border = Border(left=medium, right=medium, top=medium, bottom=medium)
                    ws.cell(row=sideStartRow, column=2).border = Border(left=medium, right=medium, top=medium, bottom=medium)

                    ws.merge_cells(start_row=sideStartRow, start_column=1, end_row=sideStartRow, end_column=2)

                    ws.cell(row=sideStartRow, column=headerStartCol).fill = PatternFill('solid', fgColor='F4B084')
                    ws.cell(row=sideStartRow, column=headerStartCol).border = Border(top=medium, bottom=medium)

                    sideStartRow += 1


                if key == 'Total':
                    cellQreLbl = ws.cell(row=sideStartRow, column=1)
                    cellQreLbl.value = val1["qreLbl"].split('. ')[0]
                    cellQreLbl.font = Font(bold=True)
                    cellQreLbl.fill = PatternFill('solid', fgColor='FFF2CC')
                    cellQreLbl.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
                    cellQreLbl.border = Border(top=medium, bottom=thin)

                for key2, val2 in val1['result'].items():

                    if key2 not in ['t2b', 'b2b', 'mean', 'std']:
                        cellValLbl = ws.cell(row=sideStartRow, column=2)
                        cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)

                        cellValLbl.value = val2['catLbl']
                        cellValLbl.alignment = Alignment(horizontal='right')

                        cellVal0.value = val2['val0']

                        cellValLbl.border = Border(left=thin, right=medium, top=dot, bottom=dot)
                        cellVal0.border = Border(left=medium, right=medium, top=dot, bottom=dot)

                        if key2 == 'base':

                            cellValLbl.value = val1["qreLbl"].split('. ')[1]
                            cellValLbl.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

                            cellValLbl.font, cellValLbl.fill = Font(bold=True), PatternFill('solid', fgColor='FFF2CC')
                            cellVal0.font, cellVal0.fill = Font(bold=True), PatternFill('solid', fgColor='FFF2CC')

                            cellValLbl.border = Border(left=thin, top=medium, right=medium, bottom=thin)
                            cellVal0.border = Border(left=thin, top=medium, right=medium, bottom=thin)

                            if ws.cell(3, headerStartCol).value is None:
                                if ws.cell(3, 1).value is None:
                                    ws.cell(3, 1).value = 'N='
                                    ws.cell(3, 1).alignment = Alignment(horizontal='right')

                                    ws.cell(3, 1).border = Border(top=medium, bottom=medium, left=medium, right=medium)
                                    ws.cell(3, 2).border = Border(top=medium, bottom=medium, left=medium, right=medium)

                                    ws.cell(3, 1).font = Font(bold=True, color='FF0000')
                                    ws.cell(3, 1).fill = PatternFill('solid', fgColor='C6E0B4')

                                    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)

                                ws.cell(row=3, column=headerStartCol).value = val2['val0']

                                ws.cell(3, headerStartCol).border = Border(top=medium, bottom=medium, left=medium, right=medium)


                                ws.cell(3, headerStartCol).font = Font(bold=True, color='FF0000')
                                ws.cell(3, headerStartCol).fill = PatternFill('solid', fgColor='C6E0B4')



                            if val2['val0'] == ws.cell(3, headerStartCol).value:
                                cellVal0.value = None



                        else:

                            if key2 in ['mean', 'std']:
                                cellVal0.number_format = '0.00'
                            else:
                                if not val1['isCount']:
                                    if self.isDisplayPctSign:
                                        cellVal0.number_format = '0%'
                                    else:
                                        cellVal0.number_format = '0'

                                        cellVal0.value = cellVal0.value * 100

                        sideStartRow += 1

                if val1['type'] in ['OL', 'JR', 'NUM']:

                    if val1['type'] == 'NUM':
                        lstAtt = ['mean']
                    else:
                        lstAtt = ['b2b', '3', 't2b', 'mean', 'std']

                    for item in lstAtt:

                        key2 = item
                        val2 = val1['result'][key2]

                        cellValLbl = ws.cell(row=sideStartRow, column=2)
                        cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)

                        cellValLbl.alignment = Alignment(horizontal='right')

                        if key2 == '3':
                            cellValLbl.value = 'Medium' if val1['type'] == 'OL' else 'JR'
                        else:
                            cellValLbl.value = val2['catLbl']

                        cellVal0.value = val2['val0']

                        cellValLbl.border = Border(left=thin, right=medium, top=dot, bottom=dot)
                        cellVal0.border = Border(left=medium, right=medium, top=dot, bottom=dot)

                        cellValLbl.font = Font(bold=True)
                        cellVal0.font = Font(bold=True)

                        if key2 in ['mean', 'std']:
                            cellVal0.number_format = '0.00'

                            cellValLbl.font = Font(bold=True, color='4472C4')
                            cellVal0.font = Font(bold=True, color='4472C4')

                        else:
                            if not val1['isCount']:
                                if self.isDisplayPctSign:
                                    cellVal0.number_format = '0%'
                                else:
                                    cellVal0.number_format = '0'

                                    cellVal0.value = cellVal0.value * 100

                        sideStartRow += 1

            headerStartCol += 1

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 70
        ws.freeze_panes = 'C4'

        icol1 = 3
        while ws.cell(row=1, column=icol1).value is not None:

            icol2 = icol1 + 1
            while ws.cell(row=1, column=icol1).value == ws.cell(row=1, column=icol2).value:
                icol2 += 1

            ws.merge_cells(start_row=1, start_column=icol1, end_row=1, end_column=icol2 - 1)

            ws.cell(row=1, column=icol1).border = Border(left=medium, top=medium, right=medium, bottom=thin)
            ws.cell(row=1, column=icol2 - 1).border = Border(left=medium, top=medium, right=medium, bottom=thin)

            icol1 = icol2

        for icol in range(1, headerStartCol):
            ws.cell(sideStartRow, icol).border = Border(top=medium)



    def toCorr(self, wb: openpyxl.workbook):

        print(f'Export Topline Correlation')
        ws = wb.create_sheet(f'6. Correlation')
        ws.sheet_properties.tabColor = '002060'

        cellTitle = ws.cell(row=1, column=1)
        cellTitle.value = self.toplineTitle
        cellTitle.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cellTitle.font = Font(bold=True, color='0070C0', size=24)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

        thin = Side(border_style='thin', color='000000')
        # medium = Side(border_style='medium', color='000000')
        dot = Side(border_style='dotted', color='000000')

        df = self.dfCorr.loc[:, self.dictCorrCols.keys()]
        dfCorr = df.corr()
        dfPval = df.corr(method=lambda x, y: pearsonr(x, y)[1]) - np.eye(*dfCorr.shape)
        dfP = dfPval.applymap(lambda x: ''.join(['*' for t in [0.01, 0.05] if x <= t]))

        irow = 2
        for key, val in self.dictCorrCols.items():

            ws.merge_cells(start_row=irow, start_column=1, end_row=irow + 1, end_column=1)
            ws.cell(row=irow, column=1).alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
            ws.cell(row=irow, column=1).value = val
            ws.cell(row=irow, column=1).border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.cell(row=irow + 1, column=1).border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.cell(row=irow, column=1).fill = PatternFill('solid', fgColor='DDEBF7')

            ws.cell(row=irow, column=2).value = 'Pearson Correlation'
            ws.cell(row=irow + 1, column=2).value = 'Sig. (2-tailed)'
            ws.cell(row=irow, column=2).border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.cell(row=irow + 1, column=2).border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.cell(row=irow, column=2).fill = PatternFill('solid', fgColor='FFF2CC')
            ws.cell(row=irow + 1, column=2).fill = PatternFill('solid', fgColor='FFF2CC')

            icol = 3
            for key2, val2 in self.dictCorrCols.items():
                ws.cell(row=1, column=icol).value = val2
                ws.cell(row=1, column=icol).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                ws.column_dimensions[get_column_letter(icol)].width = 12
                ws.cell(row=1, column=icol).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                ws.cell(row=1, column=icol).fill = PatternFill('solid', fgColor='DDEBF7')

                cellCorr = ws.cell(row=irow, column=icol)
                cellPval = ws.cell(row=irow + 1, column=icol)

                if dfCorr.at[key, key2] == 1:
                    strCorrVal = dfCorr.at[key, key2]
                else:
                    strCorrVal = str(dfCorr.at[key, key2].round(3))
                    if len(strCorrVal) == 5:
                        strCorrVal = f'{strCorrVal}{dfP.at[key, key2]}'
                    elif len(strCorrVal) == 4:
                        strCorrVal = f'{strCorrVal}0{dfP.at[key, key2]}'
                    elif len(strCorrVal) == 3:
                        strCorrVal = f'{strCorrVal}00{dfP.at[key, key2]}'
                    elif len(strCorrVal) == 2:
                        strCorrVal = f'{strCorrVal}000{dfP.at[key, key2]}'

                    strCorrVal = strCorrVal[1:]


                cellCorr.value = strCorrVal
                cellPval.value = dfPval.at[key, key2] if dfPval.at[key, key2] != 0 else None

                cellCorr.alignment = Alignment(horizontal='right', vertical='center')

                if cellPval.value is not None:
                    cellPval.number_format = '0.000'

                cellCorr.border = Border(left=thin, right=thin, top=dot, bottom=dot)
                cellPval.border = Border(left=thin, right=thin, top=dot, bottom=thin)


                icol += 1

            irow += 2

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 22
        ws.freeze_panes = 'C2'






