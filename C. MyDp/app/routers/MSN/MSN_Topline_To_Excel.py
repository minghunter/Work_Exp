import traceback
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet import worksheet as pyxl_ws


class ToplineToExcel:

    def __init__(self, topline_name: str, topline_title: str, dictTtest: dict, dictUA: dict, is_display_pct_sign: bool, is_jar_scale_3: bool, dictOE: dict):

        self.topline_name = topline_name
        self.topline_title = topline_title
        self.dictTtest = dictTtest
        self.dictUA = dictUA
        self.dictOE = dictOE

        self.is_display_pct_sign = is_display_pct_sign
        self.is_jar_scale_3 = is_jar_scale_3

        # 'Project_Information', 'OL_Summary', 'JAR_Summary_Callback', 'Tabulation_Callback', 'Tabulation_Diary_1',
        #             'Tabulation_Diary_2', 'OEs_Callback', 'OEs_Diary_1', 'OEs_Diary_2', 'Profile'

        self.lst_all_sheet = ['Project_Information', 'OL_Summary', 'JAR_Summary_Callback', 'Tabulation_Callback', 'OEs_Callback', 'Profile']


    def export_topline_excel_file(self, lstSheet, obj_prj_info):

        try:
            file_name = self.topline_name

            if 'Handcount' in lstSheet:

                self.generate_prj_info(file_name, obj_prj_info, is_handcount=True)

                wb = openpyxl.load_workbook(filename=file_name)

                self.generate_ol_summary(wb, lstQreType=['OL', 'FC'])

            else:
                self.generate_topline_skeleton(file_name)
                self.generate_prj_info(file_name, obj_prj_info, is_handcount=False)

                wb = openpyxl.load_workbook(filename=file_name)

                if 'OL_Summary' in lstSheet:
                    self.generate_ol_summary(wb)

                if 'JAR_Summary_Callback' in lstSheet:
                    self.generate_jr_summary(wb)

                if 'Tabulation_Callback' in lstSheet:
                    self.generate_tabulation(wb, 'Tabulation_Callback')

                if 'OEs_Callback' in lstSheet:
                    self.generate_tabulation(wb, 'OEs_Callback')

                if 'Profile' in lstSheet:
                    self.generate_profile(wb)

                self.hide_ungenerated_sheets(wb, lstSheet)

            wb.save(file_name)
            wb.close()

            return True, None

        except Exception:
            print(traceback.format_exc())
            return False, traceback.format_exc()


    def generate_topline_skeleton(self, file_name):
        wb = openpyxl.Workbook()

        try:

            # Generate content header
            ws_content = wb['Sheet']
            ws_content.title = 'Content'
            ws_content['A1'].value = 'CONTENT'
            ws_content['A1'].font = Font(color='FFFFFF', bold=True)
            ws_content['A1'].fill = PatternFill('solid', fgColor='5B9BD5')
            ws_content['A1'].alignment = Alignment(horizontal='center', vertical='justify')

            ws_content.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
            ws_content.column_dimensions['B'].width = 30
            ws_content.column_dimensions['C'].width = 20
            ws_content['C2'].value = 'Note'
            ws_content['C2'].font = Font(bold=True)

            self.set_border(ws_content, 'A1:C2')

            lst_sheet = self.lst_all_sheet

            for idx, ws_name in enumerate(lst_sheet):
                ws_content[f'A{idx + 3}'].value = idx + 1
                ws_content[f'B{idx + 3}'].value = ws_name
                ws_content[f'C{idx + 3}'].value = 'Theo mã SP' if 'Diary_1' in ws_name else ('Không theo mã SP' if 'Diary_2' in ws_name else '')
                self.set_border(ws_content, f'A{idx + 3}:C{idx + 3}')

                ws_new = wb.create_sheet(ws_name)

                ws_content[f'B{idx + 3}'].hyperlink = f'#{ws_name}!A1'
                ws_content[f'B{idx + 3}'].font = Font(color='0000FF')

                ws_new['A1'].value = 'Content'
                ws_new['A1'].hyperlink = '#Content!A1'
                ws_new['A1'].font = Font(color='0000FF')

            wb.save(file_name)
            wb.close()

            return True, None

        except Exception:
            print(traceback.format_exc())
            return False, traceback.format_exc()

        finally:
            wb.close()


    @staticmethod
    def set_border(ws, cell_range):
        thin = Side(border_style='thin', color='000000')
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


    def generate_prj_info(self, fileName, obj_prj_info, is_handcount):

        print('Export project information')

        df_prj_information = pd.DataFrame([
            ['1. Mục tiêu nghiên cứu'],
            [obj_prj_info['1']['val']],
            ['2. Đối tượng và phương pháp nghiên cứu'],
            [obj_prj_info['2_1']['val']],
            [obj_prj_info['2_2']['val']],
            [obj_prj_info['2_3']['val']],
            [obj_prj_info['2_4']['val']],
            ['3. Thông tin nghiên cứu (OL, JAR, Like/Dislikes, v.v…)'],
            [obj_prj_info['3']['val']],
            ['4. Action Standard'],
            [obj_prj_info['4']['val']],
            ['5. Thời gian thực hiện'],
            [obj_prj_info['5']['val']],
            ['6. Chú thích mã sản phẩm'],
            [obj_prj_info['6_1']['val']],
            [obj_prj_info['6_2']['val']],
            [obj_prj_info['6_3']['val']],
            [obj_prj_info['6_4']['val']],
            ['7. Lưu ý ký hiệu significance test '],
            ['Red'],
            ['Blue'],
            ['Green'],
        ], columns=['THÔNG TIN DỰ ÁN'])

        sheet_name = 'Project_Information'

        if is_handcount:
            with pd.ExcelWriter(fileName) as writer:
                df_prj_information.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)  # encoding='utf-8-sig'
        else:
            with pd.ExcelWriter(fileName, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df_prj_information.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)  # encoding='utf-8-sig'

        self.format_topline_sheet_prj_infor(fileName, sheet_name)


    @staticmethod
    def format_topline_sheet_prj_infor(xlsx_name, sheet_name):

        print('Export Topline Project Information')

        wb = openpyxl.load_workbook(filename=xlsx_name)
        ws = wb[sheet_name]

        ws.column_dimensions['A'].width = 150
        ws.column_dimensions['B'].width = 20
        thin = Side(border_style="thin", color="000000")

        for i in range(2, 25):

            cell = ws[f'A{i}']
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            if i in [2, 3, 5, 10, 12, 14, 16, 21]:
                cell.font = Font(color='FFFFFF', bold=True)
                cell.fill = PatternFill('solid', fgColor='5B9BD5')
            else:
                cell.fill = PatternFill('solid', fgColor='DDEBF7')
                cell.alignment = Alignment(horizontal='right', vertical='justify')

                if i == 22:
                    cell.font = Font(color='FF0000')  # red
                    ws[f'B{i}'] = 'significant at 95% CI'
                elif i == 23:
                    cell.font = Font(color='0000FF')  # blue
                    ws[f'B{i}'] = 'significant at 90% CI'
                elif i == 24:
                    cell.font = Font(color='00FF00')  # green
                    ws[f'B{i}'] = 'significant at 80% CI'


        wb.save(xlsx_name)
        wb.close()


    @staticmethod
    def format_cell(ws: pyxl_ws, props: dict):
        cell = ws.cell(row=props['row_col'][0], column=props['row_col'][1])

        for k, v in props.items():
            if k == 'value':
                cell.value = v
            elif k == 'alignment':
                cell.alignment = Alignment(wrap_text=v.get('wrap_text'), horizontal=v.get('horizontal'), vertical=v.get('vertical'))
            elif k == 'font':
                cell.font = Font(bold=v.get('bold'), color=v.get('color'), size=v.get('size'))
            elif k == 'fill':
                cell.fill = PatternFill(fill_type=v.get('fill_type'), fgColor=v.get('fgColor'))
            elif k == 'border':
                cell.border = Border(left=v.get('left'), right=v.get('right'), top=v.get('top'), bottom=v.get('bottom'))


            elif k == 'merge_cells':
                ws.merge_cells(start_row=v['start_row'], start_column=v['start_column'], end_row=v['end_row'], end_column=v['end_column'])


    @staticmethod
    def fillSigColor(sigVal, is_net=False):

        cellFont = Font(bold=is_net, color='00000000')

        if sigVal == 80:
            cellFont = Font(bold=is_net, color='0000FF00')  # green
        elif sigVal == 90:
            cellFont = Font(bold=is_net, color='000000FF')  # blue
        elif sigVal == 95:
            cellFont = Font(bold=is_net, color='00FF0000')  # red

        return cellFont


    def hide_ungenerated_sheets(self, wb: openpyxl.workbook, lst_generated_sheets):

        all_sheets = set(self.lst_all_sheet).union({'Content'})
        generated_sheets = set(lst_generated_sheets).union({'Content', 'Project_Information'})

        lst_hidden_sheet = all_sheets.difference(generated_sheets)

        ws_content = wb['Content']

        content_rows = {
            'Project_Information': 3,
            'OL_Summary': 4,
            'JAR_Summary_Callback': 5,
            'Tabulation_Callback': 6,
            'OEs_Callback': 7,
            'Profile': 8,

            # 'Tabulation_Diary_1': 7,
            # 'Tabulation_Diary_2': 8,
            # 'OEs_Diary_1': 10,
            # 'OEs_Diary_2': 11,
        }

        for sheet_name in lst_hidden_sheet:
            wb[sheet_name].sheet_state = 'hidden'
            ws_content.row_dimensions[content_rows[sheet_name]].hidden = True



    # FIND NEW WAY TO FORMAT TOPLINE - start on 24/08/2023



    def populate_data_table(self, ws: openpyxl.worksheet, props: dict):
        """
        - Need to populate data first, format excel style after that
        :return: formated worksheet
        """

        dict_data_container = props['data_container']

        if props['is_new']:
            ws.cell(1, props['start_header_col'] - 1).value = props['topline_title']
            ws.cell(2, props['start_header_col'] - 1).value = ws.title
            ws.cell(3, props['start_header_col'] - 1).value = 'Product code'
            ws.cell(4, props['start_header_col'] - 1).value = 'N='

        last_col = props['start_header_col']
        col_step = len(dict_data_container['Total']['sideQres']['1']['prod_cats'].keys())

        for hd_key, hd_val in dict_data_container.items():

            for step in range(col_step):  # will run from 0 to col_step - 1
                cur_col = last_col + step

                if props['is_new']:
                    ws.cell(1, cur_col).value = hd_val['subGroupLbl']
                    ws.cell(2, cur_col).value = hd_val['subCodeLbl']
                    ws.cell(3, cur_col).value = list(hd_val['sideQres']['1']['prod_cats'].values())[step]
                    ws.cell(4, cur_col).value = hd_val['sideQres']['1']['result']['base'][f'val{step}']

                last_row = props['start_side_row']

                for qre_key, qre_val in hd_val['sideQres'].items():

                    if qre_val['type'] not in list(props['qre_type_filter'].keys()):
                        continue

                    if cur_col == props['start_header_col']:  # only populate side axis 1 times

                        if props['extra_group_lbl']:
                            ws.cell(last_row, cur_col - 5).value = props['extra_group_lbl']

                        lst_qre_lbl = qre_val['qreLbl'].split('. ', 1)
                        ws.cell(last_row, cur_col - 4).value = qre_val['groupLbl']
                        ws.cell(last_row, cur_col - 3).value = lst_qre_lbl[0]
                        ws.cell(last_row, cur_col - 2).value = '5pt' if qre_val['type'] in ['OL'] else qre_val['type']
                        ws.cell(last_row, cur_col - 1).value = lst_qre_lbl[1]

                    for cat_key, cat_val in qre_val['result'].items():
                        if cat_key not in props['qre_type_filter'][qre_val['type']] and props['qre_type_filter'][qre_val['type']] != ['any']:
                            continue

                        cat_val_weight = 1
                        if not qre_val['isCount'] and cat_key not in ['mean', 'std', ''] and qre_val['type'] != 'NUM':
                            cat_val_weight = 100

                        cur_cell = ws.cell(last_row, cur_col)
                        cur_cell.value = cat_val[f'val{step}'] * cat_val_weight
                        cur_cell.font = self.fillSigColor(cat_val[f'sig{step}'])

                    last_row += 1





            last_col += col_step



    # END FIND NEW WAY TO FORMAT TOPLINE - start on 24/08/2023




    def generate_ol_summary(self, wb: openpyxl.workbook, lstQreType=None):

        # # TESTING HERE--------------------------------------------------------------------------------------------------
        # if lstQreType is None:
        #     print('Export Topline OL Summary')
        #     ws = wb['OL_Summary']
        #     is_handcount = False
        # else:
        #     print('Export Topline Handcount')
        #     ws = wb.create_sheet('KPI')
        #     is_handcount = True
        #
        # self.populate_data_table(ws, props={
        #     'topline_title': self.topline_title,
        #     'data_container': self.dictTtest.copy(),
        #     'is_new': True,
        #     'extra_group_lbl': 'CÁC CÂU OL(Mean)',
        #     'qre_type_filter': {
        #         'OL': ['mean']
        #     },
        #     'start_header_col': 6,
        #     'start_side_row': 5
        # })
        #
        # self.populate_data_table(ws, props={
        #     'topline_title': '',
        #     'data_container': self.dictTtest.copy(),
        #     'is_new': False,
        #     'extra_group_lbl': 'CÁC CÂU SO SÁNH',
        #     'qre_type_filter': {
        #         'FC': ['any']
        #     },
        #     'start_header_col': 6,
        #     'start_side_row': ws.max_row + 1
        # })
        #
        # if not is_handcount:
        #     self.populate_data_table(ws, props={
        #         'topline_title': '',
        #         'data_container': self.dictTtest.copy(),
        #         'is_new': False,
        #         'extra_group_lbl': 'Mức độ thích (%T2B)',
        #         'qre_type_filter': {
        #             'OL': ['t2b']
        #         },
        #         'start_header_col': 6,
        #         'start_side_row': ws.max_row + 1
        #     })
        #
        #     self.populate_data_table(ws, props={
        #         'topline_title': '',
        #         'data_container': self.dictTtest.copy(),
        #         'is_new': False,
        #         'extra_group_lbl': 'Mức độ thích (%B2B)',
        #         'qre_type_filter': {
        #             'OL': ['b2b']
        #         },
        #         'start_header_col': 6,
        #         'start_side_row': ws.max_row + 1
        #     })
        #
        #
        # # END TESTING HERE----------------------------------------------------------------------------------------------

        dictTtest = self.dictTtest

        if lstQreType is None:
            print('Export Topline OL Summary')
            lstQreType = ['OL', 'FC']  # 'JR'
            ws = wb['OL_Summary']
            is_handcount = False
        else:
            print('Export Topline Handcount')
            ws = wb.create_sheet('KPI')
            is_handcount = True

        thin = Side(border_style='thin', color='000000')
        medium = Side(border_style='medium', color='000000')
        dot = Side(border_style='dotted', color='000000')

        ws.row_dimensions[1].height, ws.row_dimensions[2].height = 22, 22

        self.format_cell(ws, props={
            'row_col': [1, 2],
            'value': self.topline_title,
            'alignment': {'wrap_text': True, 'horizontal': 'center', 'vertical': 'center'},
            'font': {'bold': True, 'color': '0070C0', 'size': 20},
            'merge_cells': {'start_row': 1, 'start_column': 2, 'end_row': 1, 'end_column': 5}
        })

        self.format_cell(ws, props={
            'row_col': [2, 1],
            'value': ws.title,
            'alignment': {'wrap_text': True, 'horizontal': 'center', 'vertical': 'center'},
            'font': {'bold': True, 'color': '000000', 'size': 20},
            'merge_cells': {'start_row': 2, 'start_column': 1, 'end_row': 2, 'end_column': 5}
        })

        self.format_cell(ws, props={
            'row_col': [3, 1],
            'value': 'Product code',
            'alignment': {'horizontal': 'right'},
            'font': {'bold': True, 'color': '000000'},
            'merge_cells': {'start_row': 3, 'start_column': 1, 'end_row': 3, 'end_column': 5}
        })

        headerStartCol = 6
        sideStartRow = -1

        for key, val in dictTtest.items():

            self.format_cell(ws, props={
                'row_col': [1, headerStartCol],
                'value': val['subGroupLbl'],
                'alignment': {'horizontal': 'center', 'vertical': 'center'},
                'font': {'bold': True},
                'fill': {'fill_type': 'solid', 'fgColor': 'FDE9D9'},
            })

            self.format_cell(ws, props={'row_col': [1, headerStartCol + 1], 'value': val['subGroupLbl']})

            self.format_cell(ws, props={
                'row_col': [2, headerStartCol],
                'value': val['subCodeLbl'],
                'alignment': {'horizontal': 'center', 'vertical': 'center'},
                'font': {'bold': True},
                'fill': {'fill_type': 'solid', 'fgColor': 'FFC000'},
                'merge_cells': {'start_row': 2, 'start_column': headerStartCol, 'end_row': 2, 'end_column': headerStartCol + 1}
            })

            self.format_cell(ws, props={'row_col': [2, headerStartCol], 'border': {'left': medium, 'top': thin}})

            self.format_cell(ws, props={'row_col': [2, headerStartCol + 1], 'border': {'right': medium, 'top': thin}})



            sideStartRow = 5

            # mean OL + JR & others Qre
            for key1, val1 in val['sideQres'].items():

                if val1['type'] in lstQreType:  # ['OL', 'JR', 'FC']

                    if key in ['Total', 'TOTAL']:
                        ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                        ws.cell(row=sideStartRow, column=5).value = val1["qreLbl"].split('. ')[1]

                        if val1['type'] == 'OL':
                            ws.cell(row=sideStartRow, column=4).value = '5pt'
                        else:
                            ws.cell(row=sideStartRow, column=4).value = val1['type']

                        if val1['type'] == 'JR':
                            ws.cell(row=sideStartRow, column=3).fill = PatternFill('solid', fgColor='FCE4D6')
                            ws.cell(row=sideStartRow, column=4).fill = PatternFill('solid', fgColor='FCE4D6')
                            ws.cell(row=sideStartRow, column=5).fill = PatternFill('solid', fgColor='FCE4D6')

                        ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                        ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                        ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)


                    cellProd1 = ws.cell(row=3, column=headerStartCol)
                    cellProd1.value = val1['prod_cats']['1']  # int(val1['prod_cats']['1'])
                    cellProd1.font = Font(bold=True, color='FFFF00')
                    cellProd1.fill = PatternFill('solid', fgColor='002060')
                    cellProd1.alignment = Alignment(horizontal='center', vertical='center')
                    cellProd1.border = Border(left=medium, right=thin, top=thin)

                    cellProd2 = ws.cell(row=3, column=headerStartCol + 1)
                    cellProd2.value = val1['prod_cats']['2']  # int(val1['prod_cats']['2'])
                    cellProd2.font = Font(bold=True, color='FFFF00')
                    cellProd2.fill = PatternFill('solid', fgColor='002060')
                    cellProd2.alignment = Alignment(horizontal='center', vertical='center')
                    cellProd2.border = Border(right=medium, top=thin)

                    for key2, val2 in val1['result'].items():

                        if key2 == 'base':
                            cellValLbl = ws.cell(row=4, column=1)
                            cellVal0 = ws.cell(row=4, column=headerStartCol)
                            cellVal1 = ws.cell(row=4, column=headerStartCol + 1)

                            if cellValLbl.value is None:
                                ws.cell(4, 1).border = Border(top=medium, bottom=medium)
                                ws.cell(4, 2).border = Border(top=medium, bottom=medium)
                                ws.cell(4, 3).border = Border(top=medium, bottom=medium)
                                ws.cell(4, 4).border = Border(top=medium, bottom=medium)
                                ws.cell(4, 5).border = Border(top=medium, bottom=medium, right=medium)

                                ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)
                                cellValLbl.value = 'N='  # val2['catLbl']
                                cellValLbl.font, cellValLbl.fill = Font(bold=True, color='FF0000'), PatternFill('solid', fgColor='C6E0B4')
                                cellValLbl.border = Border(left=medium, top=medium, right=medium, bottom=thin)
                                cellValLbl.alignment = Alignment(horizontal='right')


                            if cellVal0.value is None:
                                cellVal0.value = val2['val0']
                                cellVal0.font, cellVal0.fill = Font(bold=True, color='FF0000'), PatternFill('solid', fgColor='C6E0B4')
                                cellVal0.border = Border(left=thin, top=medium, right=thin, bottom=medium)

                            if cellVal1.value is None:
                                cellVal1.value = val2['val1']
                                cellVal1.font, cellVal1.fill = Font(bold=True, color='FF0000'), PatternFill('solid', fgColor='C6E0B4')
                                cellVal1.border = Border(left=thin, top=medium, right=medium, bottom=medium)

                        else:
                            if key2 == 'mean' or (key2 != 'mean' and val1['type'] not in ['OL', 'JR']):
                                cellValLbl = ws.cell(row=sideStartRow, column=2)
                                cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                                cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)


                                if val1['type'] in ['FC']:
                                    ws.cell(row=sideStartRow, column=1).value = 'CÁC CÂU SO SÁNH'
                                elif val1['type'] in ['SA', 'MA']:
                                    ws.cell(row=sideStartRow, column=1).value = 'CÁC CÂU KHÁC'
                                    ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                                    ws.cell(row=sideStartRow, column=4).value = val1["qreLbl"].split('. ')[1]
                                    ws.cell(row=sideStartRow, column=5).value = val2['catLbl']

                                    ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                                    ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                                    ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)
                                else:
                                    ws.cell(row=sideStartRow, column=1).value = 'CÁC CÂU OL(Mean) & JR (Mean)'  # val2['catLbl']

                                cellValLbl.value = val1['groupLbl']


                                cellVal0.value = val2['val0']
                                cellVal1.value = val2['val1']

                                cellValLbl.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                                cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                                cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                                if key2 in 'mean':
                                    cellVal0.number_format = '0.00'
                                    cellVal1.number_format = '0.00'
                                else:
                                    if not val1['isCount']:
                                        if self.is_display_pct_sign:
                                            cellVal0.number_format = '0%'
                                            cellVal1.number_format = '0%'
                                        else:
                                            cellVal0.number_format = '0'
                                            cellVal1.number_format = '0'

                                            cellVal0.value = cellVal0.value * 100
                                            cellVal1.value = cellVal1.value * 100

                                if val1['type'] in ['OL', 'FC']:
                                    cellVal0.font = self.fillSigColor(val2['sig0'])
                                    cellVal1.font = self.fillSigColor(val2['sig1'])

                                sideStartRow += 1

            # mean JR
            for key1, val1 in val['sideQres'].items():

                if val1['type'] == 'JR' and val1['type'] in lstQreType:

                    if key in ['Total', 'TOTAL']:
                        ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                        ws.cell(row=sideStartRow, column=5).value = val1["qreLbl"].split('. ')[1]
                        ws.cell(row=sideStartRow, column=4).value = val1['type']

                        ws.cell(row=sideStartRow, column=3).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=4).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=5).fill = PatternFill('solid', fgColor='FCE4D6')

                        ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                        ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                        ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)


                    for key2, val2 in val1['result'].items():

                        if key2 == 'mean':
                            cellValLbl = ws.cell(row=sideStartRow, column=2)
                            cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                            cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                            ws.cell(row=sideStartRow, column=1).value = 'CÁC CÂU JR (Mean) - Sig vs 3'  # val2['catLbl']

                            cellValLbl.value = val1['groupLbl']

                            cellVal0.value = val2['val0']
                            cellVal1.value = val2['val1']

                            cellValLbl.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                            cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                            cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                            cellVal0.number_format = '0.00'
                            cellVal1.number_format = '0.00'

                            sideStartRow += 1

            # % JR - Code 3
            for key1, val1 in val['sideQres'].items():

                if val1['type'] == 'JR' and val1['type'] in lstQreType:

                    if key in ['Total', 'TOTAL']:
                        ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                        ws.cell(row=sideStartRow, column=5).value = val1["qreLbl"].split('. ')[1]
                        ws.cell(row=sideStartRow, column=4).value = val1['type']

                        ws.cell(row=sideStartRow, column=3).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=4).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=5).fill = PatternFill('solid', fgColor='FCE4D6')

                        ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                        ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                        ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)

                    for key2, val2 in val1['result'].items():

                        if key2 == '3':
                            cellValLbl = ws.cell(row=sideStartRow, column=2)
                            cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                            cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                            ws.cell(row=sideStartRow, column=1).value = 'CÁC CÂU JR (%JR)'  # val2['catLbl']

                            cellValLbl.value = val1['groupLbl']

                            cellVal0.value = val2['val0']
                            cellVal1.value = val2['val1']

                            cellValLbl.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                            cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                            cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                            if self.is_display_pct_sign:
                                cellVal0.number_format = '0%'
                                cellVal1.number_format = '0%'
                            else:
                                cellVal0.number_format = '0'
                                cellVal1.number_format = '0'

                                cellVal0.value = cellVal0.value * 100
                                cellVal1.value = cellVal1.value * 100

                            sideStartRow += 1

            # % T2B - OL
            for key1, val1 in val['sideQres'].items():

                if val1['type'] == 'OL' and not is_handcount:

                    if key in ['Total', 'TOTAL']:
                        ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                        ws.cell(row=sideStartRow, column=5).value = val1["qreLbl"].split('. ')[1]
                        ws.cell(row=sideStartRow, column=4).value = '5pt'  # val1['type']

                        ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                        ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                        ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)

                    for key2, val2 in val1['result'].items():

                        if key2 == 't2b':
                            cellValLbl = ws.cell(row=sideStartRow, column=2)
                            cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                            cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                            ws.cell(row=sideStartRow, column=1).value = 'Mức độ thích (%T2B)'  # val2['catLbl']

                            cellValLbl.value = val1['groupLbl']

                            cellVal0.value = val2['val0']
                            cellVal1.value = val2['val1']

                            cellValLbl.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                            cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                            cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                            if self.is_display_pct_sign:
                                cellVal0.number_format = '0%'
                                cellVal1.number_format = '0%'
                            else:
                                cellVal0.number_format = '0'
                                cellVal1.number_format = '0'

                                cellVal0.value = cellVal0.value * 100
                                cellVal1.value = cellVal1.value * 100

                            cellVal0.font = self.fillSigColor(val2['sig0'])
                            cellVal1.font = self.fillSigColor(val2['sig1'])

                            sideStartRow += 1

            # % T2B - JR
            for key1, val1 in val['sideQres'].items():

                if val1['type'] == 'JR' and val1['type'] in lstQreType:

                    if key in ['Total', 'TOTAL']:
                        ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                        ws.cell(row=sideStartRow, column=5).value = val1["qreLbl"].split('. ')[1]
                        ws.cell(row=sideStartRow, column=4).value = val1['type']

                        ws.cell(row=sideStartRow, column=3).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=4).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=5).fill = PatternFill('solid', fgColor='FCE4D6')

                        ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                        ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                        ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)

                    for key2, val2 in val1['result'].items():

                        if key2 == 't2b':
                            cellValLbl = ws.cell(row=sideStartRow, column=2)
                            cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                            cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                            ws.cell(row=sideStartRow, column=1).value = 'CÁC CÂU JR (%T2B)'  # val2['catLbl']

                            cellValLbl.value = val1['groupLbl']

                            cellVal0.value = val2['val0']
                            cellVal1.value = val2['val1']

                            cellValLbl.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                            cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                            cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                            if self.is_display_pct_sign:
                                cellVal0.number_format = '0%'
                                cellVal1.number_format = '0%'
                            else:
                                cellVal0.number_format = '0'
                                cellVal1.number_format = '0'

                                cellVal0.value = cellVal0.value * 100
                                cellVal1.value = cellVal1.value * 100

                            sideStartRow += 1

            # % B2B - OL
            for key1, val1 in val['sideQres'].items():

                if val1['type'] == 'OL' and not is_handcount:

                    if key in ['Total', 'TOTAL']:
                        ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                        ws.cell(row=sideStartRow, column=5).value = val1["qreLbl"].split('. ')[1]
                        ws.cell(row=sideStartRow, column=4).value = '5pt'  # val1['type']

                        ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                        ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                        ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)

                    for key2, val2 in val1['result'].items():

                        if key2 == 'b2b':
                            cellValLbl = ws.cell(row=sideStartRow, column=2)
                            cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                            cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                            ws.cell(row=sideStartRow, column=1).value = 'Mức độ thích (%B2B)'  # val2['catLbl']

                            cellValLbl.value = val1['groupLbl']

                            cellVal0.value = val2['val0']
                            cellVal1.value = val2['val1']

                            cellValLbl.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                            cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                            cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                            if self.is_display_pct_sign:
                                cellVal0.number_format = '0%'
                                cellVal1.number_format = '0%'
                            else:
                                cellVal0.number_format = '0'
                                cellVal1.number_format = '0'

                                cellVal0.value = cellVal0.value * 100
                                cellVal1.value = cellVal1.value * 100

                            cellVal0.font = self.fillSigColor(val2['sig0'])
                            cellVal1.font = self.fillSigColor(val2['sig1'])

                            sideStartRow += 1

            # % B2B - JR
            for key1, val1 in val['sideQres'].items():

                if val1['type'] == 'JR' and val1['type'] in lstQreType:

                    if key in ['Total', 'TOTAL']:
                        ws.cell(row=sideStartRow, column=3).value = val1["qreLbl"].split('. ')[0]
                        ws.cell(row=sideStartRow, column=5).value = val1["qreLbl"].split('. ')[1]
                        ws.cell(row=sideStartRow, column=4).value = val1['type']

                        ws.cell(row=sideStartRow, column=3).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=4).fill = PatternFill('solid', fgColor='FCE4D6')
                        ws.cell(row=sideStartRow, column=5).fill = PatternFill('solid', fgColor='FCE4D6')

                        ws.cell(sideStartRow, 3).border = Border(top=dot, bottom=dot, left=medium, right=thin)
                        ws.cell(sideStartRow, 4).border = Border(top=dot, bottom=dot, left=thin, right=thin)
                        ws.cell(sideStartRow, 5).border = Border(top=dot, bottom=dot, left=thin, right=medium)

                    for key2, val2 in val1['result'].items():

                        if key2 == 'b2b':
                            cellValLbl = ws.cell(row=sideStartRow, column=2)
                            cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                            cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                            ws.cell(row=sideStartRow, column=1).value = 'CÁC CÂU JR (%B2B)'  # val2['catLbl']

                            cellValLbl.value = val1['groupLbl']

                            cellVal0.value = val2['val0']
                            cellVal1.value = val2['val1']

                            cellValLbl.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                            cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                            cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                            if self.is_display_pct_sign:
                                cellVal0.number_format = '0%'
                                cellVal1.number_format = '0%'
                            else:
                                cellVal0.number_format = '0'
                                cellVal1.number_format = '0'

                                cellVal0.value = cellVal0.value * 100
                                cellVal1.value = cellVal1.value * 100

                            sideStartRow += 1


            headerStartCol += 2

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 50
        ws.freeze_panes = 'F5'

        icol1 = 6
        while ws.cell(row=1, column=icol1).value is not None:

            icol2 = icol1 + 1
            while ws.cell(row=1, column=icol1).value == ws.cell(row=1, column=icol2).value:
                icol2 += 1

            ws.merge_cells(start_row=1, start_column=icol1, end_row=1, end_column=icol2 - 1)

            ws.cell(row=1, column=icol1).border = Border(left=medium, top=medium, right=medium, bottom=thin)
            ws.cell(row=1, column=icol2 - 1).border = Border(left=medium, top=medium, right=medium, bottom=thin)

            icol1 = icol2


        icol = 3
        irow1 = 5
        while ws.cell(row=irow1, column=icol).value is not None:

            irow2 = irow1 + 1
            while ws.cell(row=irow1, column=icol).value == ws.cell(row=irow2, column=icol).value:
                irow2 += 1

            if irow1 != irow2 - 1:
                ws.merge_cells(start_row=irow1, start_column=icol, end_row=irow2 - 1, end_column=icol)
                ws.cell(row=irow1, column=icol).alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

                ws.merge_cells(start_row=irow1, start_column=icol + 1, end_row=irow2 - 1, end_column=icol + 1)
                ws.cell(row=irow1, column=icol + 1).alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

                ws.cell(row=irow1, column=icol).border = Border(left=medium, top=thin, right=thin, bottom=thin)

                ws.cell(row=irow2 - 1, column=3).border = Border(left=medium, right=thin, bottom=thin)
                ws.cell(row=irow2 - 1, column=4).border = Border(left=thin, right=thin, bottom=thin)
                ws.cell(row=irow2 - 1, column=5).border = Border(left=thin, right=medium, bottom=thin)

                for i in range(6, headerStartCol):

                    if i % 2 == 0:
                        ws.cell(row=irow2 - 1, column=i).border = Border(left=medium, right=thin, bottom=thin, top=dot)
                    else:
                        ws.cell(row=irow2 - 1, column=i).border = Border(left=thin, right=medium, bottom=thin, top=dot)

            irow1 = irow2


        icol = 2
        irow1 = 5
        while ws.cell(row=irow1, column=icol).value is not None:

            irow2 = irow1 + 1
            while ws.cell(row=irow1, column=icol).value == ws.cell(row=irow2, column=icol).value:
                irow2 += 1

            ws.merge_cells(start_row=irow1, start_column=icol, end_row=irow2 - 1, end_column=icol)
            ws.cell(row=irow1, column=icol).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            ws.cell(row=irow1, column=icol).font = Font(bold=True)

            if icol == 1:
                ws.cell(row=irow1, column=icol).fill = PatternFill('solid', fgColor='FFE699')

            else:
                ws.cell(row=irow1, column=icol).fill = PatternFill('solid', fgColor='DDEBF7')

            ws.cell(row=irow1, column=icol).border = Border(left=medium, top=thin, right=medium, bottom=thin)

            ws.cell(row=irow2 - 1, column=1).border = Border(left=medium, bottom=thin)
            ws.cell(row=irow2 - 1, column=2).border = Border(left=medium, bottom=thin)
            ws.cell(row=irow2 - 1, column=3).border = Border(left=medium, bottom=thin)
            ws.cell(row=irow2 - 1, column=4).border = Border(left=thin, bottom=thin)
            ws.cell(row=irow2 - 1, column=5).border = Border(left=thin, bottom=thin, right=medium)

            for i in range(6, headerStartCol):

                if i % 2 == 0:
                    ws.cell(row=irow2 - 1, column=i).border = Border(bottom=thin, right=thin)
                else:
                    ws.cell(row=irow2 - 1, column=i).border = Border(bottom=thin, right=medium)


            irow1 = irow2


        icol = 1
        irow1 = 5
        while ws.cell(row=irow1, column=icol).value is not None:

            irow2 = irow1 + 1
            while ws.cell(row=irow1, column=icol).value == ws.cell(row=irow2, column=icol).value:
                irow2 += 1

            ws.merge_cells(start_row=irow1, start_column=icol, end_row=irow2 - 1, end_column=icol)
            ws.cell(row=irow1, column=icol).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            ws.cell(row=irow1, column=icol).font = Font(bold=True)

            if icol == 1:
                ws.cell(row=irow1, column=icol).fill = PatternFill('solid', fgColor='FFE699')

            else:
                ws.cell(row=irow1, column=icol).fill = PatternFill('solid', fgColor='DDEBF7')

            ws.cell(row=irow1, column=icol).border = Border(left=medium, top=medium, right=medium, bottom=thin)

            ws.cell(row=irow2 - 1, column=1).border = Border(left=medium, bottom=medium)
            ws.cell(row=irow2 - 1, column=2).border = Border(left=medium, bottom=medium)
            ws.cell(row=irow2 - 1, column=3).border = Border(left=medium, bottom=medium)
            ws.cell(row=irow2 - 1, column=4).border = Border(left=thin, bottom=medium)
            ws.cell(row=irow2 - 1, column=5).border = Border(left=thin, bottom=medium, right=medium)

            for i in range(6, headerStartCol):

                if i % 2 == 0:
                    ws.cell(row=irow2 - 1, column=i).border = Border(bottom=medium, right=thin)
                else:
                    ws.cell(row=irow2 - 1, column=i).border = Border(bottom=medium, right=medium)


            irow1 = irow2


        ws.cell(row=sideStartRow + 1, column=4).value = 'Lưu ý:'
        ws.cell(row=sideStartRow + 1, column=4).font = Font(bold=True, color='00FF0000')

        ws.cell(row=sideStartRow + 2, column=4).value = 'Chỉ chạy Sig. Test cho các câu OL, các câu JAR không chạy Sig. Test'

        ws.cell(row=sideStartRow + 3, column=4).value = 'Ký hiệu Sig. Test'
        ws.cell(row=sideStartRow + 3, column=4).font = Font(bold=True)

        ws.cell(row=sideStartRow + 4, column=4).value = 'Đỏ: Sig. ở 95% trở lên'
        ws.cell(row=sideStartRow + 4, column=4).font = Font(bold=True, color='00FF0000')

        ws.cell(row=sideStartRow + 5, column=4).value = 'Xanh dương: Sig. ở 90% đến dưới 95%'
        ws.cell(row=sideStartRow + 5, column=4).font = Font(bold=True, color='000000FF')

        ws.cell(row=sideStartRow + 6, column=4).value = 'Xanh lá: Sig. ở 80% đến dưới 90%'
        ws.cell(row=sideStartRow + 6, column=4).font = Font(bold=True, color='0000FF00')






    def generate_jr_summary(self, wb: openpyxl.workbook):

        print(f'Export Topline JR Summary')
        dictTtest = self.dictTtest
        qType = 'JR'
        ws = wb['JAR_Summary_Callback']

        thin = Side(border_style='thin', color='000000')
        medium = Side(border_style='medium', color='000000')
        dot = Side(border_style='dotted', color='000000')

        ws.row_dimensions[1].height, ws.row_dimensions[2].height = 22, 22

        cellTitle = ws.cell(row=1, column=2)
        cellTitle.value = self.topline_title
        cellTitle.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cellTitle.font = Font(bold=True, color='0070C0', size=20)
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)

        cellSheetName = ws.cell(row=2, column=1)
        cellSheetName.value = ws.title
        cellSheetName.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cellSheetName.font = Font(bold=True, color='000000', size=20)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

        cellSPCode = ws.cell(row=3, column=1)
        cellSPCode.value = 'Product code'
        cellSPCode.font = Font(bold=True, color='000000')
        cellSPCode.alignment = Alignment(horizontal='right')
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)

        cell_N = ws.cell(row=4, column=1)
        cell_N.value = 'N='
        cell_N.font, cell_N.fill = Font(bold=True, color='FF0000'), PatternFill('solid', fgColor='C6E0B4')
        cell_N.border = Border(left=medium, top=medium, right=medium, bottom=thin)
        cell_N.alignment = Alignment(horizontal='right')
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4)

        cellAtt = ws.cell(row=5, column=1)
        cellAtt.font, cellAtt.fill = Font(bold=True, color='FF0000'), PatternFill('solid', fgColor='D9E1F2')
        cellAtt.border = Border(left=medium, top=thin, right=medium, bottom=medium)
        cellAtt.alignment = Alignment(horizontal='right')
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=4)

        headerStartCol = 5
        sideStartRow = -1

        for key, val in dictTtest.items():

            subGroupCellQre = ws.cell(row=1, column=headerStartCol)
            subGroupCellQre.value = val['subGroupLbl']
            subGroupCellQre.font = Font(bold=True)
            subGroupCellQre.fill = PatternFill('solid', fgColor='FDE9D9')
            subGroupCellQre.alignment = Alignment(horizontal='center', vertical='center')

            for i in range(1, 8):
                ws.cell(row=1, column=headerStartCol + i).value = val['subGroupLbl']

            ws.merge_cells(start_row=2, start_column=headerStartCol, end_row=2, end_column=headerStartCol + 7)
            subGroupCellVal = ws.cell(row=2, column=headerStartCol)
            subGroupCellVal.value = val['subCodeLbl']
            subGroupCellVal.font = Font(bold=True)
            subGroupCellVal.fill = PatternFill('solid', fgColor='FFC000')
            subGroupCellVal.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=2, column=headerStartCol).border = Border(left=medium, top=thin)
            for i in range(1, 8):
                ws.cell(row=2, column=headerStartCol + i).border = Border(right=medium, top=thin)

            sideStartRow = 6

            for key1, val1 in val['sideQres'].items():

                if val1['type'] == qType:

                    if key in ['Total', 'TOTAL']:
                        ws.cell(row=sideStartRow, column=1).value = val1['groupLbl']
                        ws.cell(row=sideStartRow, column=2).value = val1["qreLbl"].split('. ')[0]

                        if qType == 'OL':
                            ws.cell(row=sideStartRow, column=3).value = '5pt'
                        else:
                            ws.cell(row=sideStartRow, column=3).value = 'JR'

                        ws.cell(row=sideStartRow, column=4).value = val1["qreLbl"].split('. ')[1]

                        ws.cell(row=sideStartRow, column=1).border = Border(left=medium, right=medium)
                        ws.cell(row=sideStartRow, column=2).border = Border(left=medium, right=thin, top=dot, bottom=dot)
                        ws.cell(row=sideStartRow, column=3).border = Border(left=thin, right=thin, top=dot, bottom=dot)
                        ws.cell(row=sideStartRow, column=4).border = Border(left=thin, right=medium, top=dot, bottom=dot)


                    ws.merge_cells(start_row=3, start_column=headerStartCol, end_row=3, end_column=headerStartCol + 3)
                    cellProd1 = ws.cell(row=3, column=headerStartCol)
                    cellProd1.value = val1['prod_cats']['1']  # int(val1['prod_cats']['1'])
                    cellProd1.font = Font(bold=True, color='FFFF00')
                    cellProd1.fill = PatternFill('solid', fgColor='002060')
                    cellProd1.alignment = Alignment(horizontal='center', vertical='center')
                    cellProd1.border = Border(left=medium, right=medium, top=thin)

                    ws.merge_cells(start_row=3, start_column=headerStartCol + 4, end_row=3, end_column=headerStartCol + 7)
                    cellProd2 = ws.cell(row=3, column=headerStartCol + 4)
                    cellProd2.value = val1['prod_cats']['2']  # int(val1['prod_cats']['2'])
                    cellProd2.font = Font(bold=True, color='FFFF00')
                    cellProd2.fill = PatternFill('solid', fgColor='002060')
                    cellProd2.alignment = Alignment(horizontal='center', vertical='center')
                    cellProd2.border = Border(left=medium, right=medium, top=thin)

                    for i in range(4):

                        cellSP1 = ws.cell(5, headerStartCol + i)
                        cellSP2 = ws.cell(5, headerStartCol + i + 4)

                        cellSP1.alignment = Alignment(horizontal='center', vertical='center')
                        cellSP2.alignment = Alignment(horizontal='center', vertical='center')

                        if i == 0:
                            cellSP1.value = 'Mean'
                            cellSP2.value = cellSP1.value

                        elif i == 1:
                            cellSP1.value = '% B2B'
                            cellSP2.value = cellSP1.value

                        elif i == 2:
                            cellSP1.value = 'Medium' if qType == 'OL' else 'JAR'
                            cellSP2.value = cellSP1.value

                        elif i == 3:
                            cellSP1.value = '% T2B'
                            cellSP2.value = cellSP1.value

                        if i == 0:
                            cellSP1.border = Border(left=medium, right=thin, top=thin, bottom=medium)
                            cellSP2.border = Border(left=medium, right=thin, top=thin, bottom=medium)
                        elif i == 3:
                            cellSP1.border = Border(left=thin, right=medium, top=thin, bottom=medium)
                            cellSP2.border = Border(left=thin, right=medium, top=thin, bottom=medium)
                        else:
                            cellSP1.border = Border(left=thin, right=thin, top=thin, bottom=medium)
                            cellSP2.border = Border(left=thin, right=thin, top=thin, bottom=medium)


                        cellSP1.font = Font(bold=True)
                        cellSP2.font = Font(bold=True)

                        cellSP1.fill = PatternFill('solid', fgColor='D9E1F2')
                        cellSP2.fill = PatternFill('solid', fgColor='D9E1F2')

                    for key2, val2 in val1['result'].items():

                        if key2 == 'base':

                            cellVal0 = ws.cell(row=4, column=headerStartCol)
                            cellVal1 = ws.cell(row=4, column=headerStartCol + 4)

                            if cellVal0.value is None:
                                cellVal0.value = val2['val0']
                                ws.merge_cells(start_row=4, start_column=headerStartCol, end_row=4, end_column=headerStartCol + 3)
                                cellVal0.font, cellVal0.fill = Font(bold=True, color='FF0000'), PatternFill('solid', fgColor='C6E0B4')
                                cellVal0.alignment = Alignment(horizontal='center')

                            if cellVal1.value is None:
                                cellVal1.value = val2['val1']
                                ws.merge_cells(start_row=4, start_column=headerStartCol + 4, end_row=4, end_column=headerStartCol + 7)
                                cellVal1.font, cellVal1.fill = Font(bold=True, color='FF0000'), PatternFill('solid', fgColor='C6E0B4')
                                cellVal1.alignment = Alignment(horizontal='center')


                            for i in range(4):
                                cellSP1 = ws.cell(4, headerStartCol + i)
                                cellSP2 = ws.cell(4, headerStartCol + i + 4)

                                if i == 0:
                                    cellSP1.border = Border(left=medium, right=thin, top=thin, bottom=thin)
                                    cellSP2.border = Border(left=medium, right=thin, top=thin, bottom=thin)
                                elif i == 3:
                                    cellSP1.border = Border(left=thin, right=medium, top=thin, bottom=thin)
                                    cellSP2.border = Border(left=thin, right=medium, top=thin, bottom=thin)
                                else:
                                    cellSP1.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                                    cellSP2.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                        else:
                            if key2 in ['mean', 'b2b', '3', 't2b']:

                                stepSP1 = -1
                                cellBorder = Border()

                                if key2 == 'mean':
                                    stepSP1 = 0
                                    cellBorder = Border(left=medium, right=thin, top=dot, bottom=dot)

                                elif key2 == 'b2b':
                                    stepSP1 = 1
                                    cellBorder = Border(left=thin, right=thin, top=dot, bottom=dot)

                                elif key2 == '3':
                                    stepSP1 = 2
                                    cellBorder = Border(left=thin, right=thin, top=dot, bottom=dot)

                                elif key2 == 't2b':
                                    stepSP1 = 3
                                    cellBorder = Border(left=thin, right=medium, top=dot, bottom=dot)

                                stepSP2 = stepSP1 + 4

                                cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol + stepSP1)
                                cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + stepSP2)

                                cellVal0.border = cellBorder
                                cellVal1.border = cellBorder

                                cellVal0.value = val2['val0']
                                cellVal1.value = val2['val1']

                                if key2 == 'mean':

                                    if not self.is_jar_scale_3:
                                        cellVal0.font = self.fillSigColor(val2['sig0'])
                                        cellVal1.font = self.fillSigColor(val2['sig1'])

                                    cellVal0.number_format = '0.00'
                                    cellVal1.number_format = '0.00'
                                else:
                                    if not val1['isCount']:
                                        if self.is_display_pct_sign:
                                            cellVal0.number_format = '0%'
                                            cellVal1.number_format = '0%'
                                        else:
                                            cellVal0.number_format = '0'
                                            cellVal1.number_format = '0'

                                            cellVal0.value = cellVal0.value * 100
                                            cellVal1.value = cellVal1.value * 100

                    sideStartRow += 1


            headerStartCol += 8

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 4
        ws.column_dimensions['D'].width = 45
        ws.freeze_panes = 'E6'

        icol1 = 5
        while ws.cell(row=1, column=icol1).value is not None:

            icol2 = icol1 + 1
            while ws.cell(row=1, column=icol1).value == ws.cell(row=1, column=icol2).value:
                icol2 += 1

            ws.merge_cells(start_row=1, start_column=icol1, end_row=1, end_column=icol2 - 1)

            ws.cell(row=1, column=icol1).border = Border(left=medium, top=medium, right=medium, bottom=thin)
            ws.cell(row=1, column=icol2 - 1).border = Border(left=medium, top=medium, right=medium, bottom=thin)

            icol1 = icol2


        icol = 1
        irow1 = 6
        while ws.cell(row=irow1, column=icol).value is not None:

            irow2 = irow1 + 1
            while ws.cell(row=irow1, column=icol).value == ws.cell(row=irow2, column=icol).value:
                irow2 += 1

            ws.merge_cells(start_row=irow1, start_column=icol, end_row=irow2 - 1, end_column=icol)
            ws.cell(row=irow1, column=icol).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            ws.cell(row=irow1, column=icol).font = Font(bold=True)
            ws.cell(row=irow1, column=icol).fill = PatternFill('solid', fgColor='DDEBF7')

            ws.cell(row=irow1, column=icol).border = Border(left=medium, top=medium, right=medium, bottom=thin)

            ws.cell(row=irow2 - 1, column=2).border = Border(left=medium, bottom=medium, right=thin)
            ws.cell(row=irow2 - 1, column=3).border = Border(left=thin, bottom=medium, right=thin)
            ws.cell(row=irow2 - 1, column=4).border = Border(left=thin, bottom=medium, right=medium)

            for i in range(5, headerStartCol):

                if i % 2 != 0:
                    ws.cell(row=irow2 - 1, column=i).border = Border(bottom=medium, right=thin)
                else:
                    if i % 2 == 0 and i % 4 == 0:
                        ws.cell(row=irow2 - 1, column=i).border = Border(bottom=medium, right=medium)
                    else:
                        ws.cell(row=irow2 - 1, column=i).border = Border(bottom=medium, right=thin)

            irow1 = irow2


        for icol in range(1, headerStartCol):
            ws.cell(sideStartRow, icol).border = Border(top=medium)


        # UPDATE SIG JAR WITH 3 on 07/08/2023---------------------------------------------------------------------------
        if not self.is_jar_scale_3:
            ws.cell(row=sideStartRow + 1, column=4).value = 'Lưu ý:'
            ws.cell(row=sideStartRow + 1, column=4).font = Font(bold=True, color='00FF0000')

            ws.cell(row=sideStartRow + 2, column=4).value = 'Sig. Test Mean JAR(by scale 5) so với 3'

            ws.cell(row=sideStartRow + 3, column=4).value = 'Ký hiệu Sig. Test'
            ws.cell(row=sideStartRow + 3, column=4).font = Font(bold=True)

            ws.cell(row=sideStartRow + 4, column=4).value = 'Đỏ: Sig. ở 95% trở lên'
            ws.cell(row=sideStartRow + 4, column=4).font = Font(bold=True, color='00FF0000')

            ws.cell(row=sideStartRow + 5, column=4).value = 'Xanh dương: Sig. ở 90% đến dưới 95%'
            ws.cell(row=sideStartRow + 5, column=4).font = Font(bold=True, color='000000FF')

            ws.cell(row=sideStartRow + 6, column=4).value = 'Xanh lá: Sig. ở 80% đến dưới 90%'
            ws.cell(row=sideStartRow + 6, column=4).font = Font(bold=True, color='0000FF00')
        # END UPDATE SIG JAR WITH 3 on 07/08/2023-----------------------------------------------------------------------



    def generate_tabulation(self, wb: openpyxl.workbook, sheet_name: str):

        print(f'Export Topline {sheet_name}')

        if sheet_name == 'Tabulation_Callback':
            dictTtest = self.dictTtest
        else:
            dictTtest = self.dictOE

        ws = wb[sheet_name]

        ws.column_dimensions['A'].width, ws.column_dimensions['B'].width = 10, 60
        ws.row_dimensions[1].height, ws.row_dimensions[2].height = 22, 22

        thin = Side(border_style='thin', color='000000')
        medium = Side(border_style='medium', color='000000')
        dot = Side(border_style='dotted', color='000000')

        cellTitle = ws.cell(row=1, column=2)
        cellTitle.value = self.topline_title
        cellTitle.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cellTitle.font = Font(bold=True, color='0070C0', size=20)

        cellSheetName = ws.cell(row=2, column=1)
        cellSheetName.value = ws.title
        cellSheetName.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cellSheetName.font = Font(bold=True, color='000000', size=20)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)

        cellSPCode = ws.cell(row=3, column=1)
        cellSPCode.value = 'Product code'
        cellSPCode.font = Font(bold=True, color='000000')
        cellSPCode.alignment = Alignment(horizontal='right')
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)

        headerStartCol = 3
        sideStartRow = -1

        # lst_hidden_key1 = list()
        # lst_hidden_key2 = list()

        for key, val in dictTtest.items():

            subGroupCellQre = ws.cell(row=1, column=headerStartCol)
            subGroupCellQre.value = val['subGroupLbl']
            subGroupCellQre.font = Font(bold=True)
            subGroupCellQre.fill = PatternFill('solid', fgColor='FDE9D9')
            subGroupCellQre.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=1, column=headerStartCol + 1).value = val['subGroupLbl']

            ws.merge_cells(start_row=2, start_column=headerStartCol, end_row=2, end_column=headerStartCol + 1)
            subGroupCellVal = ws.cell(row=2, column=headerStartCol)
            subGroupCellVal.value = val['subCodeLbl']
            subGroupCellVal.font = Font(bold=True)
            subGroupCellVal.fill = PatternFill('solid', fgColor='FFC000')
            subGroupCellVal.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=2, column=headerStartCol).border = Border(left=medium, top=thin)
            ws.cell(row=2, column=headerStartCol + 1).border = Border(right=medium, top=thin)

            sideStartRow = 5

            for key1, val1 in val['sideQres'].items():

                if key in ['Total', 'TOTAL']:
                    cellQreLbl = ws.cell(row=sideStartRow, column=1)
                    cellQreLbl.value = val1["qreLbl"].split('. ')[0]
                    cellQreLbl.font = Font(bold=True)
                    cellQreLbl.fill = PatternFill('solid', fgColor='FFF2CC')
                    cellQreLbl.alignment = Alignment(horizontal='center', vertical='center')
                    cellQreLbl.border = Border(top=medium, bottom=thin)

                cellProd1 = ws.cell(row=3, column=headerStartCol)
                cellProd1.value = val1['prod_cats']['1']  # int(val1['prod_cats']['1'])
                cellProd1.font = Font(bold=True, color='FFFF00')
                cellProd1.fill = PatternFill('solid', fgColor='002060')
                cellProd1.alignment = Alignment(horizontal='center', vertical='center')
                cellProd1.border = Border(left=medium, right=thin, top=thin)

                cellProd2 = ws.cell(row=3, column=headerStartCol + 1)
                cellProd2.value = val1['prod_cats']['2']  # int(val1['prod_cats']['2'])
                cellProd2.font = Font(bold=True, color='FFFF00')
                cellProd2.fill = PatternFill('solid', fgColor='002060')
                cellProd2.alignment = Alignment(horizontal='center', vertical='center')
                cellProd2.border = Border(right=medium, top=thin)

                for key2, val2 in val1['result'].items():

                    # if sheet_name == 'OEs_Callback':
                    #     if val2['val0'] == val2['val1'] == 0:
                    #
                    #         if key in ['Total', 'TOTAL']:
                    #             if key1 not in lst_hidden_key1:
                    #                 lst_hidden_key1.append(key1)
                    #
                    #             if key2 not in lst_hidden_key2:
                    #                 lst_hidden_key2.append(key2)
                    #             continue
                    #
                    #         if key1 in lst_hidden_key1 and key2 in lst_hidden_key2:
                    #             continue


                    if key2 not in ['t2b', 'b2b', 'mean', 'std', 'group: 0-4', 'group: 5', 'group: 6-10']:
                        cellValLbl = ws.cell(row=sideStartRow, column=2)
                        cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                        cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                        cellValLbl.value = val2['catLbl']
                        cellValLbl.alignment = Alignment(horizontal='right')

                        # new
                        if sheet_name == 'OEs_Callback':
                            cellVal0.font = self.fillSigColor(val2['sig0'])
                            cellVal1.font = self.fillSigColor(val2['sig1'])

                        if 'NET' in val2['catLbl'] or 'COMBINE' in val2['catLbl']:
                            cellValLbl.value = val2['catLbl'].replace('(NET)', '').replace('(COMBINE)', '')

                            cellValLbl.font = Font(bold=True)
                            cellVal0.font = Font(bold=True)
                            cellVal1.font = Font(bold=True)

                            cellVal0.font = self.fillSigColor(val2['sig0'], is_net=True)
                            cellVal1.font = self.fillSigColor(val2['sig1'], is_net=True)


                        cellVal0.value = val2['val0']
                        cellVal1.value = val2['val1']

                        cellValLbl.border = Border(left=thin, right=medium, top=dot, bottom=dot)
                        cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                        cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                        if key2 == 'base':

                            cellValLbl.value = val1["qreLbl"].split('. ')[1]
                            cellValLbl.alignment = Alignment(horizontal='left')

                            cellValLbl.font, cellValLbl.fill = Font(bold=True), PatternFill('solid', fgColor='FFF2CC')
                            cellVal0.font, cellVal0.fill = Font(bold=True), PatternFill('solid', fgColor='FFF2CC')
                            cellVal1.font, cellVal1.fill = Font(bold=True), PatternFill('solid', fgColor='FFF2CC')

                            cellValLbl.border = Border(left=thin, top=medium, right=medium, bottom=thin)
                            cellVal0.border = Border(left=thin, top=medium, right=thin, bottom=thin)
                            cellVal1.border = Border(left=thin, top=medium, right=medium, bottom=thin)

                            if ws.cell(4, headerStartCol).value is None:
                                if ws.cell(4, 1).value is None:
                                    ws.cell(4, 1).value = 'N='
                                    ws.cell(4, 1).alignment = Alignment(horizontal='right')

                                    ws.cell(4, 1).border = Border(top=medium, bottom=medium, left=medium, right=medium)
                                    ws.cell(4, 2).border = Border(top=medium, bottom=medium, left=medium, right=medium)

                                    ws.cell(4, 1).font = Font(bold=True, color='FF0000')
                                    ws.cell(4, 1).fill = PatternFill('solid', fgColor='C6E0B4')

                                    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)

                                ws.cell(row=4, column=headerStartCol).value = val2['val0']
                                ws.cell(row=4, column=headerStartCol + 1).value = val2['val1']

                                ws.cell(4, headerStartCol).border = Border(top=medium, bottom=medium, left=medium,
                                                                           right=thin)
                                ws.cell(4, headerStartCol + 1).border = Border(top=medium, bottom=medium, left=medium,
                                                                               right=medium)

                                ws.cell(4, headerStartCol).font = Font(bold=True, color='FF0000')
                                ws.cell(4, headerStartCol).fill = PatternFill('solid', fgColor='C6E0B4')

                                ws.cell(4, headerStartCol + 1).font = Font(bold=True, color='FF0000')
                                ws.cell(4, headerStartCol + 1).fill = PatternFill('solid', fgColor='C6E0B4')

                            if val2['val0'] == ws.cell(4, headerStartCol).value:
                                cellVal0.value = None

                            if val2['val1'] == ws.cell(4, headerStartCol + 1).value:
                                cellVal1.value = None

                        else:

                            if key2 in ['mean', 'std', 'min', 'max', 'median', 'first_quartile', 'third_quartile']:
                                cellVal0.number_format = '0.00'
                                cellVal1.number_format = '0.00'
                            else:

                                if not val1['isCount']:
                                    if self.is_display_pct_sign:
                                        cellVal0.number_format = '0%'
                                        cellVal1.number_format = '0%'
                                    else:
                                        cellVal0.number_format = '0'
                                        cellVal1.number_format = '0'

                                        cellVal0.value = cellVal0.value * 100
                                        cellVal1.value = cellVal1.value * 100

                        sideStartRow += 1

                if val1['type'] in ['OL', 'JR', 'NUM']:

                    if val1['type'] == 'NUM':
                        lstAtt = ['mean', 'std']
                    else:
                        if 'group: 0-4' in val1['result'] or 'group: 5' in val1['result'] or 'group: 6-10' in val1['result']:
                            lstAtt = ['group: 0-4', 'group: 5', 'group: 6-10', 'mean', 'std']
                        else:
                            lstAtt = ['b2b', '3', 't2b', 'mean', 'std']

                    for item in lstAtt:

                        key2 = item

                        val2 = val1['result'][key2]

                        cellValLbl = ws.cell(row=sideStartRow, column=2)
                        cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)
                        cellVal1 = ws.cell(row=sideStartRow, column=headerStartCol + 1)

                        cellValLbl.alignment = Alignment(horizontal='right')

                        if key2 == '3':
                            cellValLbl.value = 'Medium' if val1['type'] == 'OL' else 'JR'
                        else:
                            cellValLbl.value = val2['catLbl']

                        cellVal0.value = val2['val0']
                        cellVal1.value = val2['val1']

                        cellValLbl.border = Border(left=thin, right=medium, top=dot, bottom=dot)
                        cellVal0.border = Border(left=medium, right=thin, top=dot, bottom=dot)
                        cellVal1.border = Border(left=thin, right=medium, top=dot, bottom=dot)

                        cellValLbl.font = Font(bold=True)
                        cellVal0.font = Font(bold=True)
                        cellVal1.font = Font(bold=True)

                        if key2 in ['mean', 'std']:
                            cellVal0.number_format = '0.00'
                            cellVal1.number_format = '0.00'

                            cellValLbl.font = Font(bold=True, color='4472C4')
                            cellVal0.font = Font(bold=True, color='4472C4')
                            cellVal1.font = Font(bold=True, color='4472C4')

                        else:
                            if not val1['isCount']:
                                if self.is_display_pct_sign:
                                    cellVal0.number_format = '0%'
                                    cellVal1.number_format = '0%'
                                else:
                                    cellVal0.number_format = '0'
                                    cellVal1.number_format = '0'

                                    cellVal0.value = cellVal0.value * 100
                                    cellVal1.value = cellVal1.value * 100

                        sideStartRow += 1

            headerStartCol += 2

        ws.freeze_panes = 'C5'

        icol1 = 3
        while ws.cell(row=1, column=icol1).value is not None:

            icol2 = icol1 + 1
            while ws.cell(row=1, column=icol1).value == ws.cell(row=1, column=icol2).value:
                icol2 += 1

            ws.merge_cells(start_row=1, start_column=icol1, end_row=1, end_column=icol2 - 1)

            ws.cell(row=1, column=icol1).border = Border(left=medium, top=medium, right=medium, bottom=thin)
            ws.cell(row=1, column=icol2 - 1).border = Border(left=medium, top=medium, right=medium, bottom=thin)

            icol1 = icol2

        for icol in range(1, headerStartCol):
            ws.cell(sideStartRow, icol).border = Border(top=medium)


    def generate_profile(self, wb: openpyxl.workbook):

        print(f'Export Topline Profile')
        ws = wb['Profile']
        dictUA = self.dictUA

        thin = Side(border_style='thin', color='000000')
        medium = Side(border_style='medium', color='000000')
        dot = Side(border_style='dotted', color='000000')

        ws.column_dimensions['A'].width, ws.column_dimensions['B'].width = 10, 60
        ws.row_dimensions[1].height, ws.row_dimensions[2].height = 22, 22

        cellTitle = ws.cell(row=1, column=2)
        cellTitle.value = self.topline_title
        cellTitle.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cellTitle.font = Font(bold=True, color='0070C0', size=20)

        cellSheetName = ws.cell(row=2, column=1)
        cellSheetName.value = ws.title
        cellSheetName.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cellSheetName.font = Font(bold=True, color='000000', size=20)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)

        cell_N = ws.cell(row=3, column=1)
        cell_N.border = Border(top=medium, bottom=medium, left=medium, right=medium)
        cell_N.value = 'N='
        cell_N.alignment = Alignment(horizontal='right')
        cell_N.font = Font(bold=True, color='FF0000')
        cell_N.fill = PatternFill('solid', fgColor='C6E0B4')
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)

        headerStartCol = 3
        sideStartRow = -1

        for key, val in dictUA.items():

            subGroupCellQre = ws.cell(row=1, column=headerStartCol)
            subGroupCellQre.value = val['subGroupLbl']
            subGroupCellQre.font = Font(bold=True)
            subGroupCellQre.fill = PatternFill('solid', fgColor='FDE9D9')
            subGroupCellQre.alignment = Alignment(horizontal='center', vertical='center')

            subGroupCellVal = ws.cell(row=2, column=headerStartCol)
            subGroupCellVal.value = val['subCodeLbl']
            subGroupCellVal.font = Font(bold=True)
            subGroupCellVal.fill = PatternFill('solid', fgColor='FFC000')
            subGroupCellVal.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=2, column=headerStartCol).border = Border(left=medium, top=thin, right=medium)

            sideStartRow = 4
            groupLbl = ''
            for key1, val1 in val['sideQres'].items():

                if groupLbl != val1['groupLbl']:
                    groupLbl = val1['groupLbl']
                    ws.cell(row=sideStartRow, column=1).value = groupLbl
                    ws.cell(row=sideStartRow, column=1).font = Font(bold=True)
                    ws.cell(row=sideStartRow, column=1).fill = PatternFill('solid', fgColor='F4B084')
                    ws.cell(row=sideStartRow, column=1).alignment = Alignment(horizontal='center', vertical='center')

                    ws.cell(row=sideStartRow, column=1).border = Border(left=medium, right=medium, top=medium,
                                                                        bottom=medium)
                    ws.cell(row=sideStartRow, column=2).border = Border(left=medium, right=medium, top=medium,
                                                                        bottom=medium)

                    ws.merge_cells(start_row=sideStartRow, start_column=1, end_row=sideStartRow, end_column=2)

                    ws.cell(row=sideStartRow, column=headerStartCol).fill = PatternFill('solid', fgColor='F4B084')
                    ws.cell(row=sideStartRow, column=headerStartCol).border = Border(top=medium, bottom=medium)

                    sideStartRow += 1

                if key in ['Total', 'TOTAL']:
                    cellQreLbl = ws.cell(row=sideStartRow, column=1)
                    cellQreLbl.value = val1["qreLbl"].split('. ')[0]
                    cellQreLbl.font = Font(bold=True)
                    cellQreLbl.fill = PatternFill('solid', fgColor='FFF2CC')
                    cellQreLbl.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
                    cellQreLbl.border = Border(top=medium, bottom=thin)

                for key2, val2 in val1['result'].items():

                    if key2 not in ['t2b', 'b2b', 'mean', 'std']:
                        cellValLbl = ws.cell(row=sideStartRow, column=2)
                        cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)

                        cellValLbl.alignment = Alignment(horizontal='right')

                        if not cellValLbl.value:
                            cellValLbl.value = val2['catLbl']

                        cellVal0.value = val2['val0']

                        cellValLbl.border = Border(left=thin, right=medium, top=dot, bottom=dot)
                        cellVal0.border = Border(left=medium, right=medium, top=dot, bottom=dot)

                        if key2 == 'base':

                            cellValLbl.value = val1["qreLbl"].split('. ')[1]
                            cellValLbl.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

                            cellValLbl.font, cellValLbl.fill = Font(bold=True), PatternFill('solid', fgColor='FFF2CC')
                            cellVal0.font, cellVal0.fill = Font(bold=True), PatternFill('solid', fgColor='FFF2CC')

                            cellValLbl.border = Border(left=thin, top=medium, right=medium, bottom=thin)
                            cellVal0.border = Border(left=thin, top=medium, right=medium, bottom=thin)

                            if ws.cell(3, headerStartCol).value is None:

                                ws.cell(row=3, column=headerStartCol).value = val2['val0']

                                ws.cell(3, headerStartCol).border = Border(top=medium, bottom=medium, left=medium,
                                                                           right=medium)

                                ws.cell(3, headerStartCol).font = Font(bold=True, color='FF0000')
                                ws.cell(3, headerStartCol).fill = PatternFill('solid', fgColor='C6E0B4')

                            if val2['val0'] == ws.cell(3, headerStartCol).value:
                                cellVal0.value = None

                        else:

                            if key2 in ['mean', 'std', 'min', 'max', 'median', 'first_quartile', 'third_quartile']:
                                cellVal0.number_format = '0.00'
                            else:
                                if not val1['isCount']:
                                    if self.is_display_pct_sign:
                                        cellVal0.number_format = '0%'
                                    else:
                                        cellVal0.number_format = '0'

                                        cellVal0.value = cellVal0.value * 100

                        sideStartRow += 1

                if val1['type'] in ['OL', 'JR', 'NUM']:

                    if val1['type'] == 'NUM':
                        lstAtt = ['mean', 'std']
                    else:
                        if 'b2b' not in val1['result'].keys() and 't2b' not in val1['result'].keys():
                            lstAtt = ['mean', 'std']
                        else:
                            lstAtt = ['b2b', '3', 't2b', 'mean', 'std']

                    for item in lstAtt:

                        key2 = item

                        val2 = val1['result'][key2]

                        cellValLbl = ws.cell(row=sideStartRow, column=2)
                        cellVal0 = ws.cell(row=sideStartRow, column=headerStartCol)

                        cellValLbl.alignment = Alignment(horizontal='right')

                        if key2 == '3':
                            cellValLbl.value = 'Medium' if val1['type'] == 'OL' else 'JR'
                        else:
                            cellValLbl.value = val2['catLbl']

                        cellVal0.value = val2['val0']

                        cellValLbl.border = Border(left=thin, right=medium, top=dot, bottom=dot)
                        cellVal0.border = Border(left=medium, right=medium, top=dot, bottom=dot)

                        cellValLbl.font = Font(bold=True)
                        cellVal0.font = Font(bold=True)

                        if key2 in ['mean', 'std']:
                            cellVal0.number_format = '0.00'

                            cellValLbl.font = Font(bold=True, color='4472C4')
                            cellVal0.font = Font(bold=True, color='4472C4')

                        else:
                            if not val1['isCount']:
                                if self.is_display_pct_sign:
                                    cellVal0.number_format = '0%'
                                else:
                                    cellVal0.number_format = '0'

                                    cellVal0.value = cellVal0.value * 100

                        sideStartRow += 1

            headerStartCol += 1

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 70
        ws.freeze_panes = 'C4'

        icol1 = 3
        while ws.cell(row=1, column=icol1).value is not None:

            icol2 = icol1 + 1
            while ws.cell(row=1, column=icol1).value == ws.cell(row=1, column=icol2).value:
                icol2 += 1

            ws.merge_cells(start_row=1, start_column=icol1, end_row=1, end_column=icol2 - 1)

            ws.cell(row=1, column=icol1).border = Border(left=medium, top=medium, right=medium, bottom=thin)
            ws.cell(row=1, column=icol2 - 1).border = Border(left=medium, top=medium, right=medium, bottom=thin)

            icol1 = icol2

        for icol in range(1, headerStartCol):
            ws.cell(sideStartRow, icol).border = Border(top=medium)



