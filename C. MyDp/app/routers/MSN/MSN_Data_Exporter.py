from .MSN_Data_Converter import QMeFileConvert
from .MSN_Topline_Exporter_v2 import ToplineExporterV2
from .MSN_Addin_Variables_v2 import AddinVariables

import pandas as pd
import numpy as np
import pyreadstat
import zipfile
import traceback
import os
import io

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl import load_workbook


class DataExporter(QMeFileConvert, ToplineExporterV2):

    def __init__(self):

        QMeFileConvert.__init__(self)

        # self.prj = dict()
        # self.obj_section = dict()
        # self.obj_prj_info = dict()

        self.strProjectName = str()

        self.is_export_option_general = False
        self.is_export_option_scr_only = False
        self.is_export_option_codeframe = False

        self.strRidColName = str()
        self.strOrderColName = str()
        self.is_split_callback = False
        self.lst_split_with_cols = list()

        self.is_has_data_plm = False

        self.strFilter = str()
        self.strRotColName = str()
        self.strProdColName = str()

        # self.dict_prod_cats = dict()

        self.lst_force_choice = list()

        self.df_combined_oe = pd.DataFrame()

        self.df_scr_format = pd.DataFrame()

        self.df_plm_scr_format = pd.DataFrame()
        self.df_plm_main_format = pd.DataFrame()

        self.df_main_format = pd.DataFrame()

        self.df_pre_format = pd.DataFrame()

        self.df_data_scr_org = pd.DataFrame()
        self.df_info_scr_org = pd.DataFrame()

        self.df_data_plm_org = pd.DataFrame()
        self.df_info_plm_org = pd.DataFrame()

        self.df_data_main_org = pd.DataFrame()
        self.df_info_main_org = pd.DataFrame()

        self.df_data_scr_out, self.df_info_scr_out = pd.DataFrame(), pd.DataFrame()
        self.df_data_plm_scr_out, self.df_info_plm_scr_out = pd.DataFrame(), pd.DataFrame()

        self.df_data_plm_prod_1_out, self.df_info_plm_prod_1_out = pd.DataFrame(), pd.DataFrame()
        self.df_data_plm_prod_2_out, self.df_info_plm_prod_2_out = pd.DataFrame(), pd.DataFrame()

        self.df_data_main_prod_1_out, self.df_info_main_prod_1_out = pd.DataFrame(), pd.DataFrame()
        self.df_data_main_prod_2_out, self.df_info_main_prod_2_out = pd.DataFrame(), pd.DataFrame()

        self.df_data_pre_out, self.df_info_pre_out = pd.DataFrame(), pd.DataFrame()

        self.df_data_scr_plm_split, self.df_info_scr_plm_split = pd.DataFrame(), pd.DataFrame()

        self.df_data_stacked, self.df_info_stacked = pd.DataFrame(), pd.DataFrame()

        self.df_data_unstacked, self.df_info_unstacked = pd.DataFrame(), pd.DataFrame()

        self.sav_scr_plm_split_name = str()

        self.sav_stacked_name = str()
        self.sav_Unstacked_name = str()

        self.xlsx_name = str()

        self.zip_name = str()


    def init_data_exporter_variables(self, prj: dict, export_section: str, export_option: str = 'general'):

        self.prj = prj
        self.obj_section = prj['detail']['sections'][export_section]
        self.obj_prj_info = prj['detail']['prj_info']

        strProjectName = f"{self.prj['internal_id']}_{self.prj['name']}_{self.obj_section['name']}"
        self.strProjectName = strProjectName.replace(' - ', '_').replace(' ', '_').replace('-', '')

        self.strRidColName = self.prj['detail']['join_col']
        self.strOrderColName = ''
        self.is_split_callback = False
        self.lst_split_with_cols = list()

        self.is_has_data_plm = False if self.prj['type'] != 'HUT' else (True if self.prj['placement']['data'] else False)

        if self.prj['type'] == 'HUT':
            self.strOrderColName = self.prj['detail']['order_col']
            self.is_split_callback = self.prj['detail']['split_callback']['is_split_callback']
            self.lst_split_with_cols = self.prj['detail']['split_callback']['split_with_cols'].split(',')

        self.strFilter = self.obj_section['filter']

        self.strRotColName = self.obj_section['rotation']['name']

        # self.strProdColName = self.obj_section['product_qres']['1']['name']
        self.strProdColName = self.obj_section['product']['name']

        # dict_cats = self.obj_section['product_qres']['1']['cats']
        # self.dict_prod_cats = {
        #     int(dict_cats['1']['val']): dict_cats['1']['lbl'],
        #     int(dict_cats['2']['val']): dict_cats['2']['lbl']
        # }

        dict_cats = self.obj_section['product']['cats']
        self.dict_prod_cats = {int(k): v[0] for k, v in dict_cats.items()}

        # MỚI ADD
        self.lst_force_choice = self.obj_section['force_choice']['qres'][0].split('|')

        # GET COMBINE FORMAT
        print('GET COMBINE FORMAT')
        self.df_combined_oe = pd.DataFrame.from_dict(self.prj['detail']['oe_combine_cols'], orient='index',
                                                     columns=['combined_name', 'qre_name_1', 'qre_name_2'])

        # GET FORMAT
        print('GET SCR FORMAT')
        self.df_scr_format = pd.DataFrame.from_dict(self.prj['detail']['scr_cols'], orient='index',
                                                    columns=['output_name', 'input_name', 'qre_type'])

        if self.prj['type'] == 'HUT':
            print('GET PLM_SCR FORMAT')
            self.df_plm_scr_format = pd.DataFrame.from_dict(self.prj['detail']['plm_to_scr_cols'], orient='index',
                                                            columns=['output_name', 'input_name', 'qre_type'])

            print('GET PLM_MAIN FORMAT')
            self.df_plm_main_format = pd.DataFrame.from_dict(self.prj['detail']['plm_to_prod_cols'], orient='index',
                                                             columns=['output_name', 'input_name', 'qre_type'])

        print('GET MAIN FORMAT')
        self.df_main_format = pd.DataFrame.from_dict(self.prj['detail']['product_cols'], orient='index',
                                                     columns=['output_name', 'input_name_sp1', 'input_name_sp2', 'qre_type'])

        print('GET PRE FORMAT')
        self.df_pre_format = pd.DataFrame.from_dict(self.prj['detail']['fc_cols'], orient='index',
                                                    columns=['output_name', 'input_name', 'qre_type'])

        # GET ORIGINAL DATA
        print('GET ORIGINAL DATA')
        self.df_data_scr_org = pd.DataFrame.from_dict(self.prj['screener']['data'])
        self.df_info_scr_org = pd.DataFrame.from_dict(self.prj['screener']['qreInfo'])
        self.df_data_scr_org[self.strRidColName] = self.df_data_scr_org[self.strRidColName].astype(int)

        self.df_data_plm_org = pd.DataFrame()
        self.df_info_plm_org = pd.DataFrame()
        if self.prj['type'] == 'HUT':
            self.df_data_plm_org = pd.DataFrame.from_dict(self.prj['placement']['data'])
            self.df_info_plm_org = pd.DataFrame.from_dict(self.prj['placement']['qreInfo'])

            if not self.df_data_plm_org.empty:
                self.df_data_plm_org[self.strRidColName] = self.df_data_plm_org[self.strRidColName].astype(int)


        self.df_data_main_org = pd.DataFrame.from_dict(self.prj['main']['data'])
        self.df_info_main_org = pd.DataFrame.from_dict(self.prj['main']['qreInfo'])
        self.df_data_main_org[self.strRidColName] = self.df_data_main_org[self.strRidColName].astype(int)


        # DEFINE OUTPUT FORMATED DATAFRAME
        print('DEFINE OUTPUT FORMATED DATAFRAME')
        self.df_data_scr_out, self.df_info_scr_out = pd.DataFrame(), pd.DataFrame()
        self.df_data_plm_scr_out, self.df_info_plm_scr_out = pd.DataFrame(), pd.DataFrame()

        self.df_data_plm_prod_1_out, self.df_info_plm_prod_1_out = pd.DataFrame(), pd.DataFrame()
        self.df_data_plm_prod_2_out, self.df_info_plm_prod_2_out = pd.DataFrame(), pd.DataFrame()

        self.df_data_main_prod_1_out, self.df_info_main_prod_1_out = pd.DataFrame(), pd.DataFrame()
        self.df_data_main_prod_2_out, self.df_info_main_prod_2_out = pd.DataFrame(), pd.DataFrame()

        self.df_data_pre_out, self.df_info_pre_out = pd.DataFrame(), pd.DataFrame()

        # OUTPUT SCR_PLM SPLIT DATAFRAME
        self.df_data_scr_plm_split, self.df_info_scr_plm_split = pd.DataFrame(), pd.DataFrame()

        # OUTPUT STACKED DATAFRAME
        self.df_data_stacked, self.df_info_stacked = pd.DataFrame(), pd.DataFrame()

        # OUTPUT UNSTACKED DATAFRAME
        self.df_data_unstacked, self.df_info_unstacked = pd.DataFrame(), pd.DataFrame()

        # FILES NAME

        if export_option == 'screener_only':
            self.is_export_option_general = False
            self.is_export_option_scr_only = True
            self.is_export_option_codeframe = False

            self.sav_scr_plm_split_name = ""
            self.sav_stacked_name = f"{self.prj['internal_id']}_{self.prj['name']}_{export_option}.sav"
            self.sav_Unstacked_name = ""
            self.xlsx_name = f"{self.prj['internal_id']}_{self.prj['name']}_{export_option}.xlsx"
            self.zip_name = f"{self.prj['internal_id']}_{self.prj['name']}_{export_option}.zip"

        elif export_option == 'codeframe':
            self.is_export_option_general = False
            self.is_export_option_scr_only = False
            self.is_export_option_codeframe = True

            self.sav_scr_plm_split_name = ""
            self.sav_stacked_name = ""
            self.sav_Unstacked_name = ""
            self.xlsx_name = f'{self.strProjectName}_{export_option}.xlsx'
            self.zip_name = f"{self.strProjectName}_{export_option}.zip"

        else:
            self.is_export_option_general = True
            self.is_export_option_scr_only = False
            self.is_export_option_codeframe = False

            self.sav_scr_plm_split_name = f'{self.strProjectName}_Recruit_Placement.sav'
            self.sav_stacked_name = f'{self.strProjectName}_Stacked.sav'
            self.sav_Unstacked_name = f'{self.strProjectName}_Unstacked.sav'
            self.xlsx_name = f'{self.strProjectName}_ExcelData.xlsx'
            self.zip_name = f"{self.strProjectName}.zip"


    def init_topline_exporter_variables(self, prj: dict, df_data: pd.DataFrame, df_info: pd.DataFrame, export_section: str):

        ToplineExporterV2.__init__(self, prj, df_data, df_info, export_section)


    def export_data(self, is_export_raw=True, is_export_stacked=True, is_export_unstacked=True, is_fc_yn=True, is_export_sav=True):

        try:
            # RID duplicate Checking
            print('CHECK RID DUPLICATE')
            is_valid_rid, err_rid_lbl = self.rid_dup_validation()
            if not is_valid_rid:
                return is_valid_rid, err_rid_lbl

            # CHECK RID
            if not self.is_export_option_scr_only:
                print('CHECK RID')
                is_valid_rid, err_rid_lbl = self.rid_validation()
                if not is_valid_rid:
                    return is_valid_rid, err_rid_lbl

            self.generate_all_output_df()

            # # COPY df_info_scr_out FOR EXPORT EXCEL FILE
            # df_info_scr_out = self.df_info_scr_out.copy()

            # GENERATE OUTPUT SCR_PLM DATAFRAME FOR SPLIT HUT DATA
            if is_export_raw and self.prj['type'] == 'HUT' and self.is_split_callback and not self.is_export_option_scr_only:
                print('GENERATE OUTPUT SCR_PLM DATAFRAME FOR SPLIT HUT DATA')
                self.df_data_scr_plm_split, self.df_info_scr_plm_split, is_valid, errlbl = self.generate_scr_plm_split_df()

                if not is_valid:
                    return is_valid, errlbl

                self.df_info_scr_plm_split['val_lbl'] = [{int(cat): str(lbl) for cat, lbl in dict_val.items()} for dict_val in self.df_info_scr_plm_split['val_lbl']]

                print('EXPORT SCR_PLM SPLIT SAV FILE')
                self.export_sav_file(self.df_data_scr_plm_split, self.df_info_scr_plm_split, self.sav_scr_plm_split_name)

                # DROP UNSELECTED COLUMNS IN STACKED AND UNSTACKED FILES
                lst_col_to_drop = set(self.df_data_scr_out.columns)
                lst_col_to_drop = list(lst_col_to_drop.difference(
                    set([self.strRidColName, self.strRotColName] + self.lst_split_with_cols)))
                self.df_data_scr_out.drop(lst_col_to_drop, axis=1, inplace=True)

                self.df_info_scr_out['idx_var_name'] = self.df_info_scr_out['var_name']
                self.df_info_scr_out.set_index('idx_var_name', inplace=True)
                self.df_info_scr_out.drop(lst_col_to_drop, axis=0, inplace=True)
                self.df_info_scr_out.reset_index(drop=True, inplace=True)

                lst_col_to_drop = set(self.df_data_plm_scr_out.columns)
                lst_col_to_drop = list(lst_col_to_drop.difference(
                    set([self.strRidColName, self.strRotColName] + self.lst_split_with_cols)))
                self.df_data_plm_scr_out.drop(lst_col_to_drop, axis=1, inplace=True)

                self.df_info_plm_scr_out['idx_var_name'] = self.df_info_plm_scr_out['var_name']
                self.df_info_plm_scr_out.set_index('idx_var_name', inplace=True)
                self.df_info_plm_scr_out.drop(lst_col_to_drop, axis=0, inplace=True)
                self.df_info_plm_scr_out.reset_index(drop=True, inplace=True)


            # GENERATE OUTPUT STACKED DATAFRAME
            if is_export_stacked:
                print('GENERATE OUTPUT STACKED DATAFRAME')

                if self.is_export_option_scr_only:
                    self.df_data_stacked, self.df_info_stacked, is_valid, errlbl = self.df_data_scr_out.copy(), self.df_info_scr_out.copy(), True, None
                else:
                    self.df_data_stacked, self.df_info_stacked, is_valid, errlbl = self.generate_stacked_df()

                if not is_valid:
                    return is_valid, errlbl

                self.df_info_stacked['val_lbl'] = [{int(cat): str(lbl) for cat, lbl in dict_val.items()} for dict_val in self.df_info_stacked['val_lbl']]

                if is_export_sav:
                    print('EXPORT STACKED SAV FILE')
                    self.export_sav_file(self.df_data_stacked, self.df_info_stacked, self.sav_stacked_name)

                if self.is_export_option_codeframe:

                    cf_name = self.xlsx_name

                    dict_codeframes = dict()
                    for key, val in self.prj['codeframes'].items():

                        if '_FC_' in val['name']:
                            fc_col_name = self.obj_section['force_choice']['qres'][0]
                            lst_fix_col = [self.strRidColName, fc_col_name]
                        else:
                            lst_fix_col = [self.strRidColName, self.strProdColName]

                        lst_col = lst_fix_col + val['qres'].split(',')

                        df_cf = pd.melt(self.df_data_stacked.copy(), id_vars=lst_fix_col, value_vars=lst_col)

                        df_cf.dropna(subset=['value'], inplace=True)

                        df_cf.insert(3, 'Output_Name', [val['name']] * df_cf.shape[0])

                        df_cf['Codes'] = [np.nan] * df_cf.shape[0]

                        if '_FC_' in val['name']:
                            df_cf.drop_duplicates(subset=lst_fix_col + ['variable'], keep='first', inplace=True)

                        if 'Dau_Tien' in val['name'] or 'Tiep_Theo' in val['name']:

                            str_cf_name = val['name'].replace('_Dau_Tien', '').replace('_Tiep_Theo', '')

                            if str_cf_name in dict_codeframes.keys():
                                dict_codeframes[str_cf_name]['df'] = pd.concat([dict_codeframes[str_cf_name]['df'], df_cf], axis=0)
                            else:
                                dict_codeframes.update({str_cf_name: {'df': df_cf}})

                            dict_codeframes[str_cf_name]['df'].sort_values(by=lst_fix_col, inplace=True)

                        else:
                            dict_codeframes.update({val['name']: {'df': df_cf}})
                            dict_codeframes[val['name']]['df'].sort_values(by=lst_fix_col, inplace=True)


                    with pd.ExcelWriter(cf_name) as writer:
                        for key, val in dict_codeframes.items():

                            pd.DataFrame(columns=['Recode', 'Code', 'Label_VN', 'Label_EN']).to_excel(writer, sheet_name=f"{key}-Codeframe", index=False)  # encoding='utf-8-sig'
                            val['df'].to_excel(writer, sheet_name=f"{key}-Coding", index=False)  # encoding='utf-8-sig'

                            # writer.sheets[f"{key}-Codeframe"].set_tab_color('#FCD5B4')
                            # writer.sheets[f"{key}-Coding"].set_tab_color('#8DB4E2')

                    return self.export_zip_files([cf_name])


            # GENERATE OUTPUT UNSTACKED DATAFRAME
            if is_export_unstacked:
                print('GENERATE OUTPUT UNSTACKED DATAFRAME')
                self.df_data_unstacked, self.df_info_unstacked, is_valid, errlbl = self.generate_unstacked_df(is_export_sav=is_export_sav)

                if not is_valid:
                    return is_valid, errlbl

                self.convert_df_info_unstacked_val_lbl()

                # for idx in self.df_info_unstacked.index:
                #     val_lbl = self.df_info_unstacked.at[idx, 'val_lbl']
                #     new_val_lbl = dict()
                #     for cat, lbl in val_lbl.items():
                #         if '|' not in str(cat):
                #             new_val_lbl.update({int(cat): str(lbl)})
                #         else:
                #             for cat2, lbl2 in lbl.items():
                #                 if cat in new_val_lbl.keys():
                #                     new_val_lbl[cat].update({int(cat2): str(lbl2)})
                #                 else:
                #                     new_val_lbl.update({cat: {int(cat2): str(lbl2)}})
                #
                #     self.df_info_unstacked.at[idx, 'val_lbl'] = new_val_lbl

                # self.df_info_unstacked['val_lbl'] = [{int(cat): str(lbl) if '|' not in cat else {cat: str(lbl)} for cat, lbl in dict_val.items()} for dict_val in self.df_info_unstacked['val_lbl']]

                if is_export_sav:
                    print('EXPORT UNSTACKED SAV FILE')
                    self.export_sav_file(self.df_data_unstacked, self.df_info_unstacked, self.sav_Unstacked_name)

            # EXPORT EXCEL RAW DATA FILE
            if (is_export_stacked and is_export_unstacked) or self.is_export_option_scr_only:
                print('EXPORT EXCEL RAW DATA FILE')
                is_valid, errlbl = self.export_excel_file(self.df_info_scr_out)

                if not is_valid:
                    return is_valid, errlbl

            # Add-in yes_no columns for FC
            if is_export_unstacked and is_fc_yn:
                print('ADDIN YN FC COLUMNS TO UNSTACKED DATAFRAME')
                self.df_data_unstacked, self.df_info_unstacked = self.addin_yn_fc(self.df_data_unstacked, self.df_info_unstacked)

            if is_export_sav:

                lst_file_to_zip = [self.xlsx_name]

                if is_export_stacked:
                    lst_file_to_zip.append(self.sav_stacked_name)

                if is_export_unstacked:
                    lst_file_to_zip.append(self.sav_Unstacked_name)

                if is_export_raw and self.prj['type'] == 'HUT' and self.is_split_callback:
                    lst_file_to_zip.append(self.sav_scr_plm_split_name)

                return self.export_zip_files(lst_file_to_zip)

            return True, None

        except Exception:
            print(traceback.format_exc())
            return False, traceback.format_exc()


    def convert_df_info_unstacked_val_lbl(self):

        for idx in self.df_info_unstacked.index:
            val_lbl = self.df_info_unstacked.at[idx, 'val_lbl']
            new_val_lbl = dict()
            for cat, lbl in val_lbl.items():
                if '|' not in str(cat):
                    new_val_lbl.update({int(cat): str(lbl)})
                else:
                    for cat2, lbl2 in lbl.items():
                        if cat in new_val_lbl.keys():
                            new_val_lbl[cat].update({int(cat2): str(lbl2)})
                        else:
                            new_val_lbl.update({cat: {int(cat2): str(lbl2)}})

            self.df_info_unstacked.at[idx, 'val_lbl'] = new_val_lbl


    def rid_dup_validation(self):

        dfRidScr = self.df_data_scr_org.loc[:, [self.strRidColName]].copy()
        dfRidScrDup = pd.DataFrame(dfRidScr[dfRidScr.astype(int).duplicated()])
        if not dfRidScrDup.empty:
            return False, f'Duplicated id in screener: {dfRidScrDup.values}'

        if self.is_export_option_scr_only:
            return True, None

        if self.prj['type'] == 'HUT':

            if self.is_has_data_plm:
                dfRidPlm = self.df_data_plm_org.loc[self.df_data_plm_org[self.strOrderColName] == 1, [self.strRidColName]].copy()
                dfRidPlmDup = pd.DataFrame(dfRidPlm[dfRidPlm.astype(int).duplicated()])
                if not dfRidPlmDup.empty:
                    return False, f'Duplicated id in placement: {dfRidPlmDup.values}'

                dfRidPlm = self.df_data_plm_org.loc[self.df_data_plm_org[self.strOrderColName] == 2, [self.strRidColName]].copy()
                dfRidPlmDup = pd.DataFrame(dfRidPlm[dfRidPlm.astype(int).duplicated()])
                if not dfRidPlmDup.empty:
                    return False, f'Duplicated id in placement: {dfRidPlmDup.values}'

            dfRidMain = self.df_data_main_org.loc[self.df_data_main_org[self.strOrderColName] == 1, [self.strRidColName]].copy()
            dfRidMainDup = pd.DataFrame(dfRidMain[dfRidMain.astype(int).duplicated()])
            if not dfRidMainDup.empty:
                return False, f'Duplicated id in main: {dfRidMainDup.values}'

            dfRidMain = self.df_data_main_org.loc[self.df_data_main_org[self.strOrderColName] == 2, [self.strRidColName]].copy()
            dfRidMainDup = pd.DataFrame(dfRidMain[dfRidMain.astype(int).duplicated()])
            if not dfRidMainDup.empty:
                return False, f'Duplicated id in main: {dfRidMainDup.values}'

        else:
            dfRidMain = self.df_data_main_org.loc[:, [self.strRidColName]].copy()
            dfRidMainDup = pd.DataFrame(dfRidMain[dfRidMain.astype(int).duplicated()])
            if not dfRidMainDup.empty:
                return False, f'Duplicated id in main: {dfRidMainDup.values}'

        return True, None


    def rid_validation(self):
        try:
            rid_scr = set(self.df_data_scr_org.loc[:, self.strRidColName].copy())

            rid_plm_1, rid_plm_2, rid_main_1, rid_main_2 = set(), set(), set(), set()
            if self.prj['type'] == 'HUT':
                if self.is_has_data_plm:
                    rid_plm_1 = set(self.df_data_plm_org.loc[self.df_data_plm_org[self.strOrderColName] == 1, self.strRidColName].copy())
                    rid_plm_2 = set(self.df_data_plm_org.loc[self.df_data_plm_org[self.strOrderColName] == 2, self.strRidColName].copy())

                rid_main_1 = set(self.df_data_main_org.loc[self.df_data_main_org[self.strOrderColName] == 1, self.strRidColName].copy())
                rid_main_2 = set(self.df_data_main_org.loc[self.df_data_main_org[self.strOrderColName] == 2, self.strRidColName].copy())

            else:
                rid_main_1 = set(self.df_data_main_org.loc[:, self.strRidColName].copy())

            if self.prj['type'] == 'HUT':

                if self.is_has_data_plm:

                    pair_checking = {
                        'SCREENER vs PLACEMENT 1': [rid_scr, rid_plm_1],
                        'PLACEMENT 1 vs PLACEMENT 2': [rid_plm_1, rid_plm_2],
                        'PLACEMENT 2 vs RECALL 1': [rid_plm_2, rid_main_1],
                        'RECALL 1 vs RECALL 2': [rid_main_1, rid_main_2],
                    }

                else:

                    pair_checking = {
                        'SCREENER vs RECALL 1': [rid_scr, rid_main_1],
                        'RECALL 1 vs RECALL 2': [rid_main_1, rid_main_2],
                    }

            else:

                pair_checking = {
                    'SCREENER vs MAIN': [rid_scr, rid_main_1],
                }

            strError = str()
            for key, pair in pair_checking.items():
                checked = self.sets_checking(pair[0], pair[1])
                if len(checked):
                    strError += f'\nCheck ID {key}: {checked}'

            if strError:
                return False, strError

            return True, None

        except Exception:

            print(traceback.format_exc())
            return False, traceback.format_exc()


    def combine_original_data_cols(self):

        mainCols = list(self.df_data_main_org.columns)

        for idx in self.df_combined_oe.index:

            str_combine_name = self.df_combined_oe.at[idx, 'combined_name']
            str_qre_name_1 = self.df_combined_oe.at[idx, 'qre_name_1']
            str_qre_name_2 = self.df_combined_oe.at[idx, 'qre_name_2']

            if str_qre_name_1 in mainCols or str_qre_name_2 in mainCols:
                df_toCombined = self.df_data_main_org
            else:
                df_toCombined = self.df_data_plm_org

            df_toCombined[str_qre_name_1].replace({np.nan: 'NULL'}, inplace=True)
            df_toCombined[str_qre_name_2].replace({np.nan: 'NULL'}, inplace=True)

            lst_CombinedOE = [(a if a != 'NULL' else (b if b != 'NULL' else np.nan)) for a, b in
                              zip(df_toCombined[str_qre_name_1], df_toCombined[str_qre_name_2])]

            df_CombinedOE = pd.DataFrame(lst_CombinedOE, columns=[str_combine_name])
            df_CombinedOE[self.strRidColName] = df_toCombined[self.strRidColName].values
            lstColToMerge = [self.strRidColName]

            if self.prj['type'] == 'HUT':
                df_CombinedOE[self.strOrderColName] = df_toCombined[self.strOrderColName].values
                lstColToMerge.append(self.strOrderColName)

            df_toCombined = pd.merge(df_toCombined, df_CombinedOE, how='inner', on=lstColToMerge)

            df_toCombined[str_qre_name_1].replace({'NULL': np.nan}, inplace=True)
            df_toCombined[str_qre_name_2].replace({'NULL': np.nan}, inplace=True)

            if str_qre_name_1 in mainCols or str_qre_name_2 in mainCols:
                self.df_data_main_org = df_toCombined

                df_new_row = self.df_info_main_org.loc[self.df_info_main_org['var_name'] == str_qre_name_1, :].copy()
                df_new_row.loc[:, ['var_name']] = [str_combine_name]

                self.df_info_main_org = pd.concat([self.df_info_main_org, df_new_row], axis=0)
            else:
                self.df_data_plm_org = df_toCombined

                df_new_row = self.df_info_plm_org.loc[self.df_info_plm_org['var_name'] == str_qre_name_1, :].copy()
                df_new_row.loc[:, ['var_name']] = [str_combine_name]

                self.df_info_plm_org = pd.concat([self.df_info_plm_org, df_new_row], axis=0)


    def generate_all_output_df(self):

        if self.is_export_option_scr_only:

            print('GENERATE SCR DATA OUTPUT FOR SCR ONLY')
            self.df_data_scr_out, self.df_info_scr_out = self.df_data_scr_org.copy(), self.df_info_scr_org.copy()

        else:
            # RUN OE COMBINE IN MAIN AND PLM ORIGINAL DATA
            print('RUN OE COMBINE IN MAIN AND PLM ORIGINAL DATA')
            self.combine_original_data_cols()

            # SCR DATA OUTPUT
            print('GENERATE SCR DATA OUTPUT')
            self.df_data_scr_out, self.df_info_scr_out = self.generate_output_df(str_df_type='SCR')

            # Update addin var to Scr (new - at 10/04/2023)
            dict_update_ata_scr_org = dict()
            for key, val in self.prj['detail']['addin_vars'].items():
                if val['name'] in self.df_data_scr_org.columns:
                    dict_update_ata_scr_org[key] = val

            addin_vars = AddinVariables(prj_addVars=dict_update_ata_scr_org, df_data=self.df_data_scr_out, df_info=self.df_info_scr_out, lstProductCode=[])
            addin_vars.addin_vars()
            # End Update addin var to Scr (new - at 10/04/2023)

            if self.prj['type'] == 'HUT':

                if self.is_has_data_plm:
                    # PLM_SCR DATA OUTPUT
                    print('GENERATE PLM_SCR DATA OUTPUT')
                    self.df_data_plm_scr_out, self.df_info_plm_scr_out = self.generate_output_df(str_df_type='PLM_SCR')

                    # PLM_PROD_1 DATA OUTPUT
                    print('GENERATE PLM_PROD_1 DATA OUTPUT')
                    self.df_data_plm_prod_1_out, self.df_info_plm_prod_1_out = self.generate_output_df(str_df_type='PLM_PROD_1')

                    # PLM_PROD_2 DATA OUTPUT
                    print('GENERATE PLM_PROD_2 DATA OUTPUT')
                    self.df_data_plm_prod_2_out, self.df_info_plm_prod_2_out = self.generate_output_df(str_df_type='PLM_PROD_2')

            # MAIN_PROD_1 DATA OUTPUT
            print('GENERATE MAIN_PROD_1 DATA OUTPUT')
            self.df_data_main_prod_1_out, self.df_info_main_prod_1_out = self.generate_output_df(str_df_type='MAIN_PROD_1')

            # MAIN_PROD_2 DATA OUTPUT
            print('GENERATE MAIN_PROD_2 DATA OUTPUT')
            self.df_data_main_prod_2_out, self.df_info_main_prod_2_out = self.generate_output_df(str_df_type='MAIN_PROD_2')

            # PRE DATA OUTPUT
            print('GENERATE PRE DATA OUTPUT')
            self.df_data_pre_out, self.df_info_pre_out = self.generate_output_df(str_df_type='PRE')

            # Update 29/06/2023
            # Add OE data & qre info to each dataframe in placement, main, force choice
            if self.prj['has_oe']:

                df_oe_codelist = pd.read_csv(io.BytesIO(self.prj['oe_info']['codelist']))
                # dict_qre_OE_info_org = eval(self.prj['oe_info']['codelist'])

                df_oe_info_coding = pd.read_csv(io.BytesIO(self.prj['oe_info']['coding']))
                df_oe_info_coding.rename(columns={'ID': self.strRidColName}, inplace=True)

                # for k, v in dict_qre_OE_info_org.items():
                for idx in df_oe_codelist.index:
                    oe_name, oe_num_col = df_oe_codelist.at[idx, 'COL_NAME'].rsplit('|', 1)
                    oe_num_col = oe_num_col.split('-')
                    oe_num_col = range(int(oe_num_col[0]), int(oe_num_col[1]) + 1)
                    str_part = str(df_oe_codelist.at[idx, 'SEC']).upper()

                    lst_qre_info_to_add = list()
                    lst_qre_colname_to_add = list()
                    for i in oe_num_col:
                        lst_qre_info = df_oe_codelist.loc[idx, ['LABEL', 'TYPE', 'CODELIST']].values.tolist()
                        lst_qre_info[-1] = eval(lst_qre_info[-1])
                        lst_qre_info_to_add.append([f'{oe_name}_{i}'] + lst_qre_info)
                        lst_qre_colname_to_add.extend([f'{oe_name}_{i}'])


                    if str_part in ['PRODUCT', 'FORCE_CHOICE']:

                        self.df_info_main_prod_1_out = pd.concat([self.df_info_main_prod_1_out, pd.DataFrame(
                            columns=['var_name', 'var_lbl', 'var_type', 'val_lbl'],
                            data=lst_qre_info_to_add
                        )], axis=0, ignore_index=True)

                        lst_qre_colname_to_add = [self.strRidColName, self.strProdColName] + lst_qre_colname_to_add
                        self.df_data_main_prod_1_out = self.df_data_main_prod_1_out.merge(df_oe_info_coding[lst_qre_colname_to_add], how='left', on=[self.strRidColName, self.strProdColName])

                        self.df_info_main_prod_2_out = pd.concat([self.df_info_main_prod_2_out, pd.DataFrame(
                            columns=['var_name', 'var_lbl', 'var_type', 'val_lbl'],
                            data=lst_qre_info_to_add
                        )], axis=0, ignore_index=True)

                        self.df_data_main_prod_2_out = self.df_data_main_prod_2_out.merge(df_oe_info_coding[lst_qre_colname_to_add], how='left', on=[self.strRidColName, self.strProdColName])

                    if str_part in ['NORMAL_OE']:
                        self.df_info_pre_out = pd.concat([self.df_info_pre_out, pd.DataFrame(
                            columns=['var_name', 'var_lbl', 'var_type', 'val_lbl'],
                            data=lst_qre_info_to_add
                        )], axis=0, ignore_index=True)

                        lst_qre_colname_to_add = [self.strRidColName] + lst_qre_colname_to_add
                        df_pre_to_merge = df_oe_info_coding[lst_qre_colname_to_add].copy().drop_duplicates(subset=[self.strRidColName], keep='first')
                        self.df_data_pre_out = self.df_data_pre_out.merge(df_pre_to_merge, how='left', on=[self.strRidColName])




    def generate_output_df(self, str_df_type: str):

        rid_name = self.strRidColName

        order_name = ''
        if self.prj['type'] == 'HUT':
            order_name = self.strOrderColName

        df_format = pd.DataFrame()
        df_data_org = pd.DataFrame()
        df_info_org = pd.DataFrame()
        df_data_out = pd.DataFrame()
        df_info_out = pd.DataFrame()

        if str_df_type in ['SCR']:
            df_format = self.df_scr_format.copy()
            df_data_org = self.df_data_scr_org.copy()
            df_info_org = self.df_info_scr_org.copy()

            # Some question in main but not run by product so need to merge data main and screener
            if self.prj['type'] == 'HUT':
                df_data_org = pd.merge(df_data_org, self.df_data_main_org.loc[self.df_data_main_org[self.strOrderColName] == 1, :], how='left', on=self.strRidColName)
                df_info_org = pd.concat([df_info_org, self.df_info_main_org], axis=0)
            else:
                df_data_org = pd.merge(df_data_org, self.df_data_main_org, how='left', on=self.strRidColName)
                df_info_org = pd.concat([df_info_org, self.df_info_main_org], axis=0)


        elif str_df_type in ['PLM_SCR']:
            df_format = self.df_plm_scr_format.copy()
            df_data_org = self.df_data_plm_org.loc[self.df_data_plm_org[order_name] == 1, :].copy()
            df_info_org = self.df_info_plm_org.copy()

        elif str_df_type in ['PLM_PROD_1', 'PLM_PROD_2']:
            df_format = self.df_plm_main_format.copy()

            filVal = 1 if str_df_type in ['PLM_PROD_1'] else 2
            df_data_org = self.df_data_plm_org.loc[self.df_data_plm_org[order_name] == filVal, :].copy()

            df_info_org = self.df_info_plm_org.copy()

        elif str_df_type in ['MAIN_PROD_1', 'MAIN_PROD_2']:

            str_input_name = 'input_name_sp1' if str_df_type in ['MAIN_PROD_1'] else 'input_name_sp2'

            df_format = self.df_main_format.loc[:, ['output_name', str_input_name]].copy()
            df_format.rename(columns={str_input_name: 'input_name'}, inplace=True)

            if self.prj['type'] == 'HUT':
                filVal = 1 if str_df_type in ['MAIN_PROD_1'] else 2
                df_data_org = self.df_data_main_org.loc[self.df_data_main_org[order_name] == filVal, :].copy()
            else:
                df_data_org = self.df_data_main_org.copy()

            df_info_org = self.df_info_main_org.copy()

        elif str_df_type in ['PRE']:
            df_format = self.df_pre_format.copy()

            if self.prj['type'] == 'HUT':
                df_data_org = self.df_data_main_org.loc[self.df_data_main_org[order_name] == 2, :].copy()
            else:
                df_data_org = self.df_data_main_org.copy()

            df_info_org = self.df_info_main_org.copy()

        # RESET INDEX OF ORIGINAL DATAFRAME
        df_data_org.reset_index(drop=True, inplace=True)

        # Create RID column
        df_data_out[rid_name] = df_data_org[rid_name]
        df_data_out.reset_index(drop=True, inplace=True)
        df_info_out = df_info_org.loc[df_info_org['var_name'] == rid_name, :]

        # Create Product code column
        if str_df_type in ['MAIN_PROD_1', 'MAIN_PROD_2']:

            # if str_df_type in ['MAIN_PROD_1']:
            #     prod_code_info = self.obj_section['product_qres']['1']
            # else:
            #     prod_code_info = self.obj_section['product_qres']['2']

            # cats_prod = {
            #     int(prod_code_info['cats']['1']['val']): prod_code_info['cats']['1']['lbl'],
            #     int(prod_code_info['cats']['2']['val']): prod_code_info['cats']['2']['lbl']
            # }

            # lstQreProd = prod_code_info['qres'][0].split('|')
            # df_data_prod = df_data_org.loc[:, lstQreProd]
            # df_data_prod.replace({np.nan: 0}, inplace=True)
            # df_data_prod[prod_code_info['name']] = df_data_prod[lstQreProd].sum(axis=1)

            # df_data_out = pd.concat([df_data_out, df_data_prod.loc[:, [prod_code_info['name']]]], axis=1)

            # df_info_prod_code = pd.DataFrame([[prod_code_info['name'], prod_code_info['lbl'], 'SA', cats_prod]],
            #                                  columns=['var_name', 'var_lbl', 'var_type', 'val_lbl'])
            # df_info_out = pd.concat([df_info_out, df_info_prod_code], axis=0)

            # ----------------------------------------------------------------------------------------------------------

            cats_prod = self.dict_prod_cats
            if str_df_type in ['MAIN_PROD_1']:
                lstQreProd = self.obj_section['product']['qres'][0].split('|')
            else:
                lstQreProd = self.obj_section['product']['qres'][1].split('|')
            df_data_prod = df_data_org.loc[:, lstQreProd]
            df_data_prod.replace({np.nan: 0}, inplace=True)
            df_data_prod[self.strProdColName] = df_data_prod[lstQreProd].sum(axis=1)

            df_data_out = pd.concat([df_data_out, df_data_prod.loc[:, [self.strProdColName]]], axis=1)

            df_info_prod_code = pd.DataFrame([[self.strProdColName, self.obj_section['product']['lbl'], 'SA', cats_prod]],
                                             columns=['var_name', 'var_lbl', 'var_type', 'val_lbl'])

            df_info_out = pd.concat([df_info_out, df_info_prod_code], axis=0)


        # Recode FC columns
        if str_df_type in ['PRE'] and len(self.lst_force_choice[0]):
            # prod_qres = self.obj_section['product_qres']
            #
            # cats_prod = {
            #     int(prod_qres['1']['cats']['1']['val']): prod_qres['1']['cats']['1']['lbl'],
            #     int(prod_qres['1']['cats']['2']['val']): prod_qres['1']['cats']['2']['lbl']
            # }
            #
            # df_prod_1st, df_prod_2nd = pd.DataFrame(), pd.DataFrame()
            # for key, item in prod_qres.items():
            #
            #     lstQreProd = [rid_name]
            #     lstQreProd.extend(item['qres'][0].split('|'))
            #
            #     if self.prj['type'] == 'HUT':
            #         df_data_prod = self.df_data_main_org.loc[self.df_data_main_org[order_name] == int(key), lstQreProd].copy()
            #     else:
            #         df_data_prod = self.df_data_main_org.loc[:, lstQreProd].copy()
            #
            #     df_data_prod.replace({np.nan: 0}, inplace=True)
            #
            #     if key == '1':
            #         df_prod_1st[rid_name] = df_data_prod[rid_name]
            #         df_prod_1st['prod_1st'] = df_data_prod[lstQreProd[1:]].sum(axis=1)
            #     else:
            #         df_prod_2nd[rid_name] = df_data_prod[rid_name]
            #         df_prod_2nd['prod_2nd'] = df_data_prod[lstQreProd[1:]].sum(axis=1)

            prod_qres = {
                1: self.obj_section['product']['qres'][0].split('|'),
                2: self.obj_section['product']['qres'][1].split('|')
            }
            cats_prod = self.dict_prod_cats

            df_prod_1st, df_prod_2nd = pd.DataFrame(), pd.DataFrame()
            for key, val in prod_qres.items():

                lstQreProd = [rid_name]
                lstQreProd.extend(val)

                if self.prj['type'] == 'HUT':
                    df_data_prod = self.df_data_main_org.loc[self.df_data_main_org[order_name] == key, lstQreProd].copy()
                else:
                    df_data_prod = self.df_data_main_org.loc[:, lstQreProd].copy()

                df_data_prod.replace({np.nan: 0}, inplace=True)

                if key == 1:
                    df_prod_1st[rid_name] = df_data_prod[rid_name]
                    df_prod_1st['prod_1st'] = df_data_prod[lstQreProd[1:]].sum(axis=1)
                else:
                    df_prod_2nd[rid_name] = df_data_prod[rid_name]
                    df_prod_2nd['prod_2nd'] = df_data_prod[lstQreProd[1:]].sum(axis=1)

            df_data_prods = pd.merge(df_prod_1st, df_prod_2nd, how="left", on=[rid_name])

            lst_fc_qre = [rid_name]
            lst_fc_qre.extend(self.lst_force_choice)

            df_data_prods = pd.merge(df_data_prods, df_data_org.loc[:, lst_fc_qre], how="left", on=[rid_name])

            for qre_fc in lst_fc_qre[1:]:

                df_data_prods[qre_fc] = [np.nan if np.isnan(sel) else (_1st if sel == 1 else (_2nd if sel == 2 else (3 if sel == 3 else np.nan))) for _1st, _2nd, sel in zip(df_data_prods['prod_1st'], df_data_prods['prod_2nd'], df_data_prods[qre_fc])]

                if 3 in df_data_prods[qre_fc].values.tolist():
                    fc_cats = cats_prod | {3: 'Thích 2 sản phẩm như nhau'}
                else:
                    fc_cats = cats_prod

                df_info_org.loc[df_info_org['var_name'] == qre_fc, ['val_lbl']] = [fc_cats]

            df_data_org.drop(lst_fc_qre[1:], axis=1, inplace=True)
            df_data_org = pd.merge(df_data_org, df_data_prods, how="left", on=[rid_name])


        # GENERATE DATA
        for idx in df_format.index:

            df_data_concat = df_data_org.loc[:, [rid_name, df_format.at[idx, 'input_name']]]
            df_data_concat.rename(columns={df_format.at[idx, 'input_name']: df_format.at[idx, 'output_name']}, inplace=True)

            df_data_out = pd.merge(df_data_out, df_data_concat, how='left', on=[rid_name])

            df_info_concat = df_info_org.loc[df_info_org['var_name'] == df_format.at[idx, 'input_name'], :].copy()
            df_info_concat.loc[:, ['var_name']] = [df_format.at[idx, 'output_name']]
            df_info_out = pd.concat([df_info_out, df_info_concat], axis=0)


        # Create Rotation code column
        if str_df_type in ['SCR']:
            dict_rot_info = self.obj_section['rotation']

            if dict_rot_info['name'] not in list(df_data_out.columns):
                dict_recode_rot = dict()
                for k, v in dict_rot_info['cats'].items():
                    for i in v[1].split('|'):
                        dict_recode_rot[int(i)] = int(k)

                df_data_out = pd.concat([df_data_out, df_data_org[dict_rot_info['qres'][0]]], axis=1)
                df_data_out.rename(columns={dict_rot_info['qres'][0]: dict_rot_info['name']}, inplace=True)

                df_data_out[dict_rot_info['name']].replace(dict_recode_rot, inplace=True)

                cats_rot = {int(k): v[0] for k, v in dict_rot_info['cats'].items()}
                df_info_rot = pd.DataFrame([[dict_rot_info['name'], dict_rot_info['lbl'], 'SA', cats_rot]],
                                           columns=['var_name', 'var_lbl', 'var_type', 'val_lbl'])
                df_info_out = pd.concat([df_info_out, df_info_rot], axis=0)


        df_data_out.reset_index(drop=True, inplace=True)
        df_info_out.reset_index(drop=True, inplace=True)

        return df_data_out, df_info_out


    def generate_scr_plm_split_df(self):

        try:
            df_data_scr_plm_split = self.df_data_scr_out.copy()
            df_data_scr_plm_split = pd.merge(df_data_scr_plm_split, self.df_data_plm_scr_out, how='left', on=self.strRidColName)
            df_data_scr_plm_split[self.strRidColName] = df_data_scr_plm_split[self.strRidColName].apply(pd.to_numeric)
            df_data_scr_plm_split.sort_values(by=[self.strRidColName], inplace=True)

            lst_df_info = [
                self.df_info_scr_out,
                self.df_info_plm_scr_out
            ]

            df_info_scr_plm_split = pd.concat(lst_df_info, axis=0, ignore_index=True)
            df_info_scr_plm_split.drop_duplicates(subset=['var_name'], keep='first', inplace=True)
            df_info_scr_plm_split.reset_index(drop=True, inplace=True)

            df_data_scr_plm_split = self.get_filter_data(df_data_scr_plm_split, self.strFilter)

            return df_data_scr_plm_split, df_info_scr_plm_split, True, None

        except Exception:
            print(traceback.format_exc())
            return None, None, False, traceback.format_exc()


    def generate_stacked_df(self):
        
        try:

            df_prod_1st, df_prod_2nd = self.df_data_scr_out.copy(), self.df_data_scr_out.copy()

            # ------------------------------------------------------------------------------
            df_prod_1st['idx_rid'] = df_prod_1st[self.strRidColName]
            df_prod_1st.set_index('idx_rid', inplace=True)

            df_prod_2nd['idx_rid'] = df_prod_2nd[self.strRidColName]
            df_prod_2nd.set_index('idx_rid', inplace=True)
            # ------------------------------------------------------------------------------

            if self.prj['type'] == 'HUT':
                if self.is_has_data_plm:
                    # df_prod_1st = pd.merge(df_prod_1st, self.df_data_plm_scr_out, how='left', on=self.strRidColName)
                    # df_prod_2nd = pd.merge(df_prod_2nd, self.df_data_plm_scr_out, how='left', on=self.strRidColName)
                    #
                    # df_prod_1st = pd.merge(df_prod_1st, self.df_data_plm_prod_1_out, how='left', on=self.strRidColName)
                    # df_prod_2nd = pd.merge(df_prod_2nd, self.df_data_plm_prod_2_out, how='left', on=self.strRidColName)

                    # ------------------------------------------------------------------------------
                    df_data_plm_scr_out = self.df_data_plm_scr_out.copy()
                    df_data_plm_scr_out.set_index(self.strRidColName, inplace=True)

                    df_data_plm_prod_1_out = self.df_data_plm_prod_1_out.copy()
                    df_data_plm_prod_1_out.set_index(self.strRidColName, inplace=True)

                    df_data_plm_prod_2_out = self.df_data_plm_prod_2_out.copy()
                    df_data_plm_prod_2_out.set_index(self.strRidColName, inplace=True)

                    df_prod_1st = pd.concat([df_prod_1st, df_data_plm_scr_out, df_data_plm_prod_1_out], axis=1)
                    df_prod_2nd = pd.concat([df_prod_2nd, df_data_plm_scr_out, df_data_plm_prod_2_out], axis=1)

                    del df_data_plm_scr_out
                    del df_data_plm_prod_1_out
                    del df_data_plm_prod_2_out
                    # ------------------------------------------------------------------------------

            # df_prod_1st = pd.merge(df_prod_1st, self.df_data_main_prod_1_out, how='left', on=self.strRidColName)
            # df_prod_2nd = pd.merge(df_prod_2nd, self.df_data_main_prod_2_out, how='left', on=self.strRidColName)
            #
            # df_prod_1st = pd.merge(df_prod_1st, self.df_data_pre_out, how='left', on=self.strRidColName)
            # df_prod_2nd = pd.merge(df_prod_2nd, self.df_data_pre_out, how='left', on=self.strRidColName)

            # ------------------------------------------------------------------------------
            df_data_main_prod_1_out = self.df_data_main_prod_1_out.copy()
            df_data_main_prod_1_out.set_index(self.strRidColName, inplace=True)

            df_data_main_prod_2_out = self.df_data_main_prod_2_out.copy()
            df_data_main_prod_2_out.set_index(self.strRidColName, inplace=True)

            df_data_pre_out = self.df_data_pre_out.copy()
            df_data_pre_out.set_index(self.strRidColName, inplace=True)

            df_prod_1st = pd.concat([df_prod_1st, df_data_main_prod_1_out, df_data_pre_out], axis=1)
            df_prod_2nd = pd.concat([df_prod_2nd, df_data_main_prod_2_out, df_data_pre_out], axis=1)

            del df_data_main_prod_1_out
            del df_data_main_prod_2_out
            del df_data_pre_out

            df_prod_1st.reset_index(inplace=True, drop=True)
            df_prod_2nd.reset_index(inplace=True, drop=True)
            # ------------------------------------------------------------------------------

            df_data_stacked = pd.concat([df_prod_1st, df_prod_2nd], axis=0, ignore_index=True)
            df_data_stacked[self.strRidColName] = df_data_stacked[self.strRidColName].apply(pd.to_numeric)
            df_data_stacked.sort_values(by=[self.strRidColName, self.strProdColName], inplace=True)
            df_data_stacked.reset_index(inplace=True, drop=True)


            if self.prj['type'] == 'HUT':
                lst_df_info = [
                    self.df_info_scr_out,
                    self.df_info_plm_scr_out,
                    self.df_info_plm_prod_1_out,
                    self.df_info_main_prod_1_out,
                    self.df_info_pre_out
                ]
                # No need to concat "self.df_data_plm_prod_2_out" cuz it is the same as "self.df_data_plm_prod_1_out"
                # No need to concat "self.df_info_main_prod_2_out" cuz it is the same as "self.df_info_main_prod_1_out"
            else:
                lst_df_info = [
                    self.df_info_scr_out,
                    self.df_info_main_prod_1_out,
                    self.df_info_pre_out
                ]

            df_info_stacked = pd.concat(lst_df_info, axis=0, ignore_index=True)
            df_info_stacked.drop_duplicates(subset=['var_name'], keep='first', inplace=True)
            df_info_stacked.reset_index(drop=True, inplace=True)

            # Remove net code in val_lbl
            df_info_stacked = self.remove_net_code_in_df_info(df_info_stacked)

            # Export CSV files for checking
            # df_prod_1st.to_csv('df_prod_1st.csv', encoding='utf-8-sig')
            # df_prod_2nd.to_csv('df_prod_2nd.csv', encoding='utf-8-sig')
            # df_data_stacked.to_csv('df_data_stacked.csv', encoding='utf-8-sig')
            # df_info_stacked.to_csv('df_info_stacked.csv', encoding='utf-8-sig')

            df_data_stacked = self.get_filter_data(df_data_stacked, self.strFilter)

            lst_col_stacked = list(df_data_stacked.columns)
            lst_col_stacked.remove(self.strProdColName)
            index_of_rot = lst_col_stacked.index(self.strRotColName)
            lst_col_stacked.insert(index_of_rot + 1, self.strProdColName)

            df_data_stacked = df_data_stacked.reindex(columns=lst_col_stacked)

            df_info_stacked['var_name_idx'] = df_info_stacked['var_name']
            df_info_stacked.set_index('var_name_idx', inplace=True)
            df_info_stacked = df_info_stacked.reindex(lst_col_stacked)
            df_info_stacked.reset_index(inplace=True, drop=True)

            return df_data_stacked, df_info_stacked, True, None

        except Exception:
            print(traceback.format_exc())
            return None, None, False, traceback.format_exc()


    def generate_unstacked_df(self, is_export_sav: bool = False):

        try:

            df_data_unstacked, df_info_unstacked = self.df_data_scr_out.copy(), self.df_info_scr_out.copy()

            if self.prj['type'] == 'HUT':
                if self.is_has_data_plm:
                    df_data_unstacked = pd.merge(df_data_unstacked, self.df_data_plm_scr_out, how='left', on=self.strRidColName)
                    df_info_unstacked = pd.concat([df_info_unstacked, self.df_info_plm_scr_out], axis=0, ignore_index=True)

                    # PLM_1 & PLM_2-----------------------------------------------------------------------------------------
                    df_data_unstacked, df_info_unstacked = self.generate_unstacked_df_by_prods(df_data_unstacked, df_info_unstacked, 'PLM')
                    # END PLM_1 & PLM_2-------------------------------------------------------------------------------------

            # MAIN_1 & MAIN_2-------------------------------------------------------------------------------------------
            df_data_unstacked, df_info_unstacked = self.generate_unstacked_df_by_prods(df_data_unstacked, df_info_unstacked, 'MAIN')
            # END MAIN_1 & MAIN_2---------------------------------------------------------------------------------------

            # PRE-------------------------------------------------------------------------------------------------------
            df_data_unstacked = pd.merge(df_data_unstacked, self.df_data_pre_out, how='left', on=self.strRidColName)
            df_info_unstacked = pd.concat([df_info_unstacked, self.df_info_pre_out], axis=0, ignore_index=True)
            # END PRE---------------------------------------------------------------------------------------------------

            df_data_unstacked[self.strRidColName] = df_data_unstacked[self.strRidColName].apply(pd.to_numeric)
            df_data_unstacked.sort_values(by=[self.strRidColName], inplace=True)

            # Remove net code in val_lbl
            if is_export_sav:
                df_info_unstacked = self.remove_net_code_in_df_info(df_info_unstacked)

            df_info_unstacked.drop_duplicates(subset=['var_name'], keep='first', inplace=True)
            df_info_unstacked.reset_index(drop=True, inplace=True)

            # Important
            df_data_unstacked = df_data_unstacked.reindex(columns=df_info_unstacked['var_name'].values.tolist())

            # df_data_unstacked.to_csv('df_data_unstacked.csv', encoding='utf-8-sig')
            # df_info_unstacked.to_csv('df_info_unstacked.csv', encoding='utf-8-sig')

            df_data_unstacked = self.get_filter_data(df_data_unstacked, self.strFilter)

            return df_data_unstacked, df_info_unstacked, True, None

        except Exception:
            print(traceback.format_exc())
            return None, None, False, traceback.format_exc()


    def generate_unstacked_df_by_prods(self, df_data_unstacked, df_info_unstacked, section):

        if section == 'MAIN':
            df_data_prod_1 = self.df_data_main_prod_1_out.copy()
            df_data_prod_2 = self.df_data_main_prod_2_out.copy()
            df_info_prod_1 = self.df_info_main_prod_1_out.copy()
        else:
            df_data_prod_1 = pd.merge(self.df_data_plm_prod_1_out, self.df_data_main_prod_1_out.loc[:, [self.strRidColName, self.strProdColName]], how='left', on=self.strRidColName)
            df_data_prod_2 = pd.merge(self.df_data_plm_prod_2_out, self.df_data_main_prod_2_out.loc[:, [self.strRidColName, self.strProdColName]], how='left', on=self.strRidColName)
            df_info_prod_1 = self.df_info_plm_prod_1_out.copy()

        df_data_merge = pd.concat([df_data_prod_1, df_data_prod_2], axis=0, ignore_index=True)

        del df_data_prod_1
        del df_data_prod_2

        df_data_A = df_data_merge.loc[df_data_merge[self.strProdColName] == 1, :].copy()
        df_data_B = df_data_merge.loc[df_data_merge[self.strProdColName] == 2, :].copy()

        df_data_A.sort_values(by=[self.strRidColName], inplace=True)
        df_data_B.sort_values(by=[self.strRidColName], inplace=True)

        df_data_A.reset_index(drop=True, inplace=True)
        df_data_B.reset_index(drop=True, inplace=True)

        dict_replace_col_name_A = dict()
        dict_replace_col_name_B = dict()
        lst_drop_ma_col_name = list()
        for idx_info in df_info_prod_1.index:
            col_name = df_info_prod_1.at[idx_info, 'var_name']
            col_lbl = df_info_prod_1.at[idx_info, 'var_lbl']
            col_type = df_info_prod_1.at[idx_info, 'var_type']
            col_val_lbl = df_info_prod_1.at[idx_info, 'val_lbl']

            if col_name not in [self.strRidColName, self.strProdColName] + lst_drop_ma_col_name:
                if 'MA' in col_type:

                    ma_name_1st = col_name.rsplit('_', 1)[0]
                    lst_ma_col = df_info_prod_1.loc[df_info_prod_1['var_name'].str.contains(f'{ma_name_1st}_[0-9]+'), 'var_name'].values.tolist()
                    lst_drop_ma_col_name.extend(lst_ma_col)  # next col of same MA ques will not run again

                    lst_info = list()
                    for k_sp, v_sp in self.dict_prod_cats.items():
                        for ma_col in lst_ma_col:
                            ma_name, ma_cat = ma_col.rsplit('_', 1)
                            v_sp_new = v_sp.replace('|', '_')

                            lst_info.append([f'{ma_name}_{v_sp_new}_{ma_cat}', f'{col_lbl}_{v_sp_new}', col_type, col_val_lbl])

                            if k_sp == 1:
                                dict_replace_col_name_A.update({ma_col: f'{ma_name}_{v_sp_new}_{ma_cat}'})
                            else:
                                dict_replace_col_name_B.update({ma_col: f'{ma_name}_{v_sp_new}_{ma_cat}'})

                else:

                    lst_info = list()
                    for k_sp, v_sp in self.dict_prod_cats.items():
                        v_sp_new = v_sp.replace('|', '_')

                        lst_info.append([f'{col_name}_{v_sp_new}', f'{col_lbl}_{v_sp_new}', col_type, col_val_lbl])

                        if k_sp == 1:
                            dict_replace_col_name_A.update({col_name: f'{col_name}_{v_sp_new}'})
                        else:
                            dict_replace_col_name_B.update({col_name: f'{col_name}_{v_sp_new}'})

                df_info_unstacked = pd.concat([df_info_unstacked, pd.DataFrame(
                    columns=['var_name', 'var_lbl', 'var_type', 'val_lbl'],
                    data=lst_info)], axis=0, ignore_index=True)

        df_info_unstacked.drop_duplicates(subset=['var_name'], keep='first', inplace=True)
        df_info_unstacked.reset_index(drop=True, inplace=True)

        df_data_A.rename(columns=dict_replace_col_name_A, inplace=True)
        df_data_B.rename(columns=dict_replace_col_name_B, inplace=True)

        df_data_A.drop(columns=[self.strProdColName], inplace=True)
        df_data_B.drop(columns=[self.strProdColName], inplace=True)

        df_data_unstacked = pd.merge(df_data_unstacked, df_data_A, how='left', on=self.strRidColName)
        df_data_unstacked = pd.merge(df_data_unstacked, df_data_B, how='left', on=self.strRidColName)

        return df_data_unstacked, df_info_unstacked


    @staticmethod
    def sets_checking(set1, set2):
        return set1.difference(set2).union(set2.difference(set1))


    @staticmethod
    def get_filter_data(df_to_fil: pd.DataFrame, strFilter: str):
        return df_to_fil.query(strFilter)


    def addin_yn_fc(self, df_data_unstacked: pd.DataFrame, df_info_unstacked: pd.DataFrame):

        if len(self.lst_force_choice[0]):
            for fc_qre in self.lst_force_choice:
                for cat, lbl in self.dict_prod_cats.items():
                    new_lbl = lbl.replace('|', '_')

                    new_fc_name = f"{fc_qre}_{new_lbl}"

                    df_data_unstacked[new_fc_name] = [np.nan if pd.isna(a) else (1 if int(a) == int(cat) else 0) for a in df_data_unstacked[fc_qre]]

                    df_info_new_fc = pd.DataFrame([[new_fc_name, new_fc_name, 'SA', {0: 'No', 1: 'Yes'}]],
                                                  columns=['var_name', 'var_lbl', 'var_type', 'val_lbl'])

                    df_info_unstacked = pd.concat([df_info_unstacked, df_info_new_fc], axis=0)

            df_info_unstacked.reset_index(drop=True, inplace=True)

        return df_data_unstacked, df_info_unstacked


    @staticmethod
    def export_sav_file(df_data: pd.DataFrame, df_info: pd.DataFrame, file_name: str):

        # Columns = ['var_name', 'var_lbl', 'var_type', 'val_lbl']

        df_info['idx_by_var'] = df_info['var_name']
        df_info.set_index('idx_by_var', inplace=True)
        df_info['measure'] = ['nominal'] * df_info.shape[0]

        lst_col_lbl = df_info.loc[:, 'var_lbl'].to_list()
        dict_val_lbl = df_info.loc[:, 'val_lbl'].to_dict()
        dict_var_measure = df_info.loc[:, 'measure'].to_dict()

        df_data.replace({None: np.nan}, inplace=True)

        pyreadstat.write_sav(df_data, file_name, column_labels=lst_col_lbl, variable_value_labels=dict_val_lbl, variable_measure=dict_var_measure)
        print('EXPORT "', file_name, '" COMPLETED')


    def export_excel_file(self, df_info_scr_out):

        try:
            df_data_scr_plm_split = self.df_data_scr_plm_split.copy()
            df_data_stacked = self.df_data_stacked.copy()
            df_data_unstacked = self.df_data_unstacked.copy()

            xlsx_name = self.xlsx_name

            df_recode = df_info_scr_out.loc[df_info_scr_out['val_lbl'] != {}, ['var_name', 'val_lbl']].copy()

            if self.strProdColName in df_data_stacked.columns:
                df_recode = pd.concat([df_recode, pd.DataFrame([[self.strProdColName, self.dict_prod_cats]], columns=['var_name', 'val_lbl'])], axis=0)

            df_recode.set_index('var_name', inplace=True)
            df_recode['val_lbl'] = [{int(cat): lbl for cat, lbl in dict_val.items()} for dict_val in df_recode['val_lbl']]
            dict_recode = df_recode.loc[:, 'val_lbl'].to_dict()

            df_data_scr_plm_split.replace(dict_recode, inplace=True)
            df_data_stacked.replace(dict_recode, inplace=True)
            df_data_unstacked.replace(dict_recode, inplace=True)

            df_prj_information = pd.DataFrame([
                ['1. Mục tiêu nghiên cứu'],
                [self.obj_prj_info['1']['val']],
                ['2. Đối tượng và phương pháp nghiên cứu'],
                [self.obj_prj_info['2_1']['val']],
                [self.obj_prj_info['2_2']['val']],
                [self.obj_prj_info['2_3']['val']],
                [self.obj_prj_info['2_4']['val']],
                ['3. Thông tin nghiên cứu (OL, JAR, Like/Dislikes, v.v…)'],
                [self.obj_prj_info['3']['val']],
                ['4. Action Standard'],
                [self.obj_prj_info['4']['val']],
                ['5. Thời gian thực hiện'],
                [self.obj_prj_info['5']['val']],
                ['6. Chú thích mã sản phẩm'],
                [self.obj_prj_info['6_1']['val']],
                [self.obj_prj_info['6_2']['val']],
                [self.obj_prj_info['6_3']['val']],
                [self.obj_prj_info['6_4']['val']]
            ], columns=['THÔNG TIN DỰ ÁN'])

            with pd.ExcelWriter(xlsx_name) as writer:
                df_prj_information.to_excel(writer, sheet_name='Thông tin dự án', index=False)  # encoding='utf-8-sig'

                if self.is_export_option_scr_only:
                    df_data_stacked.to_excel(writer, sheet_name='Screener Only', index=False)
                else:
                    if not df_data_scr_plm_split.empty:
                        df_data_scr_plm_split.to_excel(writer, sheet_name='Screener_Placement', index=False)  # encoding='utf-8-sig'

                    df_data_unstacked.to_excel(writer, sheet_name='Unstacked', index=False)  # encoding='utf-8-sig'
                    df_data_stacked.to_excel(writer, sheet_name='Stacked', index=False)  # encoding='utf-8-sig'

            # Format sheet 'Thông tin dự án'
            self.format_sheet_prj_infor(xlsx_name)

            print('EXPORT "', xlsx_name, '" COMPLETED')

            return True, None

        except Exception:
            print(traceback.format_exc())
            return False, traceback.format_exc()


    @staticmethod
    def format_sheet_prj_infor(xlsx_name):
        wb = load_workbook(filename=xlsx_name)
        ws = wb['Thông tin dự án']

        ws.column_dimensions['A'].width = 150
        thin = Side(border_style="thin", color="000000")

        for i in range(1, 20):

            cell = ws[f'A{i}']
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            if i in [1, 2, 4, 9, 11, 13, 15]:
                cell.font = Font(color='FFFFFF', bold=True)
                cell.fill = PatternFill('solid', fgColor='5B9BD5')
            else:
                cell.fill = PatternFill('solid', fgColor='DDEBF7')
                cell.alignment = Alignment(horizontal='right', vertical='justify')

        wb.save(xlsx_name)
        wb.close()


    def export_zip_files(self, lstFile: list):
        try:

            with zipfile.ZipFile(self.zip_name, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
                for f in lstFile:
                    zf.write(f)
                    os.remove(f)

            return True, None

        except Exception:
            print(traceback.format_exc())
            return False, traceback.format_exc()


    def remove_net_code_in_df_info(self, df_info: pd.DataFrame) -> pd.DataFrame:

        # Remove net_code to export sav---------------------------------------------------------------------------------
        df_info_without_net = df_info.copy()

        for idx in df_info_without_net.index:
            if df_info_without_net.at[idx, 'var_type'] in ['FT', 'NUM']:
                continue

            val_lbl = df_info_without_net.at[idx, 'val_lbl']

            for k in val_lbl.keys():
                if '|' in str(k):
                    df_info_without_net.at[idx, 'val_lbl'] = self.unnetted_qre_val(val_lbl)
                    break
        # END Remove net_code to export sav-----------------------------------------------------------------------------

        return df_info_without_net

    @staticmethod
    def unnetted_qre_val(dict_netted) -> dict:
        dict_unnetted = dict()

        for key, val in dict_netted.items():

            if isinstance(val, dict):
                dict_unnetted.update(val)
            else:
                dict_unnetted.update({key: val})

        return dict_unnetted