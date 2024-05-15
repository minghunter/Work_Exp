import pandas as pd
import openpyxl
import pyreadstat
import io
import numpy as np
import zipfile
import re
import os


class APDataConverter:

    def __init__(self, file):
        self.str_file_name = file.filename
        self.df_data_input, self.df_qres_info_input = self.read_file(file)
        self.zip_name = file.filename.replace('.xlsx', '.zip')


    @staticmethod
    def read_file(file) -> (pd.DataFrame, pd.DataFrame):

        xlsx = io.BytesIO(file.file.read())
        wb = openpyxl.load_workbook(xlsx, data_only=True)

        wsData = wb['Data']
        wsQres = wb['Question']

        mergedCells = list()
        for group in wsData.merged_cells.ranges:
            mergedCells.append(group)

        for group in mergedCells:
            wsData.unmerge_cells(str(group))

        wsData.delete_rows(1, 4)
        wsData.delete_rows(2, 1)

        for icol in range(1, wsData.max_column + 1):
            if wsData.cell(row=1, column=icol).value is None:
                wsData.cell(row=1,
                            column=icol).value = f'{wsData.cell(row=1, column=icol - 1).value}_{wsData.cell(row=2, column=icol).value}'

        wsData.delete_rows(2, 1)

        data = wsData.values
        columns = next(data)[0:]
        df_data = pd.DataFrame(data, columns=columns)

        lstDrop = [
            'Approve',
            'Reject',
            'Re - do request', 'Re-do request',
            'Reason to reject',
            'Memo',
            'No.',
            'Date',
            'Country',
            'Channel',
            'Chain / Type',
            'Distributor',
            'Method',
            'Panel FB',
            'Panel Email',
            'Panel Phone',
            'Panel Age',
            'Panel Gender',
            'Panel Area',
            'Panel Income',
            'Login ID',
            'User name',
            'Store ID',
            'Store Code',
            'Store name',
            'Store level',
            'District',
            'Ward',
            'Store address',
            'Area group',
            'Store ranking',
            'Region 2',
            'Nhóm cửa hàng',
            'Nhà phân phối',
            'Manager',
            'Telephone number',
            'Contact person',
            'Email',
            'Others 1',
            'Others 2',
            'Others 3',
            'Others 4',
            'Check in',
            'Store Latitude',
            'Store Longitude',
            'User Latitude',
            'User Longitude',
            'Check out',
            'Distance',
            'Task duration',

            'Panel ID',
            'InterviewerID',
            'InterviewerName',
            'RespondentName',

            'Edited',
            'Edited by',
            'Edited ratio',

            'Verify Status',
        ]

        for col in df_data.columns:
            if col in lstDrop or '_Images' in col:
                df_data.drop(col, inplace=True, axis=1)

        data = wsQres.values
        columns = next(data)[0:]
        df_qres = pd.DataFrame(data, columns=columns)

        wb.close()

        return df_data, df_qres


    def check_duplicate_variables(self):

        dup_vars = self.df_qres_info_input.duplicated(subset=['Name of items'])

        lst_dup_vars = list()
        if dup_vars.any():
            lst_dup_vars = self.df_qres_info_input.loc[dup_vars, 'Name of items'].values.tolist()

        return lst_dup_vars


    def convert_to_sav(self, is_md: bool):

        if is_md:
            df_data, df_qres_info = self.convert_df_md()
        else:
            df_data, df_qres_info = self.convert_df_mc()

        self.generate_sav_sps(df_data, df_qres_info, is_md)


    def convert_df_md(self) -> (pd.DataFrame, pd.DataFrame):

        df_data, df_qres_info = self.df_data_input, self.df_qres_info_input

        dictQres = dict()
        for idx in df_qres_info.index:

            strMatrix = '' if df_qres_info.loc[idx, 'Question(Matrix)'] is None else f"{df_qres_info.loc[idx, 'Question(Matrix)']}_"
            strNormal = df_qres_info.loc[idx, 'Question(Normal)'] if strMatrix == '' else f"{strMatrix}{df_qres_info.loc[idx, 'Question(Normal)']}"
            strQreName = str(df_qres_info.loc[idx, 'Name of items'])
            strQreName = strQreName.replace('Rank_', 'Rank') if 'Rank_' in strQreName else strQreName

            dictQres[strQreName] = {
                'type': df_qres_info.loc[idx, 'Question type'],
                'label': f'{strNormal}',
                'isMatrix': True if strMatrix != '' else False,
                'cats': {}
            }

            lstHeaderCol = list(df_qres_info.columns)
            lstHeaderCol.remove('Name of items')
            lstHeaderCol.remove('Question type')
            lstHeaderCol.remove('Question(Matrix)')
            lstHeaderCol.remove('Question(Normal)')

            for col in lstHeaderCol:
                if df_qres_info.loc[idx, col] is not None and len(str(df_qres_info.loc[idx, col])) > 0:
                    dictQres[strQreName]['cats'].update({str(col): self.cleanhtml(str(df_qres_info.loc[idx, col]))})

        lstMatrixHeader = list()
        for k in dictQres.keys():
            if dictQres[k]['isMatrix'] and dictQres[k]['type'] == 'MA' and len(dictQres[k]['cats'].keys()):
                lstMatrixHeader.append(k)

        if len(lstMatrixHeader):
            for i in lstMatrixHeader:
                for code in dictQres[i]['cats'].keys():
                    lstLblMatrixMA = dictQres[f'{i}_{code}']['label'].split('_')
                    dictQres[f'{i}_{code}']['cats'].update({'1': self.cleanhtml(lstLblMatrixMA[1])})
                    dictQres[f'{i}_{code}']['label'] = f"{dictQres[i]['label']}_{lstLblMatrixMA[1]}"

        df_data_output, df_qres_info_output = df_data, pd.DataFrame(data=[['ID', 'ID', 'FT', {}]],
                                                                    columns=['var_name', 'var_lbl', 'var_type', 'val_lbl'])

        for qre, qre_info in dictQres.items():

            if qre in df_data_output.columns:

                arr_row = [qre, self.cleanhtml(qre_info['label']), f"{qre_info['type']}_mtr" if qre_info['isMatrix'] else qre_info['type'], qre_info['cats']]

                df_qres_info_output = pd.concat([df_qres_info_output, pd.DataFrame(data=[arr_row], columns=['var_name', 'var_lbl', 'var_type', 'val_lbl'])])


        df_data_output.replace({None: np.nan}, inplace=True)
        df_qres_info_output.reset_index(drop=True, inplace=True)

        # df_data_output.to_csv('df_data_output.csv', encoding='utf-8-sig')
        # df_qres_info_output.to_csv('df_qres_info_output.csv', encoding='utf-8-sig')

        return df_data_output, df_qres_info_output


    def convert_df_mc(self) -> (pd.DataFrame, pd.DataFrame):  # convert data with MA questions format by columns instead of code

        df_data, df_qres_info = self.df_data_input, self.df_qres_info_input
        df_data.replace({None: np.nan}, inplace=True)

        lstFullCodelist = list(df_qres_info.columns)
        lstFullCodelist.remove('Name of items')
        lstFullCodelist.remove('Question type')
        lstFullCodelist.remove('Question(Matrix)')
        lstFullCodelist.remove('Question(Normal)')

        dictQres = dict()
        for idx in df_qres_info.index:

            strQreName = str(df_qres_info.loc[idx, 'Name of items'])
            strQreName = strQreName.replace('Rank_', 'Rank') if 'Rank_' in strQreName else strQreName
            strQreType = df_qres_info.loc[idx, 'Question type']
            isMatrix = False if df_qres_info.loc[idx, 'Question(Matrix)'] is None else True
            strMatrix = '' if df_qres_info.loc[idx, 'Question(Matrix)'] is None else self.cleanhtml(f"{df_qres_info.loc[idx, 'Question(Matrix)']}")
            strNormal = '' if df_qres_info.loc[idx, 'Question(Normal)'] is None else self.cleanhtml(f"{df_qres_info.loc[idx, 'Question(Normal)']}")

            if strQreName not in dictQres.keys():

                if strQreType == 'MA':

                    if isMatrix:

                        ser_codelist = df_qres_info.loc[idx, lstFullCodelist]
                        ser_codelist.dropna(inplace=True)
                        dict_codelist = ser_codelist.to_dict()

                        if not ser_codelist.empty:
                            dictQres[strQreName] = {
                                'type': strQreType,
                                'label': f'{strMatrix}_{strNormal}' if isMatrix else strNormal,
                                'isMatrix': isMatrix,
                                'MA_Matrix_Header': strQreName,
                                'MA_cols': [f'{strQreName}_{k}' for k in dict_codelist.keys()],
                                'cats': {str(k): self.cleanhtml(v) for k, v in dict_codelist.items()},
                            }
                    else:

                        maName, maCode = strQreName.rsplit('_', 1)
                        maLbl = self.cleanhtml(df_qres_info.at[idx, 1])

                        if maName not in dictQres.keys():

                            dictQres[maName] = {
                                'type': strQreType,
                                'label': strNormal,
                                'isMatrix': isMatrix,
                                'MA_cols': [strQreName],
                                'cats': {str(maCode): maLbl}
                            }

                        else:

                            dict_qre = dictQres[maName]
                            dict_qre['MA_cols'].append(strQreName)
                            dict_qre['cats'].update({str(maCode): maLbl})


                else:  # ['SA', 'RANKING', 'FT']

                    dictQres[strQreName] = {
                        'type': strQreType,
                        'label': str(),
                        'isMatrix': isMatrix,
                        'cats': dict(),
                    }

                    dict_qre = dictQres[strQreName]
                    dict_qre['label'] = f'{strMatrix}_{strNormal}' if isMatrix else strNormal

                    if strQreType in ['SA', 'RANKING']:
                        ser_codelist = df_qres_info.loc[idx, lstFullCodelist]
                        ser_codelist.dropna(inplace=True)
                        dict_qre['cats'] = {str(k): self.cleanhtml(v) for k, v in ser_codelist.to_dict().items()}

        df_data_output = df_data.loc[:, ['ID']].copy()

        df_qres_info_output = pd.DataFrame(data=[['ID', 'ID', 'FT', {}]], columns=['var_name', 'var_lbl', 'var_type', 'val_lbl'])

        # df_data_output.index = df_data.index

        for qre, qre_info in dictQres.items():

            if qre_info['type'] == 'MA':

                dfMA = df_data.loc[:, qre_info['MA_cols']]

                for col_name in qre_info['MA_cols']:
                    maName, maCode = col_name.rsplit('_', 1)
                    dfMA[col_name].replace({1: int(maCode)}, inplace=True)

                dfMA['combined'] = [[e for e in row if e == e] for row in dfMA[qre_info['MA_cols']].values.tolist()]
                dfMA = pd.DataFrame(dfMA['combined'].tolist(), index=dfMA.index)

                for i, col_name in enumerate(qre_info['MA_cols']):

                    if i in list(dfMA.columns):
                        dfColMA = dfMA[i].to_frame()
                        dfColMA.rename(columns={i: col_name}, inplace=True)
                    else:
                        dfColMA = pd.DataFrame([np.nan] * dfMA.shape[0], columns=[col_name])

                    df_data_output = pd.concat([df_data_output, dfColMA], axis=1)

                    dfInfoRow = pd.DataFrame([[col_name, qre_info['label'], 'MA_mtr' if qre_info['isMatrix'] else 'MA', qre_info['cats']]], columns=['var_name', 'var_lbl', 'var_type', 'val_lbl'])

                    df_qres_info_output = pd.concat([df_qres_info_output, dfInfoRow], axis=0)

            else:
                if qre in df_data.columns:
                    df_data_output = pd.concat([df_data_output, df_data[qre]], axis=1)

                    dfInfoRow = pd.DataFrame([[qre, qre_info['label'], f"{qre_info['type']}_mtr" if qre_info['isMatrix'] else qre_info['type'], qre_info['cats']]], columns=['var_name', 'var_lbl', 'var_type', 'val_lbl'])

                    df_qres_info_output = pd.concat([df_qres_info_output, dfInfoRow], axis=0)

        # dfQreInfo.set_index('var_name', inplace=True)
        df_qres_info_output.reset_index(drop=True, inplace=True)

        # dfDataOutput.to_csv('dfDataOutput.csv', encoding='utf-8-sig')
        # dfQreInfo.to_csv('dfQreInfo.csv', encoding='utf-8-sig')

        return df_data_output, df_qres_info_output


    @staticmethod
    def cleanhtml(raw_html) -> str:

        if isinstance(raw_html, str):
            CLEANR = re.compile('{.*?}|<.*?>|&([a-z0-9]+|#[0-9]{1,6}|#x[0-9a-f]{1,6});|\n|\xa0')
            cleantext = re.sub(CLEANR, '', raw_html)
            return cleantext

        return raw_html


    def generate_sav_sps(self, df_data: pd.DataFrame, df_qres_info: pd.DataFrame, is_md: bool, is_export_xlsx: bool = False):

        str_sav_name = self.str_file_name.replace('.xlsx', '.sav')
        str_sps_name = self.str_file_name.replace('.xlsx', '.sps')

        dict_val_lbl = {a: {int(k): str(v) for k, v in b.items()} for a, b in zip(df_qres_info['var_name'], df_qres_info['val_lbl'])}

        dict_measure = {a: 'nominal' for a in df_qres_info['var_name']}

        pyreadstat.write_sav(df_data, str_sav_name,
                             column_labels=df_qres_info['var_lbl'].values.tolist(),
                             variable_value_labels=dict_val_lbl, variable_measure=dict_measure)

        self.generate_sps(df_qres_info, is_md, str_sps_name)

        if is_export_xlsx:

            df_data_xlsx = df_data.copy()
            df_recode = df_qres_info.loc[df_qres_info['val_lbl'] != {}, ['var_name', 'val_lbl']].copy()

            df_recode.set_index('var_name', inplace=True)
            df_recode['val_lbl'] = [{int(cat): lbl for cat, lbl in dict_val.items()} for dict_val in
                                    df_recode['val_lbl']]
            dict_recode = df_recode.loc[:, 'val_lbl'].to_dict()

            df_data_xlsx.replace(dict_recode, inplace=True)

            xlsx_name = self.str_file_name.replace('.xlsx', '_Rawdata.xlsx')
            topline_name = self.str_file_name.replace('.xlsx', '_Topline.xlsx')

            with pd.ExcelWriter(xlsx_name, engine="openpyxl") as writer:
                df_data_xlsx.to_excel(writer, sheet_name='Rawdata', index=False, encoding='utf-8-sig')

            self.zipfiles(self.zip_name, [str_sav_name, str_sps_name, xlsx_name, topline_name])
        else:
            self.zipfiles(self.zip_name, [str_sav_name, str_sps_name])


    @staticmethod
    def generate_sps(df_qres_info: pd.DataFrame, is_md: bool, sps_name: str):

        if is_md:
            temp = """
            *{0}.
            MRSETS
            /MDGROUP NAME=${1}
                LABEL='{2}'
                CATEGORYLABELS=COUNTEDVALUES 
                VARIABLES={3}
                VALUE=1
            /DISPLAY NAME=[${4}].
            """
        else:
            temp = """
            *{0}.
            MRSETS
            /MCGROUP NAME=${1}
                LABEL='{2}' 
                VARIABLES={3}
            /DISPLAY NAME=[${4}].
            """

        df_qres_ma = df_qres_info.loc[(df_qres_info['var_type'].str.contains('MA')), :].copy()

        lst_ignore_col = list()

        dict_ma_cols = dict()
        for idx in df_qres_ma.index:

            ma_name = df_qres_ma.at[idx, 'var_name'].rsplit('_', 1)[0]

            if ma_name in lst_ignore_col:
                dict_ma_cols[ma_name]['vars'].append(df_qres_ma.at[idx, 'var_name'])
            else:
                lst_ignore_col.append(ma_name)

                dict_ma_cols[ma_name] = {
                    'name': ma_name,
                    'lbl': df_qres_ma.at[idx, 'var_lbl'],
                    'vars': [df_qres_ma.at[idx, 'var_name']],
                }

        # df_qres_ma = df_qres_info.loc[(df_qres_info['var_name'].str.contains(f'{qre}_[1-9]+')), :].copy()

        str_MRSet = '.'
        for key, val in dict_ma_cols.items():
            str_MRSet += temp.format(key, val['name'], val['lbl'], ' '.join(val['vars']), val['name'])

        with open(f'{sps_name}', 'w', encoding='utf-8-sig') as text_file:
            text_file.write(str_MRSet)


    @staticmethod
    def zipfiles(zip_name: str, lst_file_name: list):
        with zipfile.ZipFile(zip_name, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
            for f_name in lst_file_name:
                zf.write(f_name)
                os.remove(f_name)












